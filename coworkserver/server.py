"""
Claude Cowork Web v5
- JWT 인증 → userid → 폴더 격리
- 백그라운드 작업 지속 (브라우저 닫아도 계속 실행)
- 재접속 시 진행/완료 자동 복원
- 작업 시작/완료 시점 MongoDB 기록
- 스트리밍 응답
"""
import os, re, json, shutil, asyncio, sys, mimetypes, uuid, time, hashlib, base64, traceback
from pathlib import Path
from datetime import datetime, timezone, timedelta
from functools import partial
from typing import List, Optional
from collections import defaultdict

# .env 파일 로드 (python-dotenv)
_env_loaded = False
try:
    from dotenv import load_dotenv
    # 여러 경로에서 .env 탐색
    _env_candidates = []
    try:
        _env_candidates.append(Path(__file__).resolve().parent / ".env")
    except:
        pass
    _env_candidates.append(Path(".env").resolve())
    _env_candidates.append(Path(os.getcwd()) / ".env")

    for _ep in _env_candidates:
        try:
            if _ep.exists():
                load_dotenv(str(_ep), override=True)
                _env_loaded = True
                print(f"[ENV] .env 로드: {_ep}")
                break
        except:
            continue
    if not _env_loaded:
        print("[WARNING] .env 파일을 찾을 수 없습니다. os.environ만 사용합니다.")
except ImportError:
    print("[WARNING] python-dotenv 미설치. 'pip install python-dotenv' 실행 권장.")

import aiofiles, aiofiles.os, httpx
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, HTTPException, Form, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, Response
import anthropic

# ============================================================
# .env 기반 설정 (회사별 커스터마이징)
# ============================================================

# --- 브랜딩 ---
APP_TITLE = os.environ.get("APP_TITLE", "K-Portal Cowork")              # 브라우저 탭 제목
APP_BRAND = os.environ.get("APP_BRAND", "K-Portal")                     # 로고 메인 텍스트
APP_BRAND_SUB = os.environ.get("APP_BRAND_SUB", "Cowork")               # 로고 서브 텍스트
APP_BRAND_ICON = os.environ.get("APP_BRAND_ICON", "K")                  # 로고 아이콘 글자
APP_BRAND_COMPANY = os.environ.get("APP_BRAND_COMPANY", "KMSLAB")       # 로고 하단 회사명
APP_WELCOME_TITLE = os.environ.get("APP_WELCOME_TITLE", "K-Portal <span>AI Cowork</span>")  # 환영 화면 제목 (HTML 허용)
APP_ASSISTANT_NAME = os.environ.get("APP_ASSISTANT_NAME", "AI Cowork 어시스턴트")  # AI 자기소개
APP_VERSION = os.environ.get("APP_VERSION", "")                         # 정적 파일 버전 (빈값이면 mtime 자동)

# --- JWT 인증 ---
JWT_SECRET = os.environ.get("JWT_SECRET", "kmslabbox2022")              # JWT 서명 키
PORTAL_URL = os.environ.get("PORTAL_URL", "https://one.kmslab.com")     # 포털 로그인 URL (세션 만료 시 리다이렉트)

# --- MongoDB ---
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://imadmin:kmslabkm@im.k-portal.co.kr:16270/?authSource=admin")
MONGO_DB_NAME = os.environ.get("MONGO_DB_NAME", "cowork")              # 메인 DB명
MONGO_ORG_DB_NAME = os.environ.get("MONGO_ORG_DB_NAME", "im_org_info") # 조직도 DB명
MONGO_POOL_SIZE = int(os.environ.get("MONGO_POOL_SIZE", "50"))
MONGO_TIMEOUT_MS = int(os.environ.get("MONGO_TIMEOUT_MS", "5000"))

# --- 워크스페이스 ---
WORKSPACE_ROOT = os.environ.get("COWORK_WORKSPACE", "D:/cowork-workspace")

# --- 서버 ---
import socket
SERVER_ID = os.environ.get("SERVER_ID", f"{socket.gethostname()}-{os.getpid()}")

# --- 관리자 ---
# .env에서 관리자 ID를 직접 지정 (쉼표 구분). 조직도 DB의 role과 병행 사용.
ADMIN_USERS = set(u.strip() for u in os.environ.get("ADMIN_USERS", "").split(",") if u.strip())

# --- AI 모델 ---
MODEL_OPUS = os.environ.get("MODEL_OPUS", "claude-opus-4-6")
MODEL_SONNET = os.environ.get("MODEL_SONNET", "claude-sonnet-4-6")
MODEL = MODEL_SONNET
MAX_TOKENS_OPUS = int(os.environ.get("MAX_TOKENS_OPUS", "64000"))
MAX_TOKENS_SONNET = int(os.environ.get("MAX_TOKENS_SONNET", "32000"))
MAX_TOKENS = MAX_TOKENS_SONNET

# --- 토큰 가격 (USD per 1M tokens) ---
TOKEN_PRICING = {
    os.environ.get("MODEL_OPUS", "claude-opus-4-6"): {
        "input": float(os.environ.get("OPUS_INPUT_PRICE_PER_MTK", "15.0")),
        "output": float(os.environ.get("OPUS_OUTPUT_PRICE_PER_MTK", "75.0")),
    },
    os.environ.get("MODEL_SONNET", "claude-sonnet-4-6"): {
        "input": float(os.environ.get("SONNET_INPUT_PRICE_PER_MTK", "3.0")),
        "output": float(os.environ.get("SONNET_OUTPUT_PRICE_PER_MTK", "15.0")),
    },
}

def estimate_cost(model: str, input_tokens: int, output_tokens: int) -> float:
    pricing = TOKEN_PRICING.get(model, {"input": 3.0, "output": 15.0})
    return (input_tokens * pricing["input"] + output_tokens * pricing["output"]) / 1_000_000

async def record_token_usage(
    username: str, task_id: str, session_id: str,
    service_type: str, model: str,
    usage, step: int, key_index: int
):
    if not MONGO_OK or token_usage_col is None or usage is None:
        return
    try:
        input_tokens = getattr(usage, 'input_tokens', 0) or 0
        output_tokens = getattr(usage, 'output_tokens', 0) or 0
        cache_creation = getattr(usage, 'cache_creation_input_tokens', 0) or 0
        cache_read = getattr(usage, 'cache_read_input_tokens', 0) or 0
        cost = estimate_cost(model, input_tokens, output_tokens)
        await token_usage_col.insert_one({
            "username": username, "task_id": task_id, "session_id": session_id,
            "service_type": service_type, "model": model,
            "input_tokens": input_tokens, "output_tokens": output_tokens,
            "cache_creation_input_tokens": cache_creation,
            "cache_read_input_tokens": cache_read,
            "total_tokens": input_tokens + output_tokens,
            "step": step, "key_index": key_index,
            "cost_estimate": round(cost, 6),
            "created_at": datetime.now(timezone.utc),
        })
    except Exception as e:
        print(f"[TOKEN USAGE] record error: {e}")

# --- 웹 검색 ---
PERPLEXITY_API_KEY = os.environ.get("PERPLEXITY_API_KEY", "")

# --- CORS ---
CORS_ORIGINS = [o.strip() for o in os.environ.get("CORS_ORIGINS", "https://one.kmslab.com,http://one.kmslab.com").split(",") if o.strip()]

# --- 스냅샷 ---
MAX_SNAPSHOTS = int(os.environ.get("MAX_SNAPSHOTS", "5"))

# --- 지원 언어 ---
SUPPORTED_LANGS = set(os.environ.get("SUPPORTED_LANGS", "ko,en,ja,zh").split(","))

# 호환용 변수 (기존 코드에서 참조)
KPORTAL_JWT_SECRET = JWT_SECRET
KPORTAL_URL = PORTAL_URL

# ============================================================
# JWT 인증
# ============================================================
import jwt as pyjwt

def decode_kportal_jwt(token: str, verify_exp: bool = True) -> Optional[dict]:
    """K-Portal JWT 복호화. verify_exp=True면 만료 검증 수행"""
    secrets_to_try = [
        KPORTAL_JWT_SECRET.encode("utf-8"),                    # raw bytes
        base64.b64encode(KPORTAL_JWT_SECRET.encode("utf-8")),  # base64 encoded
    ]
    for secret in secrets_to_try:
        try:
            payload = pyjwt.decode(token, secret, algorithms=["HS256"],
                                   options={"verify_exp": verify_exp})
            return payload
        except pyjwt.ExpiredSignatureError:
            if not verify_exp:
                try:
                    return pyjwt.decode(token, secret, algorithms=["HS256"],
                                        options={"verify_exp": False})
                except:
                    continue
            # 만료된 토큰 → None 반환 (거부)
            return None
        except:
            continue
    return None

def userid_from_jwt(token: str) -> Optional[str]:
    """JWT에서 userid 추출 (만료 토큰 거부)"""
    p = decode_kportal_jwt(token, verify_exp=True)
    if not p: return None
    uid = p.get("userid", "")
    if "@" in uid: return uid.split("@")[0]
    return uid if uid else None

def email_from_jwt(token: str) -> str:
    """JWT에서 이메일(userid 원본) 추출 (로그용, 만료 무시)"""
    p = decode_kportal_jwt(token, verify_exp=False)
    if not p: return ""
    return p.get("userid", "") or p.get("email", "") or ""

def userid_from_jwt_unsafe(token: str) -> Optional[str]:
    """JWT에서 userid 추출 (만료 무시 - 디버깅 전용)"""
    p = decode_kportal_jwt(token, verify_exp=False)
    if not p: return None
    uid = p.get("userid", "")
    if "@" in uid: return uid.split("@")[0]
    return uid if uid else None

def is_jwt_expired(token: str) -> bool:
    """JWT가 만료되었는지 확인"""
    try:
        secrets_to_try = [
            KPORTAL_JWT_SECRET.encode("utf-8"),
            base64.b64encode(KPORTAL_JWT_SECRET.encode("utf-8")),
        ]
        for secret in secrets_to_try:
            try:
                pyjwt.decode(token, secret, algorithms=["HS256"], options={"verify_exp": True})
                return False  # 유효
            except pyjwt.ExpiredSignatureError:
                return True   # 만료
            except:
                continue
    except:
        pass
    return True  # 판별 불가 → 만료 처리

# ============================================================
# MongoDB
# ============================================================
try:
    import motor.motor_asyncio
    mongo_client = motor.motor_asyncio.AsyncIOMotorClient(MONGO_URI, serverSelectionTimeoutMS=MONGO_TIMEOUT_MS, maxPoolSize=MONGO_POOL_SIZE)
    mongo_db = mongo_client[MONGO_DB_NAME]
    chat_collection = mongo_db["chat_logs"]
    task_collection = mongo_db["tasks"]
    task_log_collection = mongo_db["task_logs"]
    user_settings_collection = mongo_db["user_settings"]
    shared_folders_collection = mongo_db["shared_folders"]
    temp_links_collection = mongo_db["temp_links"]
    skills_collection = mongo_db["skills"]
    projects_collection = mongo_db["projects"]
    # 분산 상태 관리 컬렉션
    active_sessions_col = mongo_db["active_sessions"]
    active_tasks_col = mongo_db["active_tasks"]
    api_key_state_col = mongo_db["api_key_state"]
    scheduler_lock_col = mongo_db["scheduler_locks"]
    token_usage_col = mongo_db["token_usage"]
    # 조직도 DB
    org_db = mongo_client[MONGO_ORG_DB_NAME]
    org_user_collection = org_db["user_info"]
    MONGO_OK = True
except Exception:
    MONGO_OK = False
    chat_collection = None
    task_collection = None
    task_log_collection = None
    user_settings_collection = None
    shared_folders_collection = None
    temp_links_collection = None
    skills_collection = None
    projects_collection = None
    org_user_collection = None
    active_sessions_col = None
    active_tasks_col = None
    api_key_state_col = None
    scheduler_lock_col = None
    token_usage_col = None

# ============ API 키 라운드 로빈 (MongoDB 글로벌 카운터) ============
ANTHROPIC_API_KEYS = []
_env_keys = os.environ.get("ANTHROPIC_API_KEYS", "")
if _env_keys:
    ANTHROPIC_API_KEYS = [k.strip() for k in _env_keys.split(",") if k.strip()]
elif os.environ.get("ANTHROPIC_API_KEY"):
    ANTHROPIC_API_KEYS = [os.environ["ANTHROPIC_API_KEY"]]
ANTHROPIC_API_KEYS = [k for k in ANTHROPIC_API_KEYS if k and not k.startswith("#")]

import itertools, threading
_key_cycle = itertools.cycle(ANTHROPIC_API_KEYS) if ANTHROPIC_API_KEYS else None
_key_lock = threading.Lock()


def get_next_api_key():
    """라운드 로빈으로 다음 API 키 반환 (로컬 폴백, async 버전은 아래)"""
    if not _key_cycle:
        return ""
    with _key_lock:
        return next(_key_cycle)

async def get_next_api_key_async():
    """MongoDB 글로벌 카운터로 API 키 분배 (서버 간 겹침 방지)"""
    if not ANTHROPIC_API_KEYS:
        return ""
    if MONGO_OK and api_key_state_col is not None:
        try:
            doc = await api_key_state_col.find_one_and_update(
                {"_id": "key_counter"},
                {"$inc": {"index": 1}},
                upsert=True,
                return_document=True
            )
            idx = doc.get("index", 0) % len(ANTHROPIC_API_KEYS)
            return ANTHROPIC_API_KEYS[idx]
        except:
            pass
    return get_next_api_key()

# 호환성 유지
ANTHROPIC_API_KEY = ANTHROPIC_API_KEYS[0] if ANTHROPIC_API_KEYS else ""

# ============ 모델 선택 ============
def get_max_tokens(model: str) -> int:
    return MAX_TOKENS_SONNET if model == MODEL_SONNET else MAX_TOKENS_OPUS

def select_model(user_message: str) -> str:
    """사용자 메시지를 분석하여 적절한 모델 선택"""
    msg = user_message.lower().strip()
    # Opus 사용 키워드: 콘텐츠 생성, 복잡한 작업
    opus_keywords = [
        '만들어', '생성', '작성', '디자인', '변환', '인포그래픽', '슬라이드', '프레젠테이션',
        'pptx', 'ppt', '파워포인트', '보고서', '웹사이트', '홈페이지', '랜딩', 'html',
        '크롤링', '스크립트', '코드', '프로그램', '개발', '구현', '분석해',
        '엑셀', 'xlsx', '차트', '그래프', '시각화', '요약해', '정리해',
        '번역', '리서치', '조사', '검색해서', '찾아서', 'figma',
        '기획', '제안서', '계획', '전략', '아이디어', '브레인스토밍',
        'create', 'generate', 'build', 'design', 'develop', 'write', 'make',
    ]
    # Sonnet 사용 키워드: 단순 작업
    sonnet_keywords = [
        '파일 목록', '목록 보여', '보여줘', '열어', '읽어', '파일 확인',
        '이름 변경', '이동', '삭제', '폴더', '복사',
        '몇 개', '뭐가 있', '어디에', '경로',
        'list', 'show', 'read', 'rename', 'move', 'delete',
    ]
    # Sonnet 키워드 우선 체크 (단순 작업)
    for kw in sonnet_keywords:
        if kw in msg:
            return MODEL_SONNET
    # Opus 키워드 체크 (복잡한 작업)
    for kw in opus_keywords:
        if kw in msg:
            return MODEL_OPUS
    # 기본값: Sonnet (비용 절감)
    return MODEL_SONNET

IS_WINDOWS = sys.platform == "win32"
os.makedirs(WORKSPACE_ROOT, exist_ok=True)
USERNAME_PATTERN = re.compile(r'^[a-zA-Z0-9_.-]{1,50}$')

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app):
    """서버 시작/종료 시 실행 (FastAPI lifespan)"""
    # --- startup ---
    if MONGO_OK and temp_links_collection is not None:
        try:
            await temp_links_collection.create_index("expires_at", expireAfterSeconds=0)
            await temp_links_collection.create_index("token", unique=True)
        except Exception:
            pass
    if MONGO_OK:
        try:
            if active_sessions_col is not None:
                await active_sessions_col.create_index("username", unique=True)
                await active_sessions_col.create_index("last_active", expireAfterSeconds=1800)
            if active_tasks_col is not None:
                await active_tasks_col.create_index("username", unique=True)
                await active_tasks_col.create_index("started_at", expireAfterSeconds=600)
            if scheduler_lock_col is not None:
                await scheduler_lock_col.create_index("locked_at", expireAfterSeconds=300)
        except Exception as e:
            print(f"[STARTUP] index error: {e}")
    asyncio.create_task(_temp_cleanup_scheduler())
    if MONGO_OK and skills_collection is not None:
        try:
            await skills_collection.create_index("owner")
            await skills_collection.create_index("shared_with")
            await skills_collection.create_index([("owner", 1), ("name", 1)], unique=True)
        except Exception:
            pass
    # ── 관리자 대시보드용 인덱스 ──
    if MONGO_OK:
        try:
            # chat_logs: 대시보드에서 created_at 범위 조회 + username 집계에 사용
            if chat_collection is not None:
                await chat_collection.create_index("created_at")
                await chat_collection.create_index("updated_at")
                await chat_collection.create_index("username")
                # 복합: 월별/일별 사용자 수 (created_at 범위 + username distinct)
                await chat_collection.create_index([("created_at", 1), ("username", 1)])
                # 프로젝트 대화 카운트
                await chat_collection.create_index([("created_at", 1), ("project_id", 1)])
            # task_logs: started_at 범위 조회
            if task_collection is not None:
                await task_collection.create_index("started_at")
            # task_log_collection (REST 스케줄 작업): created_at 범위 조회
            if task_log_collection is not None:
                await task_log_collection.create_index("created_at")
            # org DB: lid 조회 (관리자 확인 + 사용자 이름 조회)
            if org_user_collection is not None:
                await org_user_collection.create_index("lid")
            # token_usage: 토큰 사용량 통계
            if token_usage_col is not None:
                await token_usage_col.create_index("created_at")
                await token_usage_col.create_index("username")
                await token_usage_col.create_index("model")
                await token_usage_col.create_index("service_type")
                await token_usage_col.create_index([("created_at", 1), ("username", 1)])
                await token_usage_col.create_index([("created_at", 1), ("model", 1)])
            print("[STARTUP] dashboard indexes created")
        except Exception as e:
            print(f"[STARTUP] dashboard index error: {e}")
    print(f"[STARTUP] server started: {SERVER_ID}")
    _build_index_cache()
    print(f"[STARTUP] static version: {_app_version}")
    yield
    # --- shutdown ---
    print(f"[SHUTDOWN] server stopped: {SERVER_ID}")

app = FastAPI(title=APP_TITLE, lifespan=lifespan)

# CORS
from starlette.middleware.cors import CORSMiddleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
    allow_credentials=True,
    expose_headers=["Content-Disposition", "Content-Type", "Content-Length"],
)

# Static 파일 UTF-8 charset 보장 미들웨어
@app.middleware("http")
async def add_charset_to_static(request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/static/"):
        ct = response.headers.get("content-type", "")
        if ct and ("text/" in ct or "javascript" in ct or "json" in ct) and "charset" not in ct:
            response.headers["content-type"] = ct + "; charset=utf-8"
    return response

async def _temp_cleanup_scheduler():
    """각 사용자 워크스페이스/_temp 폴더 내 1일 경과 파일 자동 삭제 (1시간마다, MongoDB 락으로 중복 방지)"""
    TEMP_MAX_AGE_DAYS = 1
    CLEANUP_INTERVAL = 3600
    while True:
        try:
            await asyncio.sleep(60)
            # MongoDB 분산 락: 하나의 서버만 실행
            if MONGO_OK and scheduler_lock_col is not None:
                try:
                    await scheduler_lock_col.insert_one({
                        "_id": "temp_cleanup",
                        "server_id": SERVER_ID,
                        "locked_at": datetime.now(timezone.utc)
                    })
                except:
                    # 이미 다른 서버가 락 보유 → 건너뛰기
                    await asyncio.sleep(CLEANUP_INTERVAL)
                    continue
            ws_root = Path(WORKSPACE_ROOT)
            if ws_root.exists():
                now = time.time()
                cutoff = now - (TEMP_MAX_AGE_DAYS * 86400)
                cleaned = 0
                for user_dir in ws_root.iterdir():
                    if not user_dir.is_dir():
                        continue
                    temp_dir = user_dir / "_temp"
                    if not temp_dir.exists():
                        continue
                    for f in temp_dir.rglob("*"):
                        try:
                            if f.is_file() and f.stat().st_mtime < cutoff:
                                f.unlink()
                                cleaned += 1
                        except:
                            pass
                    for d in sorted(temp_dir.rglob("*"), reverse=True):
                        try:
                            if d.is_dir() and not any(d.iterdir()):
                                d.rmdir()
                        except:
                            pass
                if cleaned > 0:
                    print(f"[_temp cleanup] {SERVER_ID}: {cleaned}개 파일 삭제 (>{TEMP_MAX_AGE_DAYS}일)")
        except Exception as e:
            print(f"[_temp cleanup error] {e}")
        await asyncio.sleep(CLEANUP_INTERVAL)

# ============================================================
# 백그라운드 작업 관리자
# ============================================================
class TaskManager:
    """분산 서버 환경 지원 TaskManager.
    - 로컬: running_tasks(asyncio.Task), user_connections(WebSocket) — 서버별 로컬
    - MongoDB: task_status, user_active_task — 서버 간 공유
    """
    def __init__(self):
        self.running_tasks: dict = {}         # {task_id: asyncio.Task} — 로컬만
        self.task_buffers: dict = {}          # {task_id: [msg_dict, ...]} — 로컬 버퍼
        self.user_connections: dict = {}      # {username: set(ws)} — 로컬 WS
        self.task_username: dict = {}         # {task_id: username} — 로컬
        # 로컬 캐시 (MongoDB와 동기화, 같은 서버 내 빠른 조회용)
        self.task_status: dict = {}           # {task_id: status} — 로컬 캐시
        self.user_active_task: dict = {}      # {username: task_id} — 로컬 캐시

    def register_ws(self, username: str, ws):
        if username not in self.user_connections:
            self.user_connections[username] = set()
        self.user_connections[username].add(ws)

    def unregister_ws(self, username: str, ws):
        if username in self.user_connections:
            self.user_connections[username].discard(ws)

    async def set_active_task(self, username: str, task_id: str):
        """MongoDB + 로컬 캐시에 활성 작업 등록"""
        self.user_active_task[username] = task_id
        self.task_status[task_id] = "running"
        if MONGO_OK and active_tasks_col is not None:
            try:
                await active_tasks_col.update_one(
                    {"username": username},
                    {"$set": {"task_id": task_id, "server_id": SERVER_ID, "status": "running",
                              "started_at": datetime.now(timezone.utc)}},
                    upsert=True
                )
            except:
                pass

    async def get_active_task(self, username: str):
        """MongoDB에서 사용자의 활성 작업 조회"""
        if MONGO_OK and active_tasks_col is not None:
            try:
                doc = await active_tasks_col.find_one({"username": username})
                if doc and doc.get("status") == "running":
                    return doc.get("task_id"), doc.get("server_id")
            except:
                pass
        return None, None

    async def clear_active_task(self, username: str):
        """MongoDB + 로컬 캐시에서 활성 작업 제거"""
        self.user_active_task.pop(username, None)
        if MONGO_OK and active_tasks_col is not None:
            try:
                await active_tasks_col.delete_one({"username": username})
            except:
                pass

    async def set_task_status(self, task_id: str, status: str):
        """MongoDB + 로컬 캐시에 작업 상태 업데이트"""
        self.task_status[task_id] = status
        if MONGO_OK and active_tasks_col is not None:
            try:
                await active_tasks_col.update_one(
                    {"task_id": task_id},
                    {"$set": {"status": status, "updated_at": datetime.now(timezone.utc)}}
                )
            except:
                pass

    async def is_user_busy(self, username: str) -> bool:
        """사용자가 작업 중인지 확인 (로컬 + MongoDB 분산 체크)"""
        # 1. 로컬 캐시 먼저 확인
        local_tid = self.user_active_task.get(username)
        if local_tid:
            t = self.running_tasks.get(local_tid)
            if t and not t.done():
                return True
            # 로컬 Task 완료됨 → 정리
            self.user_active_task.pop(username, None)
            await self.clear_active_task(username)
            return False
        # 2. MongoDB에서 다른 서버의 활성 작업 확인
        task_id, srv = await self.get_active_task(username)
        if not task_id:
            return False
        if srv == SERVER_ID:
            # 같은 서버인데 로컬에 없으면 좀비 → 정리
            await self.clear_active_task(username)
            return False
        # 다른 서버에서 작업 중
        return True

    async def broadcast(self, username: str, msg: dict):
        """사용자에게 메시지 전송. 연결 없으면 버퍼에 저장"""
        # 로컬 버퍼 저장
        task_id = None
        if MONGO_OK and active_tasks_col is not None:
            try:
                doc = await active_tasks_col.find_one({"username": username, "server_id": SERVER_ID})
                if doc:
                    task_id = doc.get("task_id")
            except:
                pass
        if task_id:
            if task_id not in self.task_buffers:
                self.task_buffers[task_id] = []
            self.task_buffers[task_id].append(msg)
        # 로컬 WS에 전송
        dead = set()
        for ws in self.user_connections.get(username, set()):
            try:
                await ws.send_json(msg)
            except:
                dead.add(ws)
        for ws in dead:
            self.user_connections.get(username, set()).discard(ws)

    def get_buffered_messages(self, task_id: str) -> list:
        return self.task_buffers.get(task_id, [])

    def cleanup_task(self, task_id: str):
        self.running_tasks.pop(task_id, None)

tm = TaskManager()

# ============================================================
# Tools / Helpers
# ============================================================
TOOLS = [
    {"name":"list_files","description":"작업 공간의 파일과 디렉토리 목록을 조회합니다.","input_schema":{"type":"object","properties":{"path":{"type":"string","default":"."}}}},
    {"name":"read_file","description":"텍스트 파일 내용을 읽습니다.","input_schema":{"type":"object","properties":{"path":{"type":"string"},"encoding":{"type":"string","default":"utf-8"}},"required":["path"]}},
    {"name":"write_file","description":"파일을 생성하거나 덮어씁니다.","input_schema":{"type":"object","properties":{"path":{"type":"string"},"content":{"type":"string"},"encoding":{"type":"string","default":"utf-8"}},"required":["path","content"]}},
    {"name":"edit_file","description":"파일의 특정 텍스트를 찾아서 교체합니다.","input_schema":{"type":"object","properties":{"path":{"type":"string"},"old_text":{"type":"string"},"new_text":{"type":"string"}},"required":["path","old_text","new_text"]}},
    {"name":"delete_file","description":"파일이나 디렉토리를 삭제합니다.","input_schema":{"type":"object","properties":{"path":{"type":"string"}},"required":["path"]}},
    {"name":"create_directory","description":"디렉토리를 생성합니다.","input_schema":{"type":"object","properties":{"path":{"type":"string"}},"required":["path"]}},
    {"name":"run_command","description":"셸 명령어를 실행합니다.","input_schema":{"type":"object","properties":{"command":{"type":"string"},"timeout":{"type":"integer","default":30}},"required":["command"]}},
    {"name":"search_files","description":"파일 내용에서 텍스트를 검색합니다.","input_schema":{"type":"object","properties":{"pattern":{"type":"string"},"path":{"type":"string","default":"."},"file_pattern":{"type":"string","default":""}},"required":["pattern"]}},
    {"name":"file_info","description":"파일 상세 정보를 조회합니다.","input_schema":{"type":"object","properties":{"path":{"type":"string"}},"required":["path"]}},
    {"name":"read_excel","description":"엑셀 파일을 읽습니다.","input_schema":{"type":"object","properties":{"path":{"type":"string"},"sheet_name":{"type":"string","default":""},"max_rows":{"type":"integer","default":100}},"required":["path"]}},
    {"name":"web_search","description":"Perplexity API로 인터넷 검색합니다.","input_schema":{"type":"object","properties":{"query":{"type":"string"},"detail_level":{"type":"string","enum":["brief","detailed"],"default":"detailed"}},"required":["query"]}},
    {"name":"write_temp_file","description":"임시 스크립트를 _temp 폴더에 작성합니다.","input_schema":{"type":"object","properties":{"filename":{"type":"string","default":"script.py"},"content":{"type":"string"}},"required":["content"]}},
    {"name":"figma_get_file","description":"Figma 파일의 디자인 구조(노드 트리, 스타일, 레이아웃)를 가져옵니다. URL에서 file_key를 추출하여 사용합니다. depth로 깊이를 제한할 수 있습니다.","input_schema":{"type":"object","properties":{"file_key":{"type":"string","description":"Figma 파일 키 (URL의 /design/{file_key}/ 또는 /file/{file_key}/ 부분)"},"node_id":{"type":"string","default":"","description":"특정 노드 ID (URL의 node-id 파라미터 값, 예: 1-2). 빈 문자열이면 전체 파일"},"depth":{"type":"integer","default":3,"description":"노드 트리 탐색 깊이 (1-5)"}},"required":["file_key"]}},
    {"name":"figma_get_images","description":"Figma 노드를 이미지(PNG/SVG)로 내보냅니다. 아이콘, 일러스트 등 이미지 에셋 추출에 사용합니다.","input_schema":{"type":"object","properties":{"file_key":{"type":"string","description":"Figma 파일 키"},"node_ids":{"type":"array","items":{"type":"string"},"description":"내보낼 노드 ID 배열 (예: [\"1:2\", \"3:4\"])"},"format":{"type":"string","enum":["png","svg","jpg","pdf"],"default":"png","description":"이미지 포맷"},"scale":{"type":"number","default":2,"description":"배율 (1-4)"}},"required":["file_key","node_ids"]}},
    {"name":"figma_get_styles","description":"Figma 파일에서 사용된 색상, 폰트, 이펙트 등 스타일 정보를 추출합니다.","input_schema":{"type":"object","properties":{"file_key":{"type":"string","description":"Figma 파일 키"}},"required":["file_key"]}}
]
BLOCKED_COMMANDS = ["rm -rf /","del /f /s /q C:\\","format ","mkfs","dd if=",":(){","shutdown","reboot","halt","poweroff"]
# 서버에서 GUI/브라우저 실행을 방지할 명령어 패턴
GUI_COMMANDS = ["start ", "xdg-open", "open ", "explorer ", "sensible-browser", "gnome-open", "kde-open", "wslview",
                "chrome", "firefox", "msedge", "iexplore", "safari", "opera", "brave", "chromium", "notepad", "code "]

def get_user_workspace(username):
    if not username or not USERNAME_PATTERN.match(username): return WORKSPACE_ROOT
    d = os.path.join(WORKSPACE_ROOT, username); os.makedirs(d, exist_ok=True); return d

def safe_path(rp, ws=None):
    w = Path(ws or WORKSPACE_ROOT).resolve(); t = (w / rp).resolve()
    if not str(t).startswith(str(w)): raise ValueError("작업 공간 외부 접근 차단")
    return t

def decode_bytes(data):
    if not data: return ""
    for e in ["utf-8","utf-8-sig","cp949","euc-kr","latin-1"]:
        try: return data.decode(e)
        except: continue
    return data.decode("utf-8", errors="replace")

def fmt_size(n):
    for u in ["B","KB","MB","GB"]:
        if n < 1024: return f"{n:.1f}{u}"
        n /= 1024
    return f"{n:.1f}TB"

def ok(d): return json.dumps(d, ensure_ascii=False)
def err(m): return json.dumps({"error": m}, ensure_ascii=False)

# ============================================================
# Async I/O
# ============================================================
async def aio_read_text(fp, preferred="utf-8"):
    for enc in [preferred, "utf-8", "utf-8-sig", "cp949", "euc-kr", "latin-1"]:
        try:
            async with aiofiles.open(str(fp), "r", encoding=enc) as f: return await f.read()
        except: continue
    async with aiofiles.open(str(fp), "r", encoding="utf-8", errors="replace") as f: return await f.read()

async def aio_write_text(fp, content, enc="utf-8"):
    async with aiofiles.open(str(fp), "w", encoding=enc) as f: await f.write(content)
    return len(content)

async def aio_write_bytes(fp, data):
    async with aiofiles.open(str(fp), "wb") as f: await f.write(data)
    return len(data)

async def aio_stat(fp): return await aiofiles.os.stat(str(fp))
async def aio_exists(fp): return await asyncio.to_thread(fp.exists)
async def aio_is_dir(fp): return await asyncio.to_thread(fp.is_dir)
async def aio_is_file(fp): return await asyncio.to_thread(fp.is_file)
async def aio_listdir(dp): return await asyncio.to_thread(lambda: sorted(dp.iterdir()))
async def aio_mkdir(dp): await asyncio.to_thread(partial(dp.mkdir, parents=True, exist_ok=True))
async def aio_rmtree(dp): await asyncio.to_thread(shutil.rmtree, str(dp))
async def aio_unlink(fp): await aiofiles.os.remove(str(fp))

async def aio_run_command(command, cwd, timeout=30):
    try:
        p = await asyncio.create_subprocess_shell(command, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE, cwd=cwd)
        try: so, se = await asyncio.wait_for(p.communicate(), timeout=min(timeout, 120))
        except asyncio.TimeoutError: p.kill(); await p.communicate(); return {"stdout":"","stderr":"타임아웃","returncode":-1}
        return {"stdout": decode_bytes(so)[-5000:], "stderr": decode_bytes(se)[-2000:], "returncode": p.returncode}
    except Exception as e: return {"stdout":"","stderr":str(e),"returncode":-1}

async def aio_search_files(pattern, sp, ws_dir, ext=""):
    root = Path(ws_dir).resolve()
    def collect():
        files = []
        for r, ds, fs in os.walk(str(sp)):
            ds[:] = [d for d in ds if not d.startswith(".") and not d.startswith("_")]
            for fn in fs:
                if ext and not fn.endswith(ext): continue
                files.append(Path(r) / fn)
                if len(files) >= 500: return files
        return files
    fl = await asyncio.to_thread(collect); matches = []
    for fp in fl:
        if len(matches) >= 50: break
        try:
            txt = await aio_read_text(fp)
            for i, line in enumerate(txt.splitlines(), 1):
                if pattern.lower() in line.lower():
                    matches.append(f"{fp.relative_to(root)}:{i}: {line.strip()[:200]}")
                    if len(matches) >= 50: break
        except: continue
    return matches

async def aio_read_excel(fp, sheet="", max_rows=100):
    def _r():
        import openpyxl; wb = openpyxl.load_workbook(str(fp), read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active; rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows: break
            rows.append([str(c) if c is not None else "" for c in row])
        r = {"sheet":ws.title,"all_sheets":wb.sheetnames,"total_rows":len(rows),"headers":rows[0] if rows else[],"data":rows}
        wb.close(); return r
    return await asyncio.to_thread(_r)

async def aio_perplexity(query, detail="detailed"):
    sm = "You are a helpful search assistant. Provide accurate info with source URLs. Respond in the same language as the query."
    if detail == "brief": sm += " Keep concise, under 200 words."
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post("https://api.perplexity.ai/chat/completions", json={"model":"sonar","messages":[{"role":"system","content":sm},{"role":"user","content":query}],"max_tokens":2048 if detail=="detailed" else 512,"temperature":0.2,"return_citations":True}, headers={"Authorization":f"Bearer {PERPLEXITY_API_KEY}","Content-Type":"application/json"})
        r.raise_for_status(); d = r.json()
    return {"query":query,"answer":d["choices"][0]["message"]["content"] if d.get("choices") else "","citations":d.get("citations",[])}

# ============================================================
# Figma API 헬퍼
# ============================================================
async def get_user_figma_token(username: str) -> Optional[str]:
    """MongoDB에서 사용자의 Figma 토큰 조회"""
    if not MONGO_OK or user_settings_collection is None:
        return None
    doc = await user_settings_collection.find_one({"username": username})
    if doc and doc.get("figma_token"):
        return doc["figma_token"]
    return None

def parse_figma_url(url: str) -> dict:
    """Figma URL에서 file_key와 node_id 추출"""
    import urllib.parse
    result = {"file_key": "", "node_id": ""}
    # https://www.figma.com/design/XXXXX/Name?node-id=1-2
    # https://www.figma.com/file/XXXXX/Name
    m = re.search(r'figma\.com/(?:design|file|proto)/([a-zA-Z0-9]+)', url)
    if m:
        result["file_key"] = m.group(1)
    parsed = urllib.parse.urlparse(url)
    params = urllib.parse.parse_qs(parsed.query)
    if "node-id" in params:
        result["node_id"] = params["node-id"][0]
    return result

async def figma_api_request(endpoint: str, token: str, params: dict = None) -> dict:
    """Figma REST API 호출"""
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.get(
            f"https://api.figma.com/v1/{endpoint}",
            headers={"X-Figma-Token": token},
            params=params or {}
        )
        if r.status_code == 403:
            return {"error": "Figma 토큰 권한이 없습니다. 토큰을 확인하거나 파일 접근 권한을 확인하세요."}
        if r.status_code == 404:
            return {"error": "Figma 파일을 찾을 수 없습니다. file_key를 확인하세요."}
        if r.status_code == 429:
            return {"error": "Figma API 호출 한도 초과. 잠시 후 다시 시도하세요."}
        if r.status_code != 200:
            return {"error": f"Figma API 오류 ({r.status_code}): {r.text[:300]}"}
        return r.json()

def simplify_figma_node(node: dict, depth: int, max_depth: int) -> dict:
    """Figma 노드 트리를 Claude가 이해할 수 있는 구조로 간소화"""
    simplified = {
        "id": node.get("id"),
        "name": node.get("name"),
        "type": node.get("type"),
    }
    # 크기/위치 정보
    bbox = node.get("absoluteBoundingBox")
    if bbox:
        simplified["bounds"] = {"x": bbox.get("x"), "y": bbox.get("y"), "w": bbox.get("width"), "h": bbox.get("height")}
    # 배경/채움 색상
    fills = node.get("fills", [])
    if fills:
        simplified["fills"] = [{"type": f.get("type"), "color": f.get("color"), "opacity": f.get("opacity", 1)} for f in fills[:5]]
    # 폰트 정보
    style = node.get("style")
    if style:
        simplified["font"] = {
            "family": style.get("fontFamily"), "size": style.get("fontSize"),
            "weight": style.get("fontWeight"), "align": style.get("textAlignHorizontal"),
            "lineHeight": style.get("lineHeightPx")
        }
    # 텍스트 내용
    if node.get("characters"):
        simplified["text"] = node["characters"][:500]
    # 레이아웃 모드
    if node.get("layoutMode"):
        simplified["layout"] = {
            "mode": node.get("layoutMode"), "padding": node.get("paddingLeft"),
            "gap": node.get("itemSpacing"), "align": node.get("primaryAxisAlignItems")
        }
    # 모서리 둥글기
    if node.get("cornerRadius"):
        simplified["cornerRadius"] = node.get("cornerRadius")
    # 스트로크
    strokes = node.get("strokes", [])
    if strokes:
        simplified["strokes"] = [{"type": s.get("type"), "color": s.get("color")} for s in strokes[:3]]
    # 이펙트
    effects = node.get("effects", [])
    if effects:
        simplified["effects"] = [{"type": e.get("type"), "radius": e.get("radius"), "color": e.get("color")} for e in effects[:5]]
    # 자식 노드 (깊이 제한)
    children = node.get("children", [])
    if children and depth < max_depth:
        simplified["children"] = [simplify_figma_node(c, depth + 1, max_depth) for c in children]
    elif children:
        simplified["children_count"] = len(children)
    return simplified

async def aio_figma_get_file(file_key: str, token: str, node_id: str = "", depth: int = 3) -> dict:
    """Figma 파일 구조 조회"""
    params = {"depth": min(depth, 5)}
    if node_id:
        # 특정 노드만 조회
        nid = node_id.replace("-", ":")
        endpoint = f"files/{file_key}/nodes"
        params["ids"] = nid
    else:
        endpoint = f"files/{file_key}"
    data = await figma_api_request(endpoint, token, params)
    if "error" in data:
        return data
    # 파일 기본 정보
    result = {
        "name": data.get("name", ""),
        "lastModified": data.get("lastModified", ""),
        "version": data.get("version", ""),
    }
    # 노드 트리 간소화
    if node_id and "nodes" in data:
        nid = node_id.replace("-", ":")
        node_data = data["nodes"].get(nid, {})
        doc = node_data.get("document", {})
        if doc:
            result["node"] = simplify_figma_node(doc, 0, depth)
    elif "document" in data:
        doc = data["document"]
        result["node"] = simplify_figma_node(doc, 0, depth)
    return result

async def aio_figma_get_images(file_key: str, node_ids: list, token: str, fmt: str = "png", scale: float = 2) -> dict:
    """Figma 노드를 이미지로 내보내기"""
    ids_str = ",".join([nid.replace("-", ":") for nid in node_ids])
    data = await figma_api_request(f"images/{file_key}", token, {"ids": ids_str, "format": fmt, "scale": min(scale, 4)})
    if "error" in data:
        return data
    return {"images": data.get("images", {}), "format": fmt, "scale": scale}

async def aio_figma_get_styles(file_key: str, token: str) -> dict:
    """Figma 파일의 스타일 정보 추출"""
    data = await figma_api_request(f"files/{file_key}/styles", token)
    if "error" in data:
        return data
    styles = []
    meta = data.get("meta", {})
    for s in meta.get("styles", []):
        styles.append({
            "key": s.get("key"), "name": s.get("name"),
            "type": s.get("style_type"), "description": s.get("description", "")
        })
    return {"styles": styles, "count": len(styles)}

# ============================================================
# 프로젝트 output 파일 추적
# ============================================================
_snapshot_lock = {}  # 로컬 캐시 (같은 서버 내 중복 방지)

async def _snapshot_project_before_modify(project_id: str, username: str, ws_dir: str):
    """파일 수정 전에 프로젝트 원본 폴더 전체를 _snapshots/{날짜_시간}/ 에 백업.
    최대 5개 유지, 초과 시 가장 오래된 것 삭제.
    """
    if not MONGO_OK or projects_collection is None or not project_id:
        return
    try:
        from bson import ObjectId
        now = datetime.now(timezone.utc)
        now_local = now + timedelta(hours=9)
        folder_key = now_local.strftime("%Y-%m-%d_%H-%M")

        # 로컬 중복 방지
        if _snapshot_lock.get(project_id) == folder_key:
            return
        # MongoDB 분산 락: 다른 서버에서도 같은 분 내 스냅샷 방지
        try:
            result = await projects_collection.find_one_and_update(
                {"_id": ObjectId(project_id), "username": username,
                 "$or": [{"_snapshot_key": {"$ne": folder_key}}, {"_snapshot_key": {"$exists": False}}]},
                {"$set": {"_snapshot_key": folder_key}},
                return_document=True
            )
            if not result:
                return
        except:
            pass
        _snapshot_lock[project_id] = folder_key

        proj_dir = Path(ws_dir)
        if not proj_dir.exists():
            print(f"[SNAPSHOT] 프로젝트 폴더 없음: {proj_dir}")
            return

        # 복사할 항목 목록 확인
        items_to_copy = [item for item in proj_dir.iterdir() if item.name != "_snapshots"]
        if not items_to_copy:
            print(f"[SNAPSHOT] 복사할 파일 없음: {proj_dir}")
            _snapshot_lock.pop(project_id, None)
            return

        snapshots_base = proj_dir / "_snapshots"
        snap_dir = snapshots_base / folder_key
        os.makedirs(str(snap_dir), exist_ok=True)
        print(f"[SNAPSHOT] 스냅샷 시작: {snap_dir} (원본 항목 {len(items_to_copy)}개)")

        # 프로젝트 원본 파일/폴더를 스냅샷에 복사
        copy_errors = []
        for item in items_to_copy:
            try:
                dest = snap_dir / item.name
                if item.is_dir():
                    shutil.copytree(str(item), str(dest))
                    print(f"[SNAPSHOT]   폴더 복사: {item.name}/")
                else:
                    shutil.copy2(str(item), str(dest))
                    print(f"[SNAPSHOT]   파일 복사: {item.name} ({item.stat().st_size}B)")
            except Exception as ce:
                copy_errors.append(f"{item.name}: {ce}")
                print(f"[SNAPSHOT]   복사 실패: {item.name} - {ce}")

        # 복사된 파일 통계 계산
        file_count = 0
        total_size = 0
        for root, dirs, files in os.walk(str(snap_dir)):
            for fn in files:
                try:
                    fp = Path(root) / fn
                    total_size += fp.stat().st_size
                    file_count += 1
                except:
                    pass

        if file_count == 0:
            print(f"[SNAPSHOT] 복사된 파일 0개 - 스냅샷 취소. 에러: {copy_errors}")
            shutil.rmtree(str(snap_dir), ignore_errors=True)
            _snapshot_lock.pop(project_id, None)
            return

        # DB 업데이트
        doc = await projects_collection.find_one(
            {"_id": ObjectId(project_id), "username": username},
            {"snapshots": 1}
        )
        snapshots = doc.get("snapshots", []) if doc else []
        snapshots.append({
            "folder_key": folder_key,
            "created_at": now.isoformat(),
            "file_count": file_count,
            "total_size": total_size
        })

        # 최대 스냅샷 수 유지
        if len(snapshots) > MAX_SNAPSHOTS:
            snapshots.sort(key=lambda x: x.get("created_at", ""))
            to_remove = snapshots[:-MAX_SNAPSHOTS]
            snapshots = snapshots[-MAX_SNAPSHOTS:]
            for old in to_remove:
                old_dir = snapshots_base / old["folder_key"]
                if old_dir.exists():
                    shutil.rmtree(str(old_dir), ignore_errors=True)

        await projects_collection.update_one(
            {"_id": ObjectId(project_id), "username": username},
            {"$set": {"snapshots": snapshots, "updated_at": now}}
        )
        print(f"[SNAPSHOT] 완료: {folder_key} ({file_count}개 파일, {total_size}bytes)")
    except Exception as e:
        import traceback
        print(f"[SNAPSHOT ERROR] 프로젝트 {project_id}:")
        traceback.print_exc()

# ============================================================
# execute_tool
# ============================================================
async def execute_tool(name, inp, ws_dir=None, username=None):
    ws_dir = ws_dir or WORKSPACE_ROOT
    try:
        if name == "list_files":
            p = safe_path(inp.get("path","."), ws_dir)
            if not await aio_exists(p): return err("경로 없음")
            items = []
            for f in await aio_listdir(p):
                if f.name.startswith(".") or f.name.startswith("_"): continue
                try:
                    s = await aio_stat(f); isd = await aio_is_dir(f)
                    if isd:
                        # 폴더: 내부 아이템 수 계산
                        try:
                            children = [c for c in await aio_listdir(f) if not c.name.startswith(".") and not c.name.startswith("_")]
                            child_count = len(children)
                        except:
                            child_count = 0
                        items.append({"name":f.name,"type":"directory","size":None,"child_count":child_count,"modified":datetime.fromtimestamp(s.st_mtime).strftime("%Y-%m-%d %H:%M")})
                    else:
                        items.append({"name":f.name,"type":"file","size":s.st_size,"modified":datetime.fromtimestamp(s.st_mtime).strftime("%Y-%m-%d %H:%M")})
                except: items.append({"name":f.name,"type":"unknown","size":None,"modified":None})
            return ok({"path":inp.get("path","."),"items":items})
        elif name == "read_file":
            p = safe_path(inp["path"], ws_dir)
            if not await aio_exists(p): return err("파일 없음")
            if (await aio_stat(p)).st_size > 2_000_000: return err("2MB 초과")
            return ok({"path":inp["path"],"content":await aio_read_text(p, inp.get("encoding","utf-8"))})
        elif name == "write_file":
            p = safe_path(inp["path"], ws_dir); await aio_mkdir(p.parent)
            content = inp["content"]
            # HTML 파일에 Tailwind CDN 자동 주입
            if str(p).lower().endswith(".html") and "tailwindcss" not in content.lower() and "<head" in content.lower():
                tailwind_cdn = '<script src="https://cdn.tailwindcss.com"></script>'
                import re as _re_tw
                head_match = _re_tw.search(r'(<head[^>]*>)', content, _re_tw.IGNORECASE)
                if head_match:
                    pos = head_match.end()
                    content = content[:pos] + '\n' + tailwind_cdn + '\n' + content[pos:]
            return ok({"success":True,"path":inp["path"],"size":await aio_write_text(p, content, inp.get("encoding","utf-8"))})
        elif name == "write_temp_file":
            td = Path(ws_dir)/"_temp"; await aio_mkdir(td); fn = inp.get("filename","script.py")
            await aio_write_text(td/fn, inp["content"]); return ok({"success":True,"path":"_temp/"+fn,"temp":True})
        elif name == "edit_file":
            p = safe_path(inp["path"], ws_dir)
            if not await aio_exists(p): return err("파일 없음")
            c = await aio_read_text(p)
            if inp["old_text"] not in c: return err("교체 대상 없음")
            await aio_write_text(p, c.replace(inp["old_text"], inp["new_text"], 1)); return ok({"success":True,"path":inp["path"]})
        elif name == "delete_file":
            p = safe_path(inp["path"], ws_dir)
            if not await aio_exists(p): return err("경로 없음")
            if await aio_is_dir(p): await aio_rmtree(p)
            else: await aio_unlink(p)
            return ok({"success":True,"deleted":inp["path"]})
        elif name == "create_directory":
            await aio_mkdir(safe_path(inp["path"], ws_dir)); return ok({"success":True,"path":inp["path"]})
        elif name == "run_command":
            cmd = inp["command"]; t = min(inp.get("timeout",30), 120)
            if any(b in cmd.lower() for b in BLOCKED_COMMANDS): return err("차단된 명령어")
            # GUI/브라우저 실행 명령 차단 (서버에서 포그라운드 프로세스 방지)
            cmd_lower = cmd.lower().strip()
            if any(cmd_lower.startswith(g) or cmd_lower.startswith("cmd /c " + g) or cmd_lower.startswith("powershell " + g) for g in GUI_COMMANDS):
                return {"success": True, "note": "서버 환경에서는 GUI 프로그램 실행이 생략됩니다. 파일은 정상적으로 생성되었으며, 미리보기 버튼을 통해 확인할 수 있습니다."}
            td = Path(ws_dir)/"_temp"; await aio_mkdir(td); result = await aio_run_command(cmd, str(td), t)
            try:
                if await aio_exists(td):
                    for e in await aio_listdir(td):
                        try:
                            if await aio_is_dir(e): await aio_rmtree(e)
                            else: await aio_unlink(e)
                        except: pass
            except: pass
            return ok(result)
        elif name == "search_files":
            return ok({"pattern":inp["pattern"],"matches":await aio_search_files(inp["pattern"], safe_path(inp.get("path","."), ws_dir), ws_dir, inp.get("file_pattern",""))})
        elif name == "file_info":
            p = safe_path(inp["path"], ws_dir)
            if not await aio_exists(p): return err("경로 없음")
            s = await aio_stat(p); isd = await aio_is_dir(p)
            return ok({"path":inp["path"],"type":"directory" if isd else "file","size":s.st_size,"size_human":fmt_size(s.st_size),"modified":datetime.fromtimestamp(s.st_mtime).isoformat()})
        elif name == "read_excel":
            p = safe_path(inp["path"], ws_dir)
            if not await aio_exists(p): return err("파일 없음")
            return ok(await aio_read_excel(p, inp.get("sheet_name",""), inp.get("max_rows",100)))
        elif name == "web_search":
            if not PERPLEXITY_API_KEY: return err("PERPLEXITY_API_KEY 미설정")
            return ok(await aio_perplexity(inp["query"], inp.get("detail_level","detailed")))
        elif name == "figma_get_file":
            token = await get_user_figma_token(username) if username else None
            if not token: return err("Figma 토큰이 설정되지 않았습니다. 설정에서 Figma Personal Access Token을 등록해주세요.")
            result = await aio_figma_get_file(inp["file_key"], token, inp.get("node_id",""), inp.get("depth",3))
            return ok(result) if "error" not in result else err(result["error"])
        elif name == "figma_get_images":
            token = await get_user_figma_token(username) if username else None
            if not token: return err("Figma 토큰이 설정되지 않았습니다.")
            result = await aio_figma_get_images(inp["file_key"], inp["node_ids"], token, inp.get("format","png"), inp.get("scale",2))
            return ok(result) if "error" not in result else err(result["error"])
        elif name == "figma_get_styles":
            token = await get_user_figma_token(username) if username else None
            if not token: return err("Figma 토큰이 설정되지 않았습니다.")
            result = await aio_figma_get_styles(inp["file_key"], token)
            return ok(result) if "error" not in result else err(result["error"])
        return err(f"알 수 없는 도구: {name}")
    except ValueError as e: return err(str(e))
    except Exception as e: return err(f"{type(e).__name__}: {e}")

# ============================================================
# 대화 히스토리 자동 압축 (7턴 이상 시)
# ============================================================
AUTO_COMPRESS_THRESHOLD = int(os.environ.get("AUTO_COMPRESS_THRESHOLD", "7"))

def _count_user_turns(history: list) -> int:
    """히스토리에서 사용자 텍스트 메시지 수를 카운트 (tool_result 제외)"""
    count = 0
    for msg in history:
        if msg.get("role") != "user":
            continue
        content = msg.get("content")
        if isinstance(content, str):
            count += 1
        elif isinstance(content, list):
            # tool_result만으로 구성된 메시지는 제외
            has_text = any(
                isinstance(b, dict) and b.get("type") == "text"
                for b in content
            )
            if has_text:
                count += 1
    return count

async def _auto_compress_history(history: list, username: str) -> list:
    """대화 히스토리가 AUTO_COMPRESS_THRESHOLD 이상이면 자동으로 요약 압축.
    압축된 새 히스토리를 반환. 실패 시 원본 반환."""
    user_turns = _count_user_turns(history)
    if user_turns < AUTO_COMPRESS_THRESHOLD:
        return history

    print(f"[AUTO COMPRESS] {username}: {user_turns}턴 감지, 자동 압축 시작")
    try:
        # 대화 텍스트 추출
        conversation_text = ""
        msg_count = 0
        for m in history:
            role = m.get("role", "")
            content = m.get("content", "")
            if isinstance(content, str) and content.strip():
                conversation_text += f"\n[{role}]: {content[:2000]}\n"
                msg_count += 1
            elif isinstance(content, list):
                for block in content:
                    if isinstance(block, dict):
                        if block.get("type") == "text":
                            conversation_text += f"\n[{role}]: {block['text'][:2000]}\n"
                            msg_count += 1
                        elif block.get("type") == "tool_use":
                            conversation_text += f"\n[tool_use: {block.get('name','')}]\n"
                        elif block.get("type") == "tool_result":
                            ct = block.get("content", "")
                            if isinstance(ct, str):
                                conversation_text += f"\n[tool_result]: {ct[:500]}\n"

        api_key = await get_next_api_key_async()
        compress_client = anthropic.AsyncAnthropic(api_key=api_key)

        summary_response = await compress_client.messages.create(
            model=MODEL_SONNET,
            max_tokens=4096,
            system="You are a conversation summarizer. Summarize the entire conversation concisely but completely, preserving:\n1. All key decisions, results, and conclusions\n2. Important file paths, folder names, and technical details\n3. Any pending tasks or next steps\n4. User preferences expressed during the conversation\nWrite in the same language as the conversation. Be thorough but concise.",
            messages=[{"role": "user", "content": f"다음 대화를 요약해주세요. 핵심 내용, 결정사항, 작업 결과, 파일 경로 등 중요한 세부 사항을 모두 포함해야 합니다:\n\n{conversation_text[:50000]}"}]
        )

        if hasattr(summary_response, 'usage'):
            await record_token_usage(username=username, task_id="compress", session_id=user_histories.get(username, {}).get("session_id", ""), service_type="compress", model=MODEL_SONNET, usage=summary_response.usage, step=0, key_index=0)
        summary = ""
        for block in summary_response.content:
            if hasattr(block, "text"):
                summary += block.text

        # 새 히스토리: 요약 + 최근 2턴 유지
        new_history = [
            {"role": "user", "content": f"[이전 대화 요약]\n{summary}\n\n위 요약은 이전 대화의 압축된 컨텍스트입니다. 이 맥락을 기반으로 대화를 이어가겠습니다."},
            {"role": "assistant", "content": "네, 이전 대화 내용을 이해했습니다. 요약된 맥락을 바탕으로 계속 진행하겠습니다. 무엇을 도와드릴까요?"}
        ]
        # 최근 2턴(user+assistant) 유지
        recent = []
        for m in reversed(history):
            if isinstance(m.get("content"), str) and m.get("role") in ("user", "assistant"):
                recent.insert(0, m)
                if len(recent) >= 4:
                    break
            elif isinstance(m.get("content"), list):
                has_text = any(b.get("type") == "text" for b in m["content"] if isinstance(b, dict))
                if has_text and m.get("role") in ("user", "assistant"):
                    text_only = [b for b in m["content"] if isinstance(b, dict) and b.get("type") == "text"]
                    recent.insert(0, {"role": m["role"], "content": text_only})
                    if len(recent) >= 4:
                        break
        new_history.extend(recent)

        print(f"[AUTO COMPRESS] {username}: {len(history)}개 → {len(new_history)}개 메시지로 압축 완료")
        await tm.broadcast(username, {"type": "auto_compress", "old_count": len(history), "new_count": len(new_history), "message": f"대화가 길어져 자동 압축되었습니다. ({len(history)}개 → {len(new_history)}개)"})

        # DB에도 업데이트
        if MONGO_OK and chat_collection is not None:
            session_id = user_histories.get(username, {}).get("session_id", "")
            if session_id:
                try:
                    await chat_collection.update_one(
                        {"session_id": session_id},
                        {"$set": {"api_history": new_history, "compressed": True, "compressed_at": datetime.now(timezone.utc)}}
                    )
                except:
                    pass

        # user_histories에도 반영
        if username in user_histories:
            user_histories[username]["history"] = new_history

        return new_history
    except Exception as e:
        print(f"[AUTO COMPRESS ERROR] {username}: {type(e).__name__}: {e}")
        return history  # 실패 시 원본 반환

# ============================================================
# 백그라운드 에이전트 (브라우저 독립)
# ============================================================
async def run_agent_background(task_id: str, user_message: str, history: list, ws_dir: str, current_folder: str, username: str, images: list = None, forced_skill_name: str = "", project_id: str = ""):
    """백그라운드 작업 - WebSocket 없이 독립 실행, tm.broadcast로 전송"""
    # 사용자 이메일 조회 (로그용)
    _user_email = username  # 기본값
    if MONGO_OK and active_sessions_col is not None:
        try:
            _sess = await active_sessions_col.find_one({"username": username})
            if _sess: _user_email = _sess.get("email", username)
        except: pass
    print(f"[TASK START] {task_id[:8]} | {username} ({_user_email}) | model={select_model(user_message)} | msg={user_message[:80]}")
    selected_model = select_model(user_message)
    # 이미지가 포함된 경우 Opus 강제 선택 (Vision은 복잡한 작업)
    if images:
        selected_model = MODEL_OPUS
    api_key = await get_next_api_key_async()
    client = anthropic.AsyncAnthropic(api_key=api_key)

    # 프로젝트 모드: ws_dir을 프로젝트 폴더로 변경 (AI는 프로젝트 안의 파일만 참조/수정)
    _proj_id = project_id  # 원본 보존
    project_ctx = ""
    project_ws_dir = ws_dir  # 원래 워크스페이스 (output 추적용)
    if project_id and MONGO_OK and projects_collection is not None:
        try:
            from bson import ObjectId
            proj = await projects_collection.find_one({"_id": ObjectId(project_id)})
            if proj:
                # 프로젝트 전용 작업 디렉토리로 변경
                proj_dir = str(Path(ws_dir) / "_projects" / project_id)
                os.makedirs(proj_dir, exist_ok=True)
                ws_dir = proj_dir  # ★ 핵심: AI의 모든 도구가 프로젝트 폴더 안에서만 동작
                current_folder = "."  # 프로젝트 루트에서 시작

                project_ctx = f"\n\n<project_context>\n프로젝트: {proj.get('name','')}\n설명: {proj.get('description','')}"
                if proj.get("instructions"):
                    project_ctx += f"\n\n[프로젝트 지침 — 반드시 준수]\n{proj['instructions']}"
                    project_ctx += "\n※ 위 프로젝트 지침은 모든 응답에서 반드시 참조하고 준수해야 합니다. 지침에 명시된 규칙, 스타일, 제약조건을 항상 따르세요."
                # 참고 파일 내용 포함
                for fc_item in proj.get("files_content", []):
                    if fc_item.get("content"):
                        project_ctx += f"\n\n[참고 파일: {fc_item['name']}]\n{fc_item['content'][:20000]}"
                project_ctx += "\n\n★ 중요: 모든 파일 읽기/쓰기는 이 프로젝트 폴더 안에서만 수행하세요. 프로젝트 폴더 밖의 파일은 접근하지 마세요."
                project_ctx += "\n</project_context>"
        except:
            pass

    fc = ""
    if not project_id and current_folder and current_folder != ".":
        fc = f"\n\n현재 사용자가 선택한 작업 폴더: {current_folder}\n이 폴더 내의 파일만 우선적으로 참조하여 작업하세요."

    # 활성 스킬 로드 (사용자 메시지 기반 자동 선택 또는 강제 선택)
    skills_prompt, skill_names = await _get_active_skills_prompt(username, user_message, forced_skill_name=forced_skill_name)
    system_prompt = f"""당신은 {APP_ASSISTANT_NAME}입니다.
작업 공간({ws_dir})에서 파일 관리, 코드 작성, 명령어 실행을 도와줍니다.
규칙: 파일 경로는 상대 경로 사용, 수정 전 내용 확인, 한국어 응답, 임시 스크립트는 write_temp_file 사용
Figma 변환 요청 시: figma_get_file로 디자인 구조를 먼저 가져온 후, figma_get_styles로 스타일 정보를 확인하고, 필요한 이미지는 figma_get_images로 추출하세요. URL에서 file_key와 node-id를 파싱하세요.

[Tailwind CSS 필수 규칙]
HTML 파일을 생성할 때 반드시 Tailwind CSS를 사용하세요:
1. <head> 안에 반드시 <script src="https://cdn.tailwindcss.com"></script>를 포함하세요.
2. 인라인 <style> 대신 Tailwind 유틸리티 클래스(flex, p-4, text-lg, bg-white, rounded-xl 등)를 우선 사용하세요.
3. Tailwind으로 불가능한 커스텀 스타일만 <style> 태그나 tailwind.config로 보완하세요.
4. 반응형 디자인은 Tailwind 브레이크포인트(sm:, md:, lg:, xl:)를 적극 활용하세요.
5. 다크모드가 필요하면 dark: 접두사를 활용하세요.
이 규칙은 인포그래픽, 보고서, 웹사이트, 랜딩페이지, Figma 변환 등 모든 HTML 생성에 적용됩니다.{fc}{project_ctx}
현재: {datetime.now().isoformat()} | OS: {"Windows" if IS_WINDOWS else "Linux/Mac"}{skills_prompt}"""

    # 히스토리 자동 압축 (7턴 이상 시)
    history = await _auto_compress_history(history, username)

    # 이미지가 있으면 multimodal content block 구성
    if images:
        content_blocks = []
        for img in images:
            content_blocks.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": img.get("media_type", "image/png"),
                    "data": img["data"]
                }
            })
        content_blocks.append({"type": "text", "text": user_message})
        history.append({"role": "user", "content": content_blocks})
    else:
        history.append({"role":"user","content":user_message})
    full_response = ""
    started_at = datetime.now(timezone.utc)
    _key_idx = ANTHROPIC_API_KEYS.index(api_key) + 1 if api_key in ANTHROPIC_API_KEYS else 0
    _model_label = "Opus" if selected_model == MODEL_OPUS else "Sonnet"
    await tm.broadcast(username, {"type":"model_info","model":_model_label,"key_index":_key_idx,"key_total":len(ANTHROPIC_API_KEYS)})
    # 스킬 사용 로그 전송
    if skill_names:
        matched = [s["name"] for s in skill_names if s.get("matched")]
        print(f"[SKILLS] {username}: 매칭 {len(matched)}개 - {matched}")
        await tm.broadcast(username, {"type":"skills_info","skills":skill_names})
    elif skills_prompt:
        print(f"[SKILLS] {username}: 카탈로그 모드 (매칭 없음)")
        # 카탈로그 모드에서도 알림 전송 안 함 (매칭된 스킬 없음)
    else:
        print(f"[SKILLS] {username}: 활성 스킬 없음")

    # MongoDB에 작업 시작 기록
    if MONGO_OK and task_collection is not None:
        await task_collection.insert_one({
            "task_id": task_id, "username": username, "message": user_message,
            "status": "running", "started_at": started_at, "completed_at": None,
            "response_summary": "", "current_folder": current_folder,
            "model": selected_model, "key_index": _key_idx
        })

    try:
        for step in range(1, 11):
            await tm.broadcast(username, {"type":"progress","step":step,"message":f"분석 중... (단계 {step})"})

            # API 호출 시 이전 턴의 이미지 블록을 텍스트로 교체 (토큰 절약)
            # 현재(마지막) 사용자 메시지의 이미지만 유지
            api_history = []
            for i, msg in enumerate(history):
                if isinstance(msg.get("content"), list):
                    is_last_user = (i == len(history) - 1) or (i >= len(history) - 2 and msg.get("role") == "user" and not any(
                        h.get("role") == "user" for h in history[i+1:]
                    ))
                    if msg.get("role") == "user" and not is_last_user:
                        # 이전 턴의 이미지를 텍스트로 교체
                        new_content = []
                        for block in msg["content"]:
                            if isinstance(block, dict) and block.get("type") == "image":
                                mt = block.get("source", {}).get("media_type", "image/png")
                                new_content.append({"type": "text", "text": f"[이전에 첨부한 이미지: {mt}]"})
                            else:
                                new_content.append(block)
                        api_history.append({"role": msg["role"], "content": new_content})
                    else:
                        api_history.append(msg)
                else:
                    api_history.append(msg)

            final_message = None
            for retry in range(5):
                try:
                    async with client.messages.stream(model=selected_model, max_tokens=get_max_tokens(selected_model), system=system_prompt, tools=TOOLS, messages=api_history) as stream:
                        async for event in stream:
                            if event.type == "content_block_delta" and event.delta.type == "text_delta":
                                chunk = event.delta.text; full_response += chunk
                                await tm.broadcast(username, {"type":"text_delta","content":chunk})
                            elif event.type == "content_block_start":
                                if event.content_block.type == "text":
                                    await tm.broadcast(username, {"type":"text_start"})
                                elif event.content_block.type == "tool_use":
                                    await tm.broadcast(username, {"type":"tool_start","tool":event.content_block.name,"id":event.content_block.id})
                        final_message = await stream.get_final_message()
                        stop_reason = final_message.stop_reason
                        if final_message and hasattr(final_message, 'usage'):
                            _sid = user_histories.get(username, {}).get("session_id", task_id)
                            await record_token_usage(username=username, task_id=task_id, session_id=_sid, service_type="chat", model=selected_model, usage=final_message.usage, step=step, key_index=_key_idx)
                    break
                except anthropic.RateLimitError:
                    # 다른 API 키로 교체 시도
                    new_key = await get_next_api_key_async()
                    if new_key and new_key != api_key:
                        client = anthropic.AsyncAnthropic(api_key=new_key)
                        api_key = new_key
                        _key_idx = ANTHROPIC_API_KEYS.index(api_key) + 1 if api_key in ANTHROPIC_API_KEYS else 0
                        await tm.broadcast(username, {"type":"model_info","model":_model_label,"key_index":_key_idx,"key_total":len(ANTHROPIC_API_KEYS),"switched":True})
                    wait = min(2 ** retry * 10, 120)
                    await tm.broadcast(username, {"type":"rate_limit","wait":wait,"retry":retry+1,"max_retry":5,"message":f"API 사용량 초과, Key #{_key_idx}로 전환 후 {wait}초 대기... ({retry+1}/5)"})
                    for remaining in range(wait, 0, -1):
                        await tm.broadcast(username, {"type":"rate_limit_tick","remaining":remaining,"total":wait,"retry":retry+1,"max_retry":5})
                        await asyncio.sleep(1)
                    await tm.broadcast(username, {"type":"rate_limit_resume","retry":retry+1})
                    if retry == 4:
                        await tm.broadcast(username, {"type":"error","content":"API 사용량 제한으로 요청을 처리할 수 없습니다. 잠시 후 다시 시도해주세요."}); break
                except anthropic.APIConnectionError as e:
                    print(f"[API CONNECTION ERROR] {username}: {e}")
                    if retry < 4:
                        wait = min(2 ** retry * 5, 60)
                        await tm.broadcast(username, {"type":"error","content":f"API 연결 오류 (재시도 {retry+1}/5, {wait}초 후)..."})
                        await asyncio.sleep(wait)
                        continue
                    await tm.broadcast(username, {"type":"error","content":"API 서버 연결에 실패했습니다. 잠시 후 다시 시도해주세요."})
                    break
                except anthropic.APIStatusError as e:
                    print(f"[API STATUS ERROR] {username}: {e.status_code} - {e.message}")
                    if e.status_code == 529:  # Overloaded
                        if retry < 4:
                            wait = min(2 ** retry * 15, 120)
                            await tm.broadcast(username, {"type":"error","content":f"AI 서버가 과부하 상태입니다. {wait}초 후 재시도합니다... ({retry+1}/5)"})
                            await asyncio.sleep(wait)
                            continue
                        await tm.broadcast(username, {"type":"error","content":"AI 서버가 과부하 상태입니다. 잠시 후 다시 시도해주세요."})
                        break
                    elif "prompt is too long" in str(e).lower() or "context" in str(e).lower() or "too many tokens" in str(e).lower():
                        await tm.broadcast(username, {"type":"error","content":"컨텍스트 길이가 초과되었습니다. 대화를 압축해주세요.","suggest_compress":True})
                    else:
                        await tm.broadcast(username, {"type":"error","content":f"API 오류 ({e.status_code}): {e.message}"})
                    break
                except anthropic.APIError as e:
                    error_msg = str(e)
                    print(f"[API ERROR] {username}: {error_msg}")
                    if "overloaded" in error_msg.lower():
                        if retry < 4:
                            wait = min(2 ** retry * 15, 120)
                            await tm.broadcast(username, {"type":"error","content":f"AI 서버가 과부하 상태입니다. {wait}초 후 재시도합니다... ({retry+1}/5)"})
                            await asyncio.sleep(wait)
                            continue
                        await tm.broadcast(username, {"type":"error","content":"AI 서버가 과부하 상태입니다. 잠시 후 다시 시도해주세요."})
                    elif "prompt is too long" in error_msg.lower() or "context" in error_msg.lower() or "too many tokens" in error_msg.lower():
                        await tm.broadcast(username, {"type":"error","content":"컨텍스트 길이가 초과되었습니다. 대화를 압축해주세요.","suggest_compress":True})
                    else:
                        await tm.broadcast(username, {"type":"error","content":f"API 오류: {e}"})
                    break
                except asyncio.CancelledError:
                    raise  # 취소는 상위로 전파
                except Exception as e:
                    error_msg = str(e)
                    print(f"[STREAM ERROR] {username}: {type(e).__name__}: {error_msg}")
                    if "overloaded" in error_msg.lower():
                        if retry < 4:
                            wait = min(2 ** retry * 15, 120)
                            await tm.broadcast(username, {"type":"error","content":f"AI 서버가 과부하 상태입니다. {wait}초 후 재시도합니다... ({retry+1}/5)"})
                            await asyncio.sleep(wait)
                            continue
                        await tm.broadcast(username, {"type":"error","content":"AI 서버가 과부하 상태입니다. 잠시 후 다시 시도해주세요."})
                        break
                    if retry < 2:
                        await asyncio.sleep(3)
                        continue
                    await tm.broadcast(username, {"type":"error","content":f"예상치 못한 오류: {type(e).__name__}: {e}"})
                    break

            if final_message is None: break

            history.append({"role":"assistant","content":final_message.content})
            tool_uses = [b for b in final_message.content if b.type == "tool_use"]
            total_tools = len(tool_uses)
            for idx, b in enumerate(tool_uses):
                await tm.broadcast(username, {"type":"tool_call","step":step,"tool":b.name,"input":b.input,"id":b.id,"tool_index":idx+1,"tool_total":total_tools})
            await tm.broadcast(username, {"type":"text_end"})

            if not tool_uses: break

            tool_results = []
            for idx, tu in enumerate(tool_uses):
                await tm.broadcast(username, {"type":"tool_executing","tool":tu.name,"id":tu.id,"tool_index":idx+1,"tool_total":total_tools})
                # ★ 프로젝트 모드: 파일 수정 전에 원본 스냅샷 생성
                if project_id and tu.name in ("write_file", "edit_file"):
                    await _snapshot_project_before_modify(project_id, username, ws_dir)
                rs = await execute_tool(tu.name, tu.input, ws_dir, username); rj = json.loads(rs)
                # tool_result 크기 제한 (8000자)
                if len(rs) > 8000:
                    rs = rs[:8000] + '...(truncated)'
                tool_results.append({"type":"tool_result","tool_use_id":tu.id,"content":rs})
                await tm.broadcast(username, {"type":"tool_result","tool":tu.name,"id":tu.id,"success":"error" not in rj,"result":rj,"tool_index":idx+1,"tool_total":total_tools})
            history.append({"role":"user","content":tool_results})
            if stop_reason == "end_turn": break

        await tm.set_task_status(task_id, "done")
        _sid = user_histories.get(username, {}).get("session_id", task_id)
        await tm.broadcast(username, {"type":"done","steps":step,"task_id":task_id,"session_id":_sid})
        print(f"[TASK DONE] {task_id[:8]} | {username} ({_user_email}) | steps={step}")

    except asyncio.CancelledError:
        # 사용자 취소: 중간 히스토리를 저장하여 나중에 이어서 진행 가능
        await tm.set_task_status(task_id, "cancelled")
        # 중간까지의 응답이 있으면 assistant 메시지로 기록
        if full_response:
            history.append({"role": "assistant", "content": full_response + "\n\n[⏹ 작업이 중지되었습니다]"})
        print(f"[TASK CANCELLED] {task_id[:8]} | {username} ({_user_email}) | steps={step} | history_len={len(history)}")

    except Exception as e:
        await tm.set_task_status(task_id, "error")
        await tm.broadcast(username, {"type":"error","content":f"작업 오류: {e}"})
        _sid = user_histories.get(username, {}).get("session_id", task_id)
        await tm.broadcast(username, {"type":"done","steps":0,"task_id":task_id,"session_id":_sid})
        print(f"[TASK ERROR] {task_id[:8]} | {username} ({_user_email}) | {type(e).__name__}: {e}")

    finally:
        completed_at = datetime.now(timezone.utc)
        # ★ 취소/에러 시에도 중간 히스토리를 user_histories에 반영 (계속 진행 시 이어받기 위해)
        if username in user_histories:
            user_histories[username]["history"] = history
        # asyncio.shield로 감싸서 CancelledError 중에도 DB 저장이 완료되도록 보장
        try:
            if MONGO_OK and task_collection is not None:
                await asyncio.shield(task_collection.update_one({"task_id":task_id}, {"$set":{
                    "status": tm.task_status.get(task_id,"done"),
                    "completed_at": completed_at,
                    "duration_seconds": (completed_at - started_at).total_seconds(),
                    "response_summary": full_response[:500]
                }}))
            if MONGO_OK and chat_collection is not None:
                session_id = user_histories.get(username, {}).get("session_id", task_id)
                _proj_id = user_histories.get(username, {}).get("project_id", "")
                await asyncio.shield(save_history_to_db(session_id, username, history, user_message, full_response, started_at, completed_at, current_folder=current_folder, project_id=_proj_id))
        except (asyncio.CancelledError, Exception) as e:
            print(f"[SAVE WARNING] {task_id[:8]} | {username} | finally 저장 중 오류: {type(e).__name__}: {e}")
        await tm.clear_active_task(username)
        tm.cleanup_task(task_id)

    return history

# ============================================================
# 라우트 - JWT 기반
# ============================================================
def _jwt_expired_page():
    """JWT 만료 시 표시하는 안내 페이지"""
    return f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><title>세션 만료 - {APP_TITLE}</title>
<style>*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Noto Sans KR',sans-serif;background:#F5F7FA;height:100vh;display:flex;align-items:center;justify-content:center}}
.box{{background:#fff;border-radius:16px;padding:48px;text-align:center;box-shadow:0 8px 32px rgba(0,0,0,.08);max-width:460px}}
.icon{{width:64px;height:64px;background:linear-gradient(135deg,#FF6B6B,#EE5A24);border-radius:16px;display:inline-grid;place-items:center;color:#fff;font-size:28px;margin-bottom:20px}}
h2{{font-size:20px;color:#1A1A2E;margin-bottom:8px}}p{{color:#6B7280;font-size:14px;line-height:1.6}}
.sub{{margin-top:16px;padding:12px 16px;background:#FFF5F5;border:1px solid #FED7D7;border-radius:8px;font-size:12px;color:#C53030}}
.btn{{display:inline-block;margin-top:20px;padding:10px 32px;background:linear-gradient(135deg,#4A7CFF,#06B6D4);color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;text-decoration:none}}
.btn:hover{{opacity:.9}}</style></head>
<body><div class="box"><div class="icon">⏱</div><h2>세션이 만료되었습니다</h2>
<p>인증 토큰의 유효시간이 경과했습니다.<br>포털에서 다시 접속해주세요.</p>
<div class="sub">🔒 보안을 위해 일정 시간이 지나면 자동으로 세션이 만료됩니다</div>
<a class="btn" href="{PORTAL_URL or 'javascript:window.close();'}">{APP_BRAND}로 이동</a>
</div></body></html>"""

def _jwt_invalid_page():
    """JWT 무효 시 표시하는 안내 페이지"""
    return f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><title>인증 실패 - {APP_TITLE}</title>
<style>*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Noto Sans KR',sans-serif;background:#F5F7FA;height:100vh;display:flex;align-items:center;justify-content:center}}
.box{{background:#fff;border-radius:16px;padding:48px;text-align:center;box-shadow:0 8px 32px rgba(0,0,0,.08);max-width:460px}}
.icon{{width:64px;height:64px;background:linear-gradient(135deg,#FF6B6B,#EE5A24);border-radius:16px;display:inline-grid;place-items:center;color:#fff;font-size:28px;margin-bottom:20px}}
h2{{font-size:20px;color:#1A1A2E;margin-bottom:8px}}p{{color:#6B7280;font-size:14px;line-height:1.6}}
.sub{{margin-top:16px;padding:12px 16px;background:#FFF5F5;border:1px solid #FED7D7;border-radius:8px;font-size:12px;color:#C53030}}
.btn{{display:inline-block;margin-top:20px;padding:10px 32px;background:linear-gradient(135deg,#4A7CFF,#06B6D4);color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;text-decoration:none}}
.btn:hover{{opacity:.9}}</style></head>
<body><div class="box"><div class="icon">🔐</div><h2>인증에 실패했습니다</h2>
<p>유효하지 않은 토큰입니다.<br>포털에서 다시 접속해주세요.</p>
<div class="sub">⚠️ 직접 URL 접속이나 변조된 토큰은 허용되지 않습니다</div>
<a class="btn" href="javascript:window.close();" onclick="window.close();">창 닫기</a>
</div></body></html>"""

_rendered_index_cache = None  # 치환 완료된 HTML (기본 한국어)
_rendered_lang_cache = {}     # {lang: 치환 완료된 HTML}
_app_version = None

def _compute_app_version():
    """정적 파일 버전 결정. 우선순위:
    1) APP_VERSION 환경변수 (.env에서 설정)
    2) static/version.txt 파일 내용
    3) 없으면 주요 정적 파일의 mtime 조합으로 자동 생성
    파일이 물리적으로 교체되지 않는 한 서버를 몇 번 재시작해도 동일한 값."""
    # 1) .env 환경변수
    if APP_VERSION:
        return APP_VERSION
    # 2) 명시적 버전 파일
    try:
        with open("static/version.txt", "r") as f:
            v = f.read().strip()
            if v:
                return v
    except FileNotFoundError:
        pass
    # 3) 파일 수정시간 기반 자동 버전
    max_mtime = 0
    for fp in ["static/js/app.js", "static/css/app.css", "static/index.html",
               "static/lang/ko.json", "static/lang/en.json", "static/lang/ja.json", "static/lang/zh.json"]:
        try:
            mt = os.path.getmtime(fp)
            if mt > max_mtime:
                max_mtime = mt
        except FileNotFoundError:
            pass
    # mtime을 짧은 hex 문자열로 변환 (정수 부분만, 소수점 이하 무시)
    return format(int(max_mtime), 'x')[-8:]

def _build_index_cache():
    """서버 시작 시 한 번 호출. 치환 완료된 HTML을 메모리에 캐시."""
    global _rendered_index_cache, _rendered_lang_cache, _app_version
    _app_version = _compute_app_version()
    with open("static/index.html", "r", encoding="utf-8") as f:
        template = f.read()
    # 모든 템플릿 변수 치환
    replacements = {
        "{{PORTAL_URL}}": PORTAL_URL or "",
        "{{KPORTAL_URL}}": PORTAL_URL or "",  # 호환용
        "{{APP_VERSION}}": _app_version,
        "{{APP_TITLE}}": APP_TITLE,
        "{{APP_BRAND}}": APP_BRAND,
        "{{APP_BRAND_SUB}}": APP_BRAND_SUB,
        "{{APP_BRAND_ICON}}": APP_BRAND_ICON,
        "{{APP_BRAND_COMPANY}}": APP_BRAND_COMPANY,
        "{{APP_WELCOME_TITLE}}": APP_WELCOME_TITLE,
    }
    rendered = template
    for key, val in replacements.items():
        rendered = rendered.replace(key, val)
    _rendered_index_cache = rendered
    # 언어별 캐시
    _rendered_lang_cache = {}
    for lang in SUPPORTED_LANGS:
        h = rendered.replace('lang="ko"', f'lang="{lang}"', 1)
        lang_script = f'<script>window.__LANG__="{lang}";</script>'
        h = h.replace('</head>', lang_script + '</head>', 1)
        _rendered_lang_cache[lang] = h

def _serve_index():
    """캐시된 HTML 즉시 반환 (치환 없음, O(1))"""
    if _rendered_index_cache is None:
        _build_index_cache()
    return HTMLResponse(_rendered_index_cache, media_type="text/html; charset=utf-8")

def _serve_index_lang(lang: str):
    """언어별 캐시된 HTML 즉시 반환"""
    if not _rendered_lang_cache:
        _build_index_cache()
    html = _rendered_lang_cache.get(lang, _rendered_index_cache)
    return HTMLResponse(html, media_type="text/html; charset=utf-8")

def invalidate_index_cache():
    """파일 변경 시 캐시 무효화 (관리 API에서 호출 가능)"""
    global _rendered_index_cache, _rendered_lang_cache, _app_version
    _rendered_index_cache = None
    _rendered_lang_cache = {}
    _app_version = None

# ─── 모바일 앱 업데이트 서비스 (인증 불필요) ───
# 디렉토리 구조:
#   _mobile_apps/
#     android/version.json   ← {"version":"1.0.1","download_url":"/api/mobile/download?platform=android","release_notes":"..."}
#     ios/version.json       ← {"version":"1.0.1","download_url":"https://apps.apple.com/...","release_notes":"..."}
#     android/*.apk          ← APK 파일
#   (폴백) _mobile_apps/version.json ← 플랫폼 디렉토리 없을 때 사용
MOBILE_APPS_DIR = os.path.join(WORKSPACE_ROOT, "_mobile_apps")

@app.get("/api/mobile/version")
async def mobile_version_check(platform: str = ""):
    """모바일 앱 버전 체크. platform=android|ios 로 플랫폼별 분기.
    1순위: _mobile_apps/{platform}/version.json 파일
    2순위: _mobile_apps/version.json (폴백)
    3순위: .env의 MOBILE_VERSION 값"""
    import json as _json

    # 1순위: 플랫폼별 version.json
    if platform in ("android", "ios"):
        pf_version = os.path.join(MOBILE_APPS_DIR, platform, "version.json")
        if os.path.isfile(pf_version):
            try:
                with open(pf_version, "r", encoding="utf-8") as f:
                    data = _json.load(f)
                data.setdefault("platform", platform)
                return data
            except Exception:
                pass

    # 2순위: 공통 version.json
    common_version = os.path.join(MOBILE_APPS_DIR, "version.json")
    if os.path.isfile(common_version):
        try:
            with open(common_version, "r", encoding="utf-8") as f:
                data = _json.load(f)
            if platform:
                data["platform"] = platform
            return data
        except Exception:
            pass

    # 3순위: .env 폴백
    from dotenv import dotenv_values
    env = dotenv_values(".env")
    version = env.get("MOBILE_VERSION", "v_01")
    result = {"version": version}
    if platform:
        result["platform"] = platform
    return result

@app.get("/api/mobile/download")
async def mobile_download(platform: str = "android"):
    """모바일 앱 다운로드 (Android APK).
    _mobile_apps/{platform}/ 디렉토리에서 가장 최신 APK 파일을 반환."""
    if platform == "ios":
        # iOS는 APK 다운로드 불가 → version.json의 download_url(App Store)로 안내
        return JSONResponse(
            {"error": "iOS는 App Store에서 다운로드하세요", "redirect": True},
            status_code=400
        )

    # Android APK 찾기
    apk_dir = os.path.join(MOBILE_APPS_DIR, "android")
    if not os.path.isdir(apk_dir):
        # 폴백: _mobile_apps 루트에서 APK 탐색
        apk_dir = MOBILE_APPS_DIR

    if not os.path.isdir(apk_dir):
        raise HTTPException(404, "APK 파일이 없습니다. _mobile_apps/android/ 디렉토리를 확인하세요.")

    # .apk 파일 중 최신(수정일 기준) 선택
    apk_files = sorted(
        [f for f in Path(apk_dir).glob("*.apk") if f.is_file()],
        key=lambda f: f.stat().st_mtime,
        reverse=True
    )
    if not apk_files:
        raise HTTPException(404, "APK 파일이 없습니다. _mobile_apps/android/ 디렉토리에 APK를 업로드하세요.")

    apk = apk_files[0]
    return FileResponse(
        str(apk),
        filename=apk.name,
        media_type="application/vnd.android.package-archive"
    )


@app.post("/{token}/api/admin/reload-cache")
async def admin_reload_cache(token: str):
    """서버 재시작 없이 .env 리로드 + 정적 파일 캐시 재빌드.
    .env의 APP_VERSION 등을 변경한 후 이 API를 호출하면 즉시 반영됩니다."""
    username = userid_from_jwt(token)
    if not username:
        raise HTTPException(401, "인증 실패")
    # .env 리로드
    global APP_VERSION, _env_loaded
    try:
        from dotenv import load_dotenv
        _env_candidates = []
        try:
            _env_candidates.append(Path(__file__).resolve().parent / ".env")
        except:
            pass
        _env_candidates.append(Path(".env").resolve())
        _env_candidates.append(Path(os.getcwd()) / ".env")
        for _ep in _env_candidates:
            try:
                if _ep.exists():
                    load_dotenv(str(_ep), override=True)
                    _env_loaded = True
                    break
            except:
                continue
    except ImportError:
        pass
    # APP_VERSION 갱신
    APP_VERSION = os.environ.get("APP_VERSION", "")
    # 캐시 재빌드
    invalidate_index_cache()
    _build_index_cache()
    print(f"[ADMIN] cache reloaded by {username} | version: {_app_version}")
    return {"ok": True, "version": _app_version}

# ============================================================
# 관리자 대시보드 API
# ============================================================

async def _check_admin(token: str) -> str:
    """관리자 권한 확인. admin이면 username 반환, 아니면 HTTPException"""
    username = userid_from_jwt(token) if token.count(".") == 2 else (token if USERNAME_PATTERN.match(token) else None)
    if not username:
        raise HTTPException(401, "인증 실패")
    # 직접 접속 (개발모드) → admin 허용
    if USERNAME_PATTERN.match(token):
        return username
    # .env의 ADMIN_USERS 확인
    if username in ADMIN_USERS:
        return username
    # 조직도에서 role 확인 (여러 필드명 대응)
    if MONGO_OK and org_user_collection is not None:
        try:
            doc = await org_user_collection.find_one({"lid": username})
            if doc:
                role = (doc.get("role") or doc.get("auth") or doc.get("grade") or doc.get("level") or doc.get("type") or "")
                if role == "admin":
                    return username
        except:
            pass
    raise HTTPException(403, "관리자 권한이 필요합니다")

def _parse_date_range(start_date: str = "", end_date: str = ""):
    """start_date/end_date (YYYY-MM-DD) 파싱. 비어있으면 None 반환"""
    s = e = None
    try:
        if start_date:
            s = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        if end_date:
            e = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        pass
    return s, e

@app.get("/{token}/api/admin/dashboard/summary")
async def admin_dashboard_summary(token: str, start_date: str = "", end_date: str = ""):
    """대시보드 요약 통계 (날짜 범위 지원)"""
    await _check_admin(token)
    if not MONGO_OK:
        raise HTTPException(500, "MongoDB 미연결")
    try:
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        range_start, range_end = _parse_date_range(start_date, end_date)
        # 날짜 범위 필터
        date_filter = {}
        if range_start and range_end:
            date_filter = {"created_at": {"$gte": range_start, "$lte": range_end}}
        elif range_start:
            date_filter = {"created_at": {"$gte": range_start}}

        if date_filter:
            range_users = len(await chat_collection.distinct("username", date_filter)) if chat_collection is not None else 0
            range_chats = await chat_collection.count_documents(date_filter) if chat_collection is not None else 0
            task_f = {"started_at": {"$gte": range_start}} if range_start else {}
            if range_end:
                task_f["started_at"] = {**(task_f.get("started_at") or {}), "$lte": range_end}
            range_tasks = await task_collection.count_documents(task_f) if task_collection is not None and task_f else 0
        else:
            range_users = 0; range_chats = 0; range_tasks = 0

        total_users = len(await chat_collection.distinct("username")) if chat_collection is not None else 0
        today_users = len(await chat_collection.distinct("username", {"updated_at": {"$gte": today_start}})) if chat_collection is not None else 0
        total_chats = await chat_collection.count_documents({}) if chat_collection is not None else 0
        today_chats = await chat_collection.count_documents({"created_at": {"$gte": today_start}}) if chat_collection is not None else 0
        total_tasks = await task_collection.count_documents({}) if task_collection is not None else 0
        today_tasks = await task_collection.count_documents({"started_at": {"$gte": today_start}}) if task_collection is not None else 0
        total_projects = await projects_collection.count_documents({}) if projects_collection is not None else 0
        total_skills = await skills_collection.count_documents({}) if skills_collection is not None else 0

        return {
            "total_users": total_users, "today_users": today_users,
            "total_chats": total_chats, "today_chats": today_chats,
            "total_tasks": total_tasks, "today_tasks": today_tasks,
            "total_projects": total_projects, "total_skills": total_skills,
            "range_users": range_users, "range_chats": range_chats, "range_tasks": range_tasks,
        }
    except Exception as e:
        print(f"[ADMIN] summary error: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(500, f"Dashboard error: {str(e)}")

@app.get("/{token}/api/admin/dashboard/monthly")
async def admin_dashboard_monthly(token: str, year: int = 0):
    """월별 사용량 통계"""
    await _check_admin(token)
    try:
        if not MONGO_OK or chat_collection is None:
            return {"months": []}
        if year == 0:
            year = datetime.now(timezone.utc).year
        months = []
        for m in range(1, 13):
            start = datetime(year, m, 1, tzinfo=timezone.utc)
            end = datetime(year, m + 1, 1, tzinfo=timezone.utc) if m < 12 else datetime(year + 1, 1, 1, tzinfo=timezone.utc)
            chats = await chat_collection.count_documents({"created_at": {"$gte": start, "$lt": end}})
            tasks = await task_collection.count_documents({"started_at": {"$gte": start, "$lt": end}}) if task_collection is not None else 0
            users = len(await chat_collection.distinct("username", {"created_at": {"$gte": start, "$lt": end}}))
            months.append({"month": m, "chats": chats, "tasks": tasks, "users": users})
        return {"year": year, "months": months}
    except Exception as e:
        print(f"[ADMIN] monthly error: {e}")
        import traceback; traceback.print_exc()
        return {"year": year, "months": []}

@app.get("/{token}/api/admin/dashboard/users")
async def admin_dashboard_users(token: str, period: str = "month", start_date: str = "", end_date: str = ""):
    """사용자별 사용량 통계 (날짜 범위 지원)"""
    await _check_admin(token)
    try:
        if not MONGO_OK or chat_collection is None:
            return {"users": []}
        now = datetime.now(timezone.utc)
        range_start, range_end = _parse_date_range(start_date, end_date)
        if range_start:
            since = range_start
            until = range_end
        elif period == "today":
            since = now.replace(hour=0, minute=0, second=0, microsecond=0); until = None
        elif period == "week":
            since = now - timedelta(days=7); until = None
        elif period == "year":
            since = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0); until = None
        else:
            since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0); until = None

        match_filter = {"created_at": {"$gte": since}}
        if until:
            match_filter["created_at"]["$lte"] = until

        pipeline = [
            {"$match": match_filter},
            {"$group": {"_id": "$username", "chat_count": {"$sum": 1}, "last_active": {"$max": "$updated_at"}}},
            {"$sort": {"chat_count": -1}},
            {"$limit": 50}
        ]
        cursor = chat_collection.aggregate(pipeline)
        users = []
        async for doc in cursor:
            uname = doc["_id"]
            display_name = uname
            if org_user_collection is not None:
                try:
                    org_doc = await org_user_collection.find_one({"lid": uname}, {"nm": 1})
                    if org_doc:
                        display_name = org_doc.get("nm", uname)
                except:
                    pass
            la = doc.get("last_active")
            users.append({
                "username": uname, "display_name": display_name,
                "chat_count": doc["chat_count"],
                "last_active": la.isoformat() if hasattr(la, 'isoformat') else str(la) if la else ""
            })
        return {"period": period, "users": users}
    except Exception as e:
        print(f"[ADMIN] users error: {e}")
        import traceback; traceback.print_exc()
        return {"period": period, "users": []}

@app.get("/{token}/api/admin/dashboard/daily")
async def admin_dashboard_daily(token: str, days: int = 30, start_date: str = "", end_date: str = ""):
    """일별 사용량 추이 (날짜 범위 또는 최근 N일)"""
    await _check_admin(token)
    try:
        if not MONGO_OK or chat_collection is None:
            return {"days": []}
        range_start, range_end = _parse_date_range(start_date, end_date)
        now = datetime.now(timezone.utc)
        if range_start and range_end:
            # 날짜 범위 모드
            d = range_start
            result = []
            while d <= range_end:
                start = d.replace(hour=0, minute=0, second=0, microsecond=0)
                end = start + timedelta(days=1)
                chats = await chat_collection.count_documents({"created_at": {"$gte": start, "$lt": end}})
                users = len(await chat_collection.distinct("username", {"created_at": {"$gte": start, "$lt": end}}))
                result.append({"date": start.strftime("%m/%d"), "chats": chats, "users": users})
                d += timedelta(days=1)
            return {"days": result}
        else:
            result = []
            for i in range(days - 1, -1, -1):
                d = now - timedelta(days=i)
                start = d.replace(hour=0, minute=0, second=0, microsecond=0)
                end = start + timedelta(days=1)
                chats = await chat_collection.count_documents({"created_at": {"$gte": start, "$lt": end}})
                users = len(await chat_collection.distinct("username", {"created_at": {"$gte": start, "$lt": end}}))
                result.append({"date": start.strftime("%m/%d"), "chats": chats, "users": users})
            return {"days": result}
    except Exception as e:
        print(f"[ADMIN] daily error: {e}")
        import traceback; traceback.print_exc()
        return {"days": []}

@app.get("/{token}/api/admin/dashboard/features")
async def admin_dashboard_features(token: str, period: str = "month", start_date: str = "", end_date: str = ""):
    """기능별 사용량 (날짜 범위 지원)"""
    await _check_admin(token)
    try:
        if not MONGO_OK:
            return {"features": []}
        now = datetime.now(timezone.utc)
        range_start, range_end = _parse_date_range(start_date, end_date)
        if range_start:
            since = range_start
            until = range_end
        elif period == "today":
            since = now.replace(hour=0, minute=0, second=0, microsecond=0); until = None
        elif period == "week":
            since = now - timedelta(days=7); until = None
        elif period == "year":
            since = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0); until = None
        else:
            since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0); until = None

        chat_filter = {"created_at": {"$gte": since}}
        task_filter = {"started_at": {"$gte": since}}
        if until:
            chat_filter["created_at"]["$lte"] = until
            task_filter["started_at"]["$lte"] = until

        features = []
        # AI 대화
        if chat_collection is not None:
            cnt = await chat_collection.count_documents(chat_filter)
            if cnt > 0:
                features.append({"name": "AI 대화", "count": cnt})
        # 프로젝트 대화
        if chat_collection is not None:
            pf = dict(chat_filter)
            pf["project_id"] = {"$exists": True, "$ne": ""}
            cnt = await chat_collection.count_documents(pf)
            if cnt > 0:
                features.append({"name": "프로젝트", "count": cnt})
        # REST 스케줄 작업
        if task_collection is not None:
            cnt = await task_collection.count_documents(task_filter)
            if cnt > 0:
                features.append({"name": "스케줄 작업", "count": cnt})
        return {"period": period, "features": features}
    except Exception as e:
        print(f"[ADMIN] features error: {e}")
        import traceback; traceback.print_exc()
        return {"period": period, "features": []}

@app.get("/{token}/api/admin/dashboard/search-users")
async def admin_search_users(token: str, q: str = ""):
    """관리자용 사용자 검색 (이름/ID로 검색)"""
    await _check_admin(token)
    if not q or len(q.strip()) < 1:
        return {"users": []}
    q = q.strip()
    results = []
    # 1) 조직도에서 이름/ID 검색
    if MONGO_OK and org_user_collection is not None:
        try:
            cursor = org_user_collection.find(
                {"$or": [
                    {"lid": {"$regex": q, "$options": "i"}},
                    {"nm": {"$regex": q, "$options": "i"}},
                ]},
                {"_id": 0, "lid": 1, "nm": 1, "dp": 1}
            ).limit(20)
            async for doc in cursor:
                results.append({
                    "username": doc.get("lid", ""),
                    "name": doc.get("nm", ""),
                    "dept": doc.get("dp", ""),
                })
        except Exception:
            pass
    # 2) 조직도에 없으면 chat_logs에서 username 검색
    if not results and MONGO_OK and chat_collection is not None:
        try:
            usernames = await chat_collection.distinct("username")
            for u in usernames:
                if q.lower() in u.lower():
                    results.append({"username": u, "name": u, "dept": ""})
                if len(results) >= 20:
                    break
        except Exception:
            pass
    return {"users": results, "query": q}

@app.get("/{token}/api/admin/dashboard/user-detail")
async def admin_user_detail(token: str, username: str = "", start_date: str = "", end_date: str = ""):
    """개별 사용자 상세 대시보드 (날짜 범위 지원)"""
    await _check_admin(token)
    if not username:
        raise HTTPException(400, "username 필요")
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    range_start, range_end = _parse_date_range(start_date, end_date)

    # 사용자 기본 정보
    user_info = {"username": username, "name": username, "dept": ""}
    if MONGO_OK and org_user_collection is not None:
        try:
            doc = await org_user_collection.find_one({"lid": username}, {"_id": 0, "nm": 1, "dp": 1, "role": 1})
            if doc:
                user_info["name"] = doc.get("nm", username)
                user_info["dept"] = doc.get("dp", "")
                user_info["role"] = doc.get("role", "")
        except Exception:
            pass

    # 대화 통계
    total_chats = 0; today_chats = 0; month_chats = 0; year_chats = 0
    first_chat = None; last_chat = None
    if MONGO_OK and chat_collection is not None:
        try:
            total_chats = await chat_collection.count_documents({"username": username})
            today_chats = await chat_collection.count_documents({"username": username, "created_at": {"$gte": today_start}})
            month_chats = await chat_collection.count_documents({"username": username, "created_at": {"$gte": month_start}})
            year_chats = await chat_collection.count_documents({"username": username, "created_at": {"$gte": year_start}})
            # 첫 대화 / 마지막 대화
            first_doc = await chat_collection.find_one({"username": username}, sort=[("created_at", 1)])
            last_doc = await chat_collection.find_one({"username": username}, sort=[("created_at", -1)])
            if first_doc and first_doc.get("created_at"):
                first_chat = first_doc["created_at"].isoformat()
            if last_doc and last_doc.get("updated_at"):
                last_chat = last_doc["updated_at"].isoformat()
        except Exception:
            pass

    # 기간별 대화 수
    range_chats = 0
    if range_start and MONGO_OK and chat_collection is not None:
        try:
            rf = {"username": username, "created_at": {"$gte": range_start}}
            if range_end:
                rf["created_at"]["$lte"] = range_end
            range_chats = await chat_collection.count_documents(rf)
        except Exception:
            pass

    # 일별 추이
    daily = []
    if range_start and range_end:
        d = range_start
        while d <= range_end:
            start = d.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=1)
            cnt = 0
            if MONGO_OK and chat_collection is not None:
                try:
                    cnt = await chat_collection.count_documents({"username": username, "created_at": {"$gte": start, "$lt": end}})
                except Exception:
                    pass
            daily.append({"date": start.strftime("%m/%d"), "chats": cnt})
            d += timedelta(days=1)
    else:
        for i in range(29, -1, -1):
            d = now - timedelta(days=i)
            start = d.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=1)
            cnt = 0
            if MONGO_OK and chat_collection is not None:
                try:
                    cnt = await chat_collection.count_documents({"username": username, "created_at": {"$gte": start, "$lt": end}})
                except Exception:
                    pass
            daily.append({"date": start.strftime("%m/%d"), "chats": cnt})

    # 월별 추이 (올해)
    monthly = []
    for m in range(1, 13):
        start = datetime(now.year, m, 1, tzinfo=timezone.utc)
        end = datetime(now.year, m + 1, 1, tzinfo=timezone.utc) if m < 12 else datetime(now.year + 1, 1, 1, tzinfo=timezone.utc)
        cnt = 0
        if MONGO_OK and chat_collection is not None:
            try:
                cnt = await chat_collection.count_documents({"username": username, "created_at": {"$gte": start, "$lt": end}})
            except Exception:
                pass
        monthly.append({"month": m, "chats": cnt})

    # 프로젝트 수
    project_count = 0
    if MONGO_OK and projects_collection is not None:
        try:
            project_count = await projects_collection.count_documents({"owner": username})
        except Exception:
            pass

    # 파일 사용량
    file_stats = {"total_files": 0, "total_size": 0, "total_size_fmt": "0 B"}
    ws = os.path.join(WORKSPACE_ROOT, username)
    if os.path.isdir(ws):
        total_files = 0; total_size = 0
        for root, dirs, files in os.walk(ws):
            # _로 시작하는 폴더 제외 (스냅샷 등)
            dirs[:] = [d for d in dirs if not d.startswith("_") and not d.startswith(".")]
            for fn in files:
                if fn.startswith("."):
                    continue
                fp = os.path.join(root, fn)
                try:
                    total_files += 1
                    total_size += os.path.getsize(fp)
                except Exception:
                    pass
        file_stats = {"total_files": total_files, "total_size": total_size, "total_size_fmt": _fmt_size(total_size)}

    return {
        "user": user_info,
        "stats": {
            "total_chats": total_chats, "today_chats": today_chats,
            "month_chats": month_chats, "year_chats": year_chats,
            "range_chats": range_chats,
            "first_chat": first_chat, "last_chat": last_chat,
            "project_count": project_count,
        },
        "daily": daily,
        "monthly": monthly,
        "file_stats": file_stats,
    }

def _fmt_size(b):
    for u in ['B','KB','MB','GB','TB']:
        if b < 1024:
            return f"{b:.1f} {u}" if u != 'B' else f"{b} {u}"
        b /= 1024
    return f"{b:.1f} PB"

@app.get("/{token}/api/admin/dashboard/storage")
async def admin_dashboard_storage(token: str):
    """전체 스토리지 현황"""
    await _check_admin(token)
    total_files = 0; total_size = 0
    user_storage = []
    if os.path.isdir(WORKSPACE_ROOT):
        for entry in os.scandir(WORKSPACE_ROOT):
            if not entry.is_dir() or entry.name.startswith("."):
                continue
            uname = entry.name
            u_files = 0; u_size = 0
            for root, dirs, files in os.walk(entry.path):
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                for fn in files:
                    try:
                        u_files += 1
                        u_size += os.path.getsize(os.path.join(root, fn))
                    except Exception:
                        pass
            total_files += u_files
            total_size += u_size
            # 사용자 이름 조회
            display_name = uname
            if MONGO_OK and org_user_collection is not None:
                try:
                    doc = await org_user_collection.find_one({"lid": uname}, {"nm": 1})
                    if doc:
                        display_name = doc.get("nm", uname)
                except Exception:
                    pass
            user_storage.append({
                "username": uname, "display_name": display_name,
                "files": u_files, "size": u_size, "size_fmt": _fmt_size(u_size),
            })
    user_storage.sort(key=lambda x: x["size"], reverse=True)
    return {
        "total_files": total_files,
        "total_size": total_size,
        "total_size_fmt": _fmt_size(total_size),
        "workspace_path": WORKSPACE_ROOT,
        "users": user_storage[:50],
    }

# ============ 토큰 사용량 통계 API ============

@app.get("/{token}/api/admin/dashboard/token-summary")
async def admin_token_summary(token: str, start_date: str = "", end_date: str = ""):
    """토큰 사용량 요약 통계"""
    await _check_admin(token)
    if not MONGO_OK or token_usage_col is None:
        return {"total_input_tokens": 0, "total_output_tokens": 0, "total_tokens": 0, "total_cost": 0, "total_calls": 0, "range_input_tokens": 0, "range_output_tokens": 0, "range_tokens": 0, "range_cost": 0, "range_calls": 0, "today_tokens": 0, "today_cost": 0, "today_calls": 0}
    try:
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        range_start, range_end = _parse_date_range(start_date, end_date)
        # 전체 합계
        total_agg = await token_usage_col.aggregate([{"$group": {"_id": None, "input": {"$sum": "$input_tokens"}, "output": {"$sum": "$output_tokens"}, "tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}]).to_list(1)
        t = total_agg[0] if total_agg else {}
        # 오늘
        today_agg = await token_usage_col.aggregate([{"$match": {"created_at": {"$gte": today_start}}}, {"$group": {"_id": None, "tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}]).to_list(1)
        td = today_agg[0] if today_agg else {}
        # 기간 필터
        r = {}
        if range_start:
            date_filter = {"created_at": {"$gte": range_start}}
            if range_end:
                date_filter["created_at"]["$lte"] = range_end
            range_agg = await token_usage_col.aggregate([{"$match": date_filter}, {"$group": {"_id": None, "input": {"$sum": "$input_tokens"}, "output": {"$sum": "$output_tokens"}, "tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}]).to_list(1)
            r = range_agg[0] if range_agg else {}
        return {
            "total_input_tokens": t.get("input", 0), "total_output_tokens": t.get("output", 0),
            "total_tokens": t.get("tokens", 0), "total_cost": round(t.get("cost", 0), 2), "total_calls": t.get("calls", 0),
            "range_input_tokens": r.get("input", 0), "range_output_tokens": r.get("output", 0),
            "range_tokens": r.get("tokens", 0), "range_cost": round(r.get("cost", 0), 2), "range_calls": r.get("calls", 0),
            "today_tokens": td.get("tokens", 0), "today_cost": round(td.get("cost", 0), 2), "today_calls": td.get("calls", 0),
        }
    except Exception as e:
        print(f"[ADMIN] token-summary error: {e}")
        import traceback; traceback.print_exc()
        raise HTTPException(500, f"Token summary error: {str(e)}")

@app.get("/{token}/api/admin/dashboard/token-daily")
async def admin_token_daily(token: str, days: int = 30, start_date: str = "", end_date: str = ""):
    """일별 토큰 사용 추이"""
    await _check_admin(token)
    if not MONGO_OK or token_usage_col is None:
        return {"days": []}
    try:
        range_start, range_end = _parse_date_range(start_date, end_date)
        if range_start and range_end:
            s, e = range_start, range_end
        else:
            now = datetime.now(timezone.utc)
            e = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            s = (now - timedelta(days=days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)
        result = []
        current = s
        while current <= e:
            day_start = current.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = current.replace(hour=23, minute=59, second=59, microsecond=999999)
            agg = await token_usage_col.aggregate([
                {"$match": {"created_at": {"$gte": day_start, "$lte": day_end}}},
                {"$group": {"_id": None, "input": {"$sum": "$input_tokens"}, "output": {"$sum": "$output_tokens"}, "tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}
            ]).to_list(1)
            d = agg[0] if agg else {}
            result.append({"date": current.strftime("%m/%d"), "input_tokens": d.get("input", 0), "output_tokens": d.get("output", 0), "total_tokens": d.get("tokens", 0), "cost": round(d.get("cost", 0), 2), "calls": d.get("calls", 0)})
            current += timedelta(days=1)
        return {"days": result}
    except Exception as e:
        print(f"[ADMIN] token-daily error: {e}")
        return {"days": []}

@app.get("/{token}/api/admin/dashboard/token-monthly")
async def admin_token_monthly(token: str, year: int = 0):
    """월별 토큰 사용량"""
    await _check_admin(token)
    if not MONGO_OK or token_usage_col is None:
        return {"year": year, "months": []}
    try:
        if year == 0:
            year = datetime.now(timezone.utc).year
        months = []
        for m in range(1, 13):
            s = datetime(year, m, 1, tzinfo=timezone.utc)
            e = datetime(year, m + 1, 1, tzinfo=timezone.utc) if m < 12 else datetime(year + 1, 1, 1, tzinfo=timezone.utc)
            agg = await token_usage_col.aggregate([
                {"$match": {"created_at": {"$gte": s, "$lt": e}}},
                {"$group": {"_id": None, "input": {"$sum": "$input_tokens"}, "output": {"$sum": "$output_tokens"}, "tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}
            ]).to_list(1)
            d = agg[0] if agg else {}
            months.append({"month": m, "input_tokens": d.get("input", 0), "output_tokens": d.get("output", 0), "total_tokens": d.get("tokens", 0), "cost": round(d.get("cost", 0), 2), "calls": d.get("calls", 0)})
        return {"year": year, "months": months}
    except Exception as e:
        print(f"[ADMIN] token-monthly error: {e}")
        return {"year": year, "months": []}

@app.get("/{token}/api/admin/dashboard/token-by-model")
async def admin_token_by_model(token: str, start_date: str = "", end_date: str = ""):
    """모델별 토큰 사용 비율"""
    await _check_admin(token)
    if not MONGO_OK or token_usage_col is None:
        return {"models": []}
    try:
        range_start, range_end = _parse_date_range(start_date, end_date)
        match_filter = {}
        if range_start:
            match_filter["created_at"] = {"$gte": range_start}
            if range_end:
                match_filter["created_at"]["$lte"] = range_end
        pipeline = [{"$group": {"_id": "$model", "input_tokens": {"$sum": "$input_tokens"}, "output_tokens": {"$sum": "$output_tokens"}, "total_tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}]
        if match_filter:
            pipeline.insert(0, {"$match": match_filter})
        agg = await token_usage_col.aggregate(pipeline).to_list(100)
        models = []
        for doc in agg:
            model_name = doc["_id"] or "unknown"
            label = "Opus" if "opus" in model_name.lower() else ("Sonnet" if "sonnet" in model_name.lower() else model_name)
            models.append({"model": model_name, "label": label, "input_tokens": doc["input_tokens"], "output_tokens": doc["output_tokens"], "total_tokens": doc["total_tokens"], "cost": round(doc["cost"], 2), "calls": doc["calls"]})
        models.sort(key=lambda x: x["total_tokens"], reverse=True)
        return {"models": models}
    except Exception as e:
        print(f"[ADMIN] token-by-model error: {e}")
        return {"models": []}

@app.get("/{token}/api/admin/dashboard/token-by-user")
async def admin_token_by_user(token: str, period: str = "month", start_date: str = "", end_date: str = "", limit: int = 50):
    """사용자별 토큰 사용량 랭킹"""
    await _check_admin(token)
    if not MONGO_OK or token_usage_col is None:
        return {"users": []}
    try:
        range_start, range_end = _parse_date_range(start_date, end_date)
        if not range_start:
            now = datetime.now(timezone.utc)
            if period == "today":
                range_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            elif period == "week":
                range_start = (now - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
            elif period == "year":
                range_start = datetime(now.year, 1, 1, tzinfo=timezone.utc)
            else:
                range_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
            range_end = now
        match_filter = {"created_at": {"$gte": range_start}}
        if range_end:
            match_filter["created_at"]["$lte"] = range_end
        # 사용자+모델별 집계
        pipeline = [
            {"$match": match_filter},
            {"$group": {"_id": {"username": "$username", "model": "$model"}, "input_tokens": {"$sum": "$input_tokens"}, "output_tokens": {"$sum": "$output_tokens"}, "total_tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}
        ]
        agg = await token_usage_col.aggregate(pipeline).to_list(10000)
        # 사용자별로 재그룹핑
        user_map = {}
        for doc in agg:
            uname = doc["_id"]["username"]
            model = doc["_id"]["model"] or ""
            if uname not in user_map:
                user_map[uname] = {"username": uname, "input_tokens": 0, "output_tokens": 0, "total_tokens": 0, "cost": 0, "calls": 0, "opus_calls": 0, "sonnet_calls": 0, "opus_tokens": 0, "sonnet_tokens": 0}
            u = user_map[uname]
            u["input_tokens"] += doc["input_tokens"]
            u["output_tokens"] += doc["output_tokens"]
            u["total_tokens"] += doc["total_tokens"]
            u["cost"] += doc["cost"]
            u["calls"] += doc["calls"]
            if "opus" in model.lower():
                u["opus_calls"] += doc["calls"]; u["opus_tokens"] += doc["total_tokens"]
            else:
                u["sonnet_calls"] += doc["calls"]; u["sonnet_tokens"] += doc["total_tokens"]
        users = sorted(user_map.values(), key=lambda x: x["total_tokens"], reverse=True)[:limit]
        # 표시 이름 조회
        if org_user_collection is not None:
            for u in users:
                org = await org_user_collection.find_one({"lid": u["username"]})
                u["display_name"] = org.get("name", u["username"]) if org else u["username"]
                u["cost"] = round(u["cost"], 2)
        else:
            for u in users:
                u["display_name"] = u["username"]
                u["cost"] = round(u["cost"], 2)
        return {"users": users}
    except Exception as e:
        print(f"[ADMIN] token-by-user error: {e}")
        import traceback; traceback.print_exc()
        return {"users": []}

@app.get("/{token}/api/admin/dashboard/token-by-service")
async def admin_token_by_service(token: str, start_date: str = "", end_date: str = ""):
    """서비스 유형별 토큰 사용 비율"""
    await _check_admin(token)
    if not MONGO_OK or token_usage_col is None:
        return {"services": []}
    try:
        range_start, range_end = _parse_date_range(start_date, end_date)
        match_filter = {}
        if range_start:
            match_filter["created_at"] = {"$gte": range_start}
            if range_end:
                match_filter["created_at"]["$lte"] = range_end
        pipeline = [{"$group": {"_id": "$service_type", "total_tokens": {"$sum": "$total_tokens"}, "cost": {"$sum": "$cost_estimate"}, "calls": {"$sum": 1}}}]
        if match_filter:
            pipeline.insert(0, {"$match": match_filter})
        agg = await token_usage_col.aggregate(pipeline).to_list(100)
        label_map = {"chat": "AI 대화", "rest_task": "스케줄 작업", "compress": "컨텍스트 압축"}
        services = []
        for doc in agg:
            stype = doc["_id"] or "unknown"
            services.append({"service_type": stype, "label": label_map.get(stype, stype), "total_tokens": doc["total_tokens"], "cost": round(doc["cost"], 2), "calls": doc["calls"]})
        services.sort(key=lambda x: x["total_tokens"], reverse=True)
        return {"services": services}
    except Exception as e:
        print(f"[ADMIN] token-by-service error: {e}")
        return {"services": []}

@app.get("/{token}/api/admin/dashboard")
async def admin_dashboard_page(token: str):
    """관리자 대시보드 페이지"""
    await _check_admin(token)
    try:
        async with aiofiles.open("static/admin_dashboard.html", "r", encoding="utf-8") as f:
            html = await f.read()
    except FileNotFoundError:
        raise HTTPException(404, "Dashboard page not found")
    html = html.replace("{{BASE_URL}}", f"/{token}")
    html = html.replace("{{APP_TITLE}}", APP_TITLE)
    html = html.replace("{{APP_BRAND_ICON}}", APP_BRAND_ICON)
    return HTMLResponse(html)

@app.get("/")
async def index():
    return HTMLResponse("""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><title>K-Portal Cowork</title>
<style>*{margin:0;padding:0;box-sizing:border-box}body{font-family:'Noto Sans KR',sans-serif;background:#F5F7FA;height:100vh;display:flex;align-items:center;justify-content:center}
.box{background:#fff;border-radius:16px;padding:48px;text-align:center;box-shadow:0 8px 32px rgba(0,0,0,.08);max-width:460px}
.icon{width:64px;height:64px;background:linear-gradient(135deg,#4A7CFF,#06B6D4);border-radius:16px;display:inline-grid;place-items:center;color:#fff;font-size:28px;font-weight:700;margin-bottom:20px}
h2{font-size:20px;color:#1A1A2E;margin-bottom:8px}p{color:#6B7280;font-size:14px;line-height:1.6}
.sub{margin-top:16px;padding:12px 16px;background:#F5F7FA;border-radius:8px;font-size:12px;color:#9CA3AF}</style></head>
<body><div class="box"><div class="icon">K</div><h2>K-Portal Cowork</h2>
<p>이 서비스는 K-Portal 통합인증을 통해 접속할 수 있습니다.<br>K-Portal에서 Cowork 메뉴를 이용해주세요.</p>
<div class="sub">🔒 직접 URL 접속은 허용되지 않습니다</div></div></body></html>""", status_code=403)

# favicon
@app.get("/favicon.ico")
async def favicon():
    return FileResponse("static/img/favicon.ico")

# ============================================================
# 테스트 로그인 (개발/디버깅용)
# ============================================================
@app.get("/api/auth/testuser")
async def test_login():
    """테스트용 JWT를 생성하여 메인 페이지로 리다이렉트"""
    payload = {
        "sub": "auth",
        "depths": "000000^000000^000006",
        "exp": 9999999999,  # 2286년까지 유효 (사실상 무한)
        "userid": "ygkim@kmslab.com",
        "email": "KM0035"
    }
    token = pyjwt.encode(payload, KPORTAL_JWT_SECRET.encode("utf-8"), algorithm="HS256")
    from starlette.responses import RedirectResponse
    return RedirectResponse(url=f"/{token}/ko", status_code=302)

# JWT 토큰으로 접속: /{jwt_token} 또는 /{jwt_token}/{lang}
SUPPORTED_LANGS = {'ko', 'en', 'ja', 'zh'}

@app.get("/{token}")
async def user_index_jwt(token: str):
    if token in ("api","static","ws","favicon.ico"): raise HTTPException(404)
    if token.count(".") == 2:
        username = userid_from_jwt(token)
        if not username:
            # JWT 만료 vs 무효 구분
            expired = is_jwt_expired(token)
            _email = email_from_jwt(token)
            print(f"[ACCESS DENIED] {_email or 'unknown'} - {'expired' if expired else 'invalid'}")
            if expired:
                return HTMLResponse(_jwt_expired_page(), status_code=401)
            return HTMLResponse(_jwt_invalid_page(), status_code=401)
        get_user_workspace(username)
        _email = email_from_jwt(token)
        print(f"[ACCESS] {username} ({_email})")
        return _serve_index()
    if USERNAME_PATTERN.match(token):
        get_user_workspace(token)
        return _serve_index()
    raise HTTPException(400, "유효하지 않은 접근")

@app.get("/{token}/{lang}")
async def user_index_lang(token: str, lang: str):
    """언어 지정 접속: /{token}/en, /{token}/ja, /{token}/zh, /{token}/ko"""
    if lang not in SUPPORTED_LANGS:
        raise HTTPException(404)
    if token.count(".") == 2:
        username = userid_from_jwt(token)
        if not username:
            expired = is_jwt_expired(token)
            if expired:
                return HTMLResponse(_jwt_expired_page(), status_code=401)
            return HTMLResponse(_jwt_invalid_page(), status_code=401)
        get_user_workspace(username)
    elif USERNAME_PATTERN.match(token):
        get_user_workspace(token)
    else:
        raise HTTPException(400, "유효하지 않은 접근")
    return _serve_index_lang(lang)

# JWT 정보 확인 API
@app.get("/{token}/api/auth-info")
async def auth_info(token: str):
    if token.count(".") == 2:
        username = userid_from_jwt(token)
        if not username:
            expired = is_jwt_expired(token)
            return {"authenticated": False, "expired": expired, "reason": "만료됨" if expired else "유효하지 않은 토큰"}
        payload = decode_kportal_jwt(token)
        if payload and username:
            # 조직도에서 role 조회 (여러 필드명 대응)
            role = ""
            if MONGO_OK and org_user_collection is not None:
                try:
                    org_doc = await org_user_collection.find_one({"lid": username})
                    if org_doc:
                        # role 필드 여러 후보에서 탐색
                        role = (org_doc.get("role") or org_doc.get("auth") or
                                org_doc.get("grade") or org_doc.get("level") or
                                org_doc.get("type") or "")
                        print(f"[AUTH] user={username} role={role} org_fields={[k for k in org_doc.keys() if k != '_id']}")
                    else:
                        print(f"[AUTH] user={username} not found in org DB")
                except Exception as e:
                    print(f"[AUTH] org lookup error: {e}")
            return {"username": username, "userid": payload.get("userid",""), "email": payload.get("email",""), "authenticated": True,
                    "role": role if role else ("admin" if username in ADMIN_USERS else "")}
    elif USERNAME_PATTERN.match(token):
        return {"username": token, "userid": token, "email": "", "authenticated": True, "role": "admin"}
    return {"authenticated": False}


# 사용자 설정 API (Figma 토큰 등)
@app.get("/{token}/api/settings")
async def get_user_settings(token: str):
    u, _ = _resolve_user(token)
    if not MONGO_OK or user_settings_collection is None:
        return {"figma_token": "", "has_figma_token": False}
    doc = await user_settings_collection.find_one({"username": u})
    if doc and doc.get("figma_token"):
        # 토큰은 앞 8자만 보여주고 나머지는 마스킹
        t = doc["figma_token"]
        masked = t[:8] + "*" * (len(t) - 8) if len(t) > 8 else "****"
        return {"figma_token_masked": masked, "has_figma_token": True}
    return {"figma_token_masked": "", "has_figma_token": False}

@app.post("/{token}/api/settings")
async def save_user_settings(token: str, payload: dict):
    u, _ = _resolve_user(token)
    if not MONGO_OK or user_settings_collection is None:
        raise HTTPException(500, "MongoDB 연결 필요")
    update_fields = {}
    if "figma_token" in payload:
        ft = payload["figma_token"].strip()
        if ft:
            # 토큰 유효성 검증
            async with httpx.AsyncClient(timeout=10) as c:
                r = await c.get("https://api.figma.com/v1/me", headers={"X-Figma-Token": ft})
                if r.status_code != 200:
                    raise HTTPException(400, "Figma 토큰이 유효하지 않습니다. Personal Access Token을 확인하세요.")
                figma_user = r.json()
            update_fields["figma_token"] = ft
            update_fields["figma_user"] = figma_user.get("email", figma_user.get("handle", ""))
        else:
            update_fields["figma_token"] = ""
            update_fields["figma_user"] = ""
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc)
        await user_settings_collection.update_one(
            {"username": u}, {"$set": update_fields}, upsert=True
        )
    return {"success": True, "message": "설정이 저장되었습니다."}

@app.delete("/{token}/api/settings/figma")
async def delete_figma_token(token: str):
    u, _ = _resolve_user(token)
    if not MONGO_OK or user_settings_collection is None:
        raise HTTPException(500, "MongoDB 연결 필요")
    await user_settings_collection.update_one(
        {"username": u}, {"$set": {"figma_token": "", "figma_user": "", "updated_at": datetime.now(timezone.utc)}}
    )
    return {"success": True}

# ============================================================
# Skills API
# ============================================================
@app.get("/{token}/api/skills")
async def get_skills(token: str):
    """내 스킬 + 공유받은 스킬 목록"""
    u, ws = _resolve_user(token)
    if not MONGO_OK or skills_collection is None:
        return {"my_skills": [], "shared_skills": []}
    # 내 스킬
    my_cursor = skills_collection.find({"owner": u}, {"md_contents": 0}).sort("name", 1)
    my_skills = []
    async for doc in my_cursor:
        doc["_id"] = str(doc["_id"])
        my_skills.append(doc)
    # 공유받은 스킬
    shared_cursor = skills_collection.find({"shared_with": u}, {"md_contents": 0}).sort("name", 1)
    shared_skills = []
    async for doc in shared_cursor:
        doc["_id"] = str(doc["_id"])
        shared_skills.append(doc)
    return {"my_skills": my_skills, "shared_skills": shared_skills}

@app.post("/{token}/api/skills")
async def register_skill(token: str, payload: dict):
    """스킬 등록 - 워크스페이스 내 폴더의 .md 파일 내용을 DB에 저장 (공유 폴더 지원)"""
    u, ws = _resolve_user(token)
    if not MONGO_OK or skills_collection is None:
        raise HTTPException(500, "MongoDB 연결 필요")
    name = payload.get("name", "").strip()
    folder = payload.get("folder", "").strip()
    description = payload.get("description", "").strip()
    skill_owner = payload.get("owner", "").strip()  # 공유 폴더의 소유자
    if not name or not folder:
        raise HTTPException(400, "스킬 이름과 폴더 경로를 입력하세요")
    # 공유 폴더인 경우 소유자 워크스페이스 사용
    target_ws = ws
    if skill_owner and skill_owner != u:
        if MONGO_OK and shared_folders_collection is not None:
            share = await shared_folders_collection.find_one({"owner": skill_owner, "shared_with": u})
            if share:
                target_ws = get_user_workspace(skill_owner)
            else:
                raise HTTPException(403, "공유 권한이 없습니다")
        else:
            raise HTTPException(400, "공유 폴더 접근 불가")
    # 폴더 존재 확인 + 하위 포함 .md 파일 재귀 읽기
    folder_path = safe_path(folder, target_ws)
    if not folder_path.is_dir():
        raise HTTPException(400, f"폴더가 존재하지 않습니다: {folder}")
    md_files_data = []
    for root, dirs, files in os.walk(str(folder_path)):
        dirs.sort(key=str.lower)
        for fname in sorted(files, key=str.lower):
            if fname.endswith(".md"):
                fpath = Path(root) / fname
                try:
                    content = fpath.read_text(encoding="utf-8")
                    # 상대 경로로 저장 (폴더 기준)
                    rel = str(fpath.relative_to(folder_path)).replace("\\", "/")
                    md_files_data.append({"name": rel, "content": content})
                except:
                    pass
    if not md_files_data:
        raise HTTPException(400, f"폴더에 .md 파일이 없습니다: {folder}")
    doc = {
        "owner": u,
        "name": name,
        "description": description,
        "folder": folder,
        "folder_owner": skill_owner if skill_owner and skill_owner != u else "",
        "md_files": [m["name"] for m in md_files_data],
        "md_contents": md_files_data,
        "active": True,
        "shared_with": [],
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    try:
        result = await skills_collection.insert_one(doc)
        return {"success": True, "id": str(result.inserted_id), "md_files": doc["md_files"]}
    except Exception as e:
        if "duplicate" in str(e).lower():
            raise HTTPException(400, f"같은 이름의 스킬이 이미 존재합니다: {name}")
        raise HTTPException(500, str(e))

@app.put("/{token}/api/skills/{skill_id}")
async def update_skill(token: str, skill_id: str, payload: dict):
    """스킬 수정 (이름, 설명, 활성화 상태, 폴더 재스캔 → DB 내용 갱신)"""
    u, ws = _resolve_user(token)
    if not MONGO_OK or skills_collection is None:
        raise HTTPException(500, "MongoDB 연결 필요")
    from bson import ObjectId
    doc = await skills_collection.find_one({"_id": ObjectId(skill_id)})
    if not doc:
        raise HTTPException(404, "스킬을 찾을 수 없습니다")
    if doc["owner"] != u:
        raise HTTPException(403, "본인의 스킬만 수정할 수 있습니다")
    update = {"updated_at": datetime.now(timezone.utc)}
    if "name" in payload: update["name"] = payload["name"].strip()
    if "description" in payload: update["description"] = payload["description"].strip()
    if "active" in payload: update["active"] = bool(payload["active"])

    def _scan_folder(folder_path):
        """폴더 + 하위 폴더에서 .md 파일 재귀적으로 읽어서 이름+내용 반환"""
        md_data = []
        if folder_path.is_dir():
            for root, dirs, files in os.walk(str(folder_path)):
                dirs.sort(key=str.lower)
                for fname in sorted(files, key=str.lower):
                    if fname.endswith(".md"):
                        fpath = Path(root) / fname
                        try:
                            rel = str(fpath.relative_to(folder_path)).replace("\\", "/")
                            md_data.append({"name": rel, "content": fpath.read_text(encoding="utf-8")})
                        except:
                            pass
        return md_data

    if "folder" in payload:
        folder = payload["folder"].strip()
        folder_path = safe_path(folder, ws)
        md_data = _scan_folder(folder_path)
        update["folder"] = folder
        update["md_files"] = [m["name"] for m in md_data]
        update["md_contents"] = md_data

    if payload.get("rescan"):
        # folder_owner가 있으면 해당 사용자의 워크스페이스에서 재스캔
        rescan_ws = ws
        fo = doc.get("folder_owner", "")
        if fo and fo != u:
            rescan_ws = get_user_workspace(fo)
        folder_path = safe_path(doc["folder"], rescan_ws)
        if folder_path.is_dir():
            md_data = _scan_folder(folder_path)
            update["md_files"] = [m["name"] for m in md_data]
            update["md_contents"] = md_data
        # 폴더가 없으면 기존 DB 내용 유지 (경고만)

    await skills_collection.update_one({"_id": ObjectId(skill_id)}, {"$set": update})
    return {"success": True, "md_count": len(update.get("md_files", doc.get("md_files", [])))}

@app.delete("/{token}/api/skills/{skill_id}")
async def delete_skill(token: str, skill_id: str):
    """스킬 삭제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or skills_collection is None:
        raise HTTPException(500, "MongoDB 연결 필요")
    from bson import ObjectId
    doc = await skills_collection.find_one({"_id": ObjectId(skill_id)})
    if not doc or doc["owner"] != u:
        raise HTTPException(403, "본인의 스킬만 삭제할 수 있습니다")
    await skills_collection.delete_one({"_id": ObjectId(skill_id)})
    return {"success": True}

@app.post("/{token}/api/skills/{skill_id}/share")
async def share_skill(token: str, skill_id: str, payload: dict):
    """스킬 공유/해제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or skills_collection is None:
        raise HTTPException(500, "MongoDB 연결 필요")
    from bson import ObjectId
    doc = await skills_collection.find_one({"_id": ObjectId(skill_id)})
    if not doc or doc["owner"] != u:
        raise HTTPException(403, "본인의 스킬만 공유할 수 있습니다")
    users = payload.get("users", [])
    action = payload.get("action", "add")  # add or remove
    if action == "add":
        await skills_collection.update_one(
            {"_id": ObjectId(skill_id)},
            {"$addToSet": {"shared_with": {"$each": users}}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    elif action == "remove":
        await skills_collection.update_one(
            {"_id": ObjectId(skill_id)},
            {"$pull": {"shared_with": {"$in": users}}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    return {"success": True}

@app.post("/{token}/api/skills/{skill_id}/toggle")
async def toggle_shared_skill(token: str, skill_id: str, payload: dict):
    """공유받은 스킬의 활성화/비활성화 (개인 설정에 저장)"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or user_settings_collection is None:
        raise HTTPException(500, "MongoDB 연결 필요")
    active = payload.get("active", True)
    # user_settings에 disabled_shared_skills 배열로 관리
    if not active:
        await user_settings_collection.update_one(
            {"username": u},
            {"$addToSet": {"disabled_shared_skills": skill_id}, "$set": {"updated_at": datetime.now(timezone.utc)}},
            upsert=True
        )
    else:
        await user_settings_collection.update_one(
            {"username": u},
            {"$pull": {"disabled_shared_skills": skill_id}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    return {"success": True}

@app.get("/{token}/api/skills/active")
async def get_active_skills_content(token: str):
    """현재 활성화된 모든 스킬의 .md 내용을 반환 (DB 기반)"""
    u, ws = _resolve_user(token)
    if not MONGO_OK or skills_collection is None:
        return {"skills": []}
    disabled = set()
    if user_settings_collection is not None:
        settings = await user_settings_collection.find_one({"username": u})
        if settings:
            disabled = set(settings.get("disabled_shared_skills", []))
    skills = []
    async for doc in skills_collection.find({"owner": u, "active": True}):
        skills.append(doc)
    async for doc in skills_collection.find({"shared_with": u, "active": True}):
        if str(doc["_id"]) not in disabled:
            skills.append(doc)
    result = []
    for skill in skills:
        md_contents = skill.get("md_contents", [])
        if md_contents:
            contents = [{"file": m["name"], "content": m["content"][:10000]} for m in md_contents]
            result.append({"name": skill["name"], "description": skill.get("description", ""), "owner": skill["owner"], "files": contents})
    return {"skills": result}

def _strip_korean_particles(word):
    """한국어 조사/어미를 제거하여 어근 추출"""
    particles = ['입니다','습니다','을','를','이','가','은','는','에','의','로','으로','에서','까지','부터',
                 '와','과','나','도','만','든','들','님','하여','해서','하고','인','적','해','임',
                 '이나','이란','에게','한테','께서']
    w = word
    for p in sorted(particles, key=len, reverse=True):
        if len(w) > len(p) + 1 and w.endswith(p):
            w = w[:-len(p)]
            break
    return w

def _match_skill_score(skill: dict, user_message: str) -> int:
    """사용자 메시지와 스킬의 관련도 점수 계산 (0~100)"""
    msg = user_message.lower()
    # 메시지 키워드 추출 (조사 제거)
    msg_words = set()
    for w in re.split(r'[\s/\-_·,\.!?]+', msg):
        if len(w) >= 2:
            msg_words.add(w)
            msg_words.add(_strip_korean_particles(w))
    score = 0
    # 스킬 이름 전체 매칭 (가장 강력)
    skill_name = skill.get("name", "").lower()
    if skill_name and skill_name in msg:
        score += 50
    # 스킬 이름의 각 단어 매칭 (3글자 이상만)
    name_words = [_strip_korean_particles(w) for w in re.split(r'[\s/\-_·,]+', skill_name) if len(w) >= 3]
    for w in name_words:
        if w in msg_words:
            score += 15
    # description 키워드 매칭 (3글자 이상만, 조사 제거)
    desc = skill.get("description", "").lower()
    raw_desc_words = [_strip_korean_particles(w) for w in re.split(r'[\s/\-_·,\.]+', desc) if len(w) >= 3]
    # 불용어 제거 (일반적으로 자주 쓰이는 단어)
    stopwords = {'있습니다','만들어','하는','위한','대한','등록','사용','활용','작성','이런','저런','그런',
                 '만들때','합니다','입니다','스킬','방법','정의한','순서',
                 '통해','결과','데이터','주어진','통한','작업','분석','보고','보고서',
                 '만들','검색','인터넷','다양한','협업','때문',
                 'the','and','for','with','this','that','from','are','was','will','can'}
    desc_words = [w for w in raw_desc_words if w not in stopwords]
    matched_desc = sum(1 for w in desc_words if w in msg_words)
    if desc_words:
        score += min(30, int(matched_desc / max(len(desc_words), 1) * 60))
    # md 파일명 매칭 (3글자 이상만)
    for md_name in skill.get("md_files", []):
        fname = md_name.lower().replace(".md", "").replace("/", " ").replace("-", " ").replace("_", " ")
        for w in fname.split():
            if len(w) >= 3 and w in msg_words:
                score += 8
    return min(score, 100)

async def _get_active_skills_prompt(username: str, user_message: str = "", forced_skill_name: str = "") -> tuple:
    """사용자 메시지와 관련된 스킬만 선택하여 system prompt 블록으로 조합. (prompt, skill_names) 반환
    forced_skill_name이 지정되면 해당 스킬을 강제로 포함합니다."""
    if not MONGO_OK or skills_collection is None:
        return "", []
    try:
        disabled = set()
        if user_settings_collection is not None:
            settings = await user_settings_collection.find_one({"username": username})
            if settings:
                disabled = set(settings.get("disabled_shared_skills", []))
        all_skills = []
        async for doc in skills_collection.find({"owner": username, "active": True}):
            all_skills.append(doc)
        async for doc in skills_collection.find({"shared_with": username, "active": True}):
            if str(doc["_id"]) not in disabled:
                all_skills.append(doc)
        if not all_skills:
            return "", []
        # forced_skill_name이 지정되면 해당 스킬을 강제 선택
        if forced_skill_name:
            forced = [s for s in all_skills if s.get("name", "").lower() == forced_skill_name.lower()]
            if not forced:
                forced = [s for s in all_skills if forced_skill_name.lower() in s.get("name", "").lower()]
            if forced:
                selected_skills = forced
            else:
                selected_skills = all_skills[:3]
        # user_message가 없으면 전부 사용 (초기 로딩 등)
        elif not user_message.strip():
            selected_skills = all_skills
        else:
            # 항상 관련성 점수로 필터링 (스킬 수와 무관)
            scored = []
            for skill in all_skills:
                score = _match_skill_score(skill, user_message)
                scored.append((score, skill))
            scored.sort(key=lambda x: -x[0])
            selected_skills = [s for score, s in scored if score >= 15][:5]
            if not selected_skills:
                # 매칭되는 스킬이 없으면 카탈로그만 제공
                catalog = "\n".join([f"- {s.get('name','')}: {s.get('description','')}" for s in all_skills])
                prompt = f"\n\n<available_skills_catalog>\n사용자가 등록한 스킬 목록입니다. 아래 스킬이 작업에 필요하다고 판단되면 사용자에게 해당 스킬 활용을 제안하세요.\n{catalog}\n</available_skills_catalog>"
                all_names = [{"name": s["name"], "owner": s["owner"], "shared": s["owner"] != username, "files": len(s.get("md_contents", [])), "matched": False} for s in all_skills]
                return prompt, all_names
        blocks = []
        skill_names = []
        for skill in selected_skills:
            md_contents = skill.get("md_contents", [])
            if not md_contents:
                fo = skill.get("folder_owner", "") or skill["owner"]
                owner_ws = get_user_workspace(fo)
                folder_path = Path(owner_ws) / skill["folder"]
                if folder_path.is_dir():
                    for md_name in skill.get("md_files", []):
                        md_path = folder_path / md_name
                        if md_path.is_file():
                            try:
                                md_contents.append({"name": md_name, "content": md_path.read_text(encoding="utf-8")})
                            except:
                                pass
            if not md_contents:
                continue
            parts = [f"<skill name=\"{skill['name']}\" description=\"{skill.get('description','')}\">"]
            for md in md_contents:
                text = md.get("content", "")
                if len(text) > 10000:
                    text = text[:10000] + "\n...(truncated)"
                parts.append(f"### {md['name']}\n{text}")
            parts.append("</skill>")
            blocks.append("\n".join(parts))
            is_shared = skill["owner"] != username
            skill_names.append({"name": skill["name"], "owner": skill["owner"], "shared": is_shared, "files": len(md_contents), "matched": True})
        if not blocks:
            return "", []
        prompt = "\n\n<available_skills>\n아래는 사용자의 요청과 관련된 스킬(지침서)입니다. 이 스킬의 내용을 참고하여 최적의 결과를 만들어주세요.\n\n" + "\n\n".join(blocks) + "\n</available_skills>"
        return prompt, skill_names
    except Exception as e:
        print(f"[SKILLS ERROR] {type(e).__name__}: {e}")
        traceback.print_exc()
        return "", []

def _resolve_user(token: str):
    """토큰에서 username과 workspace 반환 (JWT 만료 시 401)"""
    if token.count(".") == 2:
        u = userid_from_jwt(token)
        if not u:
            expired = is_jwt_expired(token)
            if expired:
                raise HTTPException(401, "세션이 만료되었습니다. K-Portal에서 다시 접속해주세요.")
            raise HTTPException(401, "유효하지 않은 토큰입니다.")
        return u, get_user_workspace(u)
    if USERNAME_PATTERN.match(token):
        return token, get_user_workspace(token)
    raise HTTPException(400, "유효하지 않은 접근")

async def _resolve_workspace(token: str, owner: str = None):
    """파일 API용: owner가 지정되면 공유 권한 확인 후 소유자 워크스페이스 반환"""
    user, my_ws = _resolve_user(token)
    if not owner or owner == user:
        return user, my_ws, None  # (user, workspace, share_perm)
    # 공유 권한 확인
    if not MONGO_OK or shared_folders_collection is None:
        raise HTTPException(403, "공유 기능 사용 불가")
    share = await shared_folders_collection.find_one({"owner": owner, "shared_with": user})
    if not share:
        raise HTTPException(403, "공유받지 않은 사용자입니다")
    return user, get_user_workspace(owner), share.get("permission", "read")

@app.get("/{token}/api/workspace")
async def u_workspace(token: str):
    u, ws = _resolve_user(token)
    return {"workspace": ws, "exists": os.path.exists(ws), "username": u}

@app.get("/{token}/api/files")
async def u_files(token: str, path: str = ".", owner: str = None):
    _, ws, _ = await _resolve_workspace(token, owner)
    return JSONResponse(json.loads(await execute_tool("list_files", {"path":path}, ws)))

@app.get("/{token}/api/folders")
async def u_folders(token: str, path: str = ".", owner: str = None):
    """폴더 목록만 반환 (스킬 등록용 폴더 브라우저). owner 지정 시 공유 폴더 탐색"""
    u, ws = _resolve_user(token)
    if owner and owner != u:
        # 공유 폴더: 권한 확인
        if MONGO_OK and shared_folders_collection is not None:
            share = await shared_folders_collection.find_one({"owner": owner, "shared_with": u})
            if share:
                ws = get_user_workspace(owner)
            else:
                return {"folders": [], "path": path, "md_count": 0, "error": "공유 권한 없음"}
        else:
            return {"folders": [], "path": path, "md_count": 0}
    t = safe_path(path, ws)
    if not t.is_dir():
        return {"folders": [], "path": path, "md_count": 0, "total_size": 0}
    folders = []
    # 현재 폴더 직하위 폴더 목록
    for entry in sorted(t.iterdir(), key=lambda x: x.name.lower()):
        if entry.name.startswith(".") or entry.name.startswith("_"):
            continue
        if entry.is_dir():
            folders.append(entry.name)
    # 하위 전체 재귀: md 카운트 + 전체 사이즈
    md_count = 0
    total_size = 0
    for root, dirs, files in os.walk(str(t)):
        dirs[:] = [d for d in dirs if not d.startswith(".") and not d.startswith("_")]
        for fname in files:
            fpath = Path(root) / fname
            try:
                total_size += fpath.stat().st_size
            except:
                pass
            if fname.endswith(".md"):
                md_count += 1
    return {"folders": folders, "path": path, "md_count": md_count, "total_size": total_size}

@app.get("/{token}/api/file")
async def u_file(token: str, path: str, owner: str = None):
    _, ws, _ = await _resolve_workspace(token, owner)
    return JSONResponse(json.loads(await execute_tool("read_file", {"path":path}, ws)))

@app.put("/{token}/api/file")
async def u_save_file(token: str, request: Request, owner: str = None):
    """파일 내용 저장 (웹 에디터용)"""
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read":
        raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    body = await request.json()
    file_path = body.get("path", "")
    content = body.get("content", "")
    if not file_path:
        raise HTTPException(400, "path 필수")
    p = safe_path(file_path, ws)
    if not await aio_exists(p):
        raise HTTPException(404, "파일 없음")
    size = await aio_write_text(p, content, body.get("encoding", "utf-8"))
    return {"success": True, "path": file_path, "size": size}

@app.post("/{token}/api/upload")
async def u_upload(token: str, file: UploadFile = File(...), path: str = Form("."), owner: str = Form(None)):
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read":
        raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    fname = file.filename or "file"
    # 중복 파일명 처리: name(1).ext, name(2).ext ...
    target_dir = safe_path(path, ws)
    stem = Path(fname).stem
    ext = Path(fname).suffix
    final_name = fname
    counter = 1
    while (target_dir / final_name).exists():
        final_name = f"{stem}({counter}){ext}"
        counter += 1
    t = safe_path(os.path.join(path, final_name), ws); await aio_mkdir(t.parent)
    c = await file.read(); await aio_write_bytes(t, c)
    return {"success":True,"path":str(t.relative_to(Path(ws).resolve())),"size":len(c),"name":final_name}

@app.post("/{token}/api/upload-attach")
async def u_upload_attach(token: str, file: UploadFile = File(...), owner: str = None):
    """대화창 첨부 전용 업로드 - 항상 사용자 _temp 폴더에 저장"""
    u, _ = _resolve_user(token)
    ws = get_user_workspace(u)  # 항상 본인 워크스페이스의 _temp
    temp_dir = Path(ws) / "_temp"
    os.makedirs(str(temp_dir), exist_ok=True)
    # 타임스탬프로 중복 방지
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = file.filename or "file"
    stem, ext = os.path.splitext(name)
    safe_name = f"{stem}_{ts}{ext}"
    t = temp_dir / safe_name
    c = await file.read()
    await aio_write_bytes(t, c)
    rel_path = f"_temp/{safe_name}"
    return {"success": True, "path": rel_path, "name": safe_name, "size": len(c)}

@app.post("/{token}/api/upload-folder")
async def u_upload_folder(token: str, request: Request, files: List[UploadFile] = File(...), basePath: str = Form("."), owner: str = Form(None)):
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read": raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    results = []
    for f in files:
        try:
            rel = f.filename; full_rel = os.path.join(basePath, rel) if basePath != "." else rel
            t = safe_path(full_rel, ws); await aio_mkdir(t.parent)
            # 중복 파일명 처리
            if t.exists():
                stem = t.stem
                ext = t.suffix
                parent = t.parent
                counter = 1
                while t.exists():
                    t = parent / f"{stem}({counter}){ext}"
                    counter += 1
                full_rel = str(t.relative_to(Path(ws).resolve()))
            async with aiofiles.open(str(t), 'wb') as out:
                while True:
                    chunk = await f.read(1024 * 1024)
                    if not chunk: break
                    await out.write(chunk)
            results.append({"path":full_rel,"success":True})
        except Exception as e: results.append({"path":f.filename,"success":False,"error":str(e)})
    return {"results":results,"count":len([r for r in results if r["success"]])}

@app.get("/{token}/api/download")
async def u_download(token: str, path: str, owner: str = None):
    _, ws, _ = await _resolve_workspace(token, owner); t = safe_path(path, ws)
    if not await aio_exists(t) or not await aio_is_file(t): raise HTTPException(404)
    return FileResponse(str(t), filename=t.name)

@app.post("/{token}/api/temp-link")
async def u_temp_link(token: str, payload: dict):
    """오피스 뷰어용 임시 공개 다운로드 링크 생성 (10분 유효)"""
    if not MONGO_OK or temp_links_collection is None:
        raise HTTPException(500, "DB 미연결")
    owner = payload.get("owner")
    _, ws, _ = await _resolve_workspace(token, owner)
    path = payload.get("path", "").strip()
    if not path: raise HTTPException(400, "path 필요")
    t = safe_path(path, ws)
    if not await aio_exists(t) or not await aio_is_file(t): raise HTTPException(404)
    # 임시 토큰 생성
    tk = hashlib.sha256((str(t) + str(time.time()) + str(uuid.uuid4())).encode()).hexdigest()
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
    await temp_links_collection.insert_one({
        "token": tk,
        "path": str(t),
        "name": t.name,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc),
    })
    return {"token": tk, "url": f"/pub/dl/{tk}/{t.name}"}

@app.get("/pub/dl/{tk}/{filename}")
async def pub_download(tk: str, filename: str):
    """인증 없이 접근 가능한 임시 다운로드 (오피스 뷰어 연동, MongoDB 기반)"""
    if not MONGO_OK or temp_links_collection is None:
        raise HTTPException(500, "DB 미연결")
    entry = await temp_links_collection.find_one({"token": tk})
    if not entry:
        raise HTTPException(410, "링크가 만료되었거나 존재하지 않습니다")
    # MongoDB TTL이 삭제하기 전이라도 만료 시각 체크
    if entry.get("expires_at") and entry["expires_at"].replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
        await temp_links_collection.delete_one({"_id": entry["_id"]})
        raise HTTPException(410, "링크가 만료되었습니다")
    t = Path(entry["path"])
    if not t.exists() or not t.is_file(): raise HTTPException(404)
    mime, _ = mimetypes.guess_type(entry["name"])
    resp = FileResponse(
        str(t),
        filename=entry["name"],
        media_type=mime or "application/octet-stream",
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Expose-Headers": "Content-Disposition, Content-Type, Content-Length",
        }
    )
    return resp

@app.options("/pub/dl/{tk}/{filename}")
async def pub_download_options(tk: str, filename: str):
    """CORS preflight 처리"""
    from starlette.responses import Response as StarletteResponse
    resp = StarletteResponse(status_code=204)
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "*"
    return resp

@app.get("/{token}/api/preview/{file_path:path}")
async def u_preview(token: str, file_path: str, owner: str = None, nav: str = None, nonav: str = None, edit: str = None):
    _, ws, perm = await _resolve_workspace(token, owner); t = safe_path(file_path, ws)
    if not await aio_exists(t) or not await aio_is_file(t): raise HTTPException(404)

    # 편집 모드: 편집 가능한 텍스트 파일이면 CodeMirror 에디터 반환
    EDITABLE_EXT = {'html','htm','css','js','json','md','txt','xml','svg','csv','yml','yaml','sh','py','java','ts','jsx','tsx','sql','log','ini','cfg','conf','env'}
    file_ext = t.name.rsplit('.',1)[-1].lower() if '.' in t.name else ''
    if edit and file_ext in EDITABLE_EXT:
        file_content = await aio_read_text(t)
        # JSON 직렬화 후 </script> 이스케이프 (f-string 안에서 백슬래시 불가하므로 미리 처리)
        file_content_js = json.dumps(file_content).replace("</script>", "<\\/script>")
        save_url = f"/{token}/api/file"
        owner_param = f'"owner":"{owner}",' if owner else ''
        is_readonly = (perm == "read")
        # CodeMirror 언어 매핑
        cm_mode_map = {'html':'htmlmixed','htm':'htmlmixed','css':'css','js':'javascript','json':'javascript',
                       'md':'markdown','txt':'text/plain','xml':'xml','svg':'xml','csv':'text/plain',
                       'yml':'yaml','yaml':'yaml','sh':'shell','py':'python','java':'text/x-java',
                       'ts':'javascript','jsx':'jsx','tsx':'javascript','sql':'sql',
                       'log':'text/plain','ini':'text/plain','cfg':'text/plain','conf':'text/plain','env':'text/plain'}
        cm_mode = cm_mode_map.get(file_ext, 'text/plain')
        editor_html = f'''<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>✏️ {t.name}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/codemirror.min.css">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/theme/dracula.min.css">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/dialog/dialog.min.css">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/search/matchesonscrollbar.min.css">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans KR',sans-serif;height:100vh;display:flex;flex-direction:column;background:#282a36}}
.toolbar{{display:flex;align-items:center;gap:8px;padding:8px 16px;background:#21222c;border-bottom:1px solid #44475a;flex-shrink:0}}
.toolbar .filename{{font-size:14px;font-weight:600;color:#f8f8f2;flex:1;display:flex;align-items:center;gap:8px}}
.toolbar .filename .ext-badge{{padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;text-transform:uppercase}}
.toolbar .filename .modified{{color:#ffb86c;font-size:11px;display:none}}
.toolbar button{{padding:6px 16px;border-radius:6px;border:none;font-size:13px;font-weight:600;cursor:pointer;display:flex;align-items:center;gap:5px;transition:all .15s}}
.btn-save{{background:#50fa7b;color:#282a36}}.btn-save:hover{{background:#69ff94}}
.btn-save:disabled{{background:#44475a;color:#6272a4;cursor:default}}
.btn-preview{{background:#8be9fd;color:#282a36}}.btn-preview:hover{{background:#a4f0ff}}
.btn-close{{background:#44475a;color:#f8f8f2}}.btn-close:hover{{background:#6272a4}}
.btn-undo{{background:#44475a;color:#f8f8f2}}.btn-undo:hover{{background:#6272a4}}
.status-bar{{display:flex;align-items:center;justify-content:space-between;padding:4px 16px;background:#21222c;border-top:1px solid #44475a;font-size:11px;color:#6272a4;flex-shrink:0}}
.status-bar .saved{{color:#50fa7b}}
.status-bar .saving{{color:#ffb86c}}
.status-bar .error{{color:#ff5555}}
.cm-wrap{{flex:1;overflow:hidden}}
.CodeMirror{{height:100%;font-size:14px;line-height:1.6;font-family:'JetBrains Mono','Fira Code','Consolas',monospace}}
.readonly-bar{{background:#ff5555;color:#fff;text-align:center;padding:4px;font-size:12px;font-weight:600}}
@media(max-width:640px){{
.toolbar{{padding:6px 10px;gap:4px}}
.toolbar button{{padding:5px 10px;font-size:12px}}
.toolbar .filename{{font-size:12px}}
.CodeMirror{{font-size:13px}}
}}
</style>
</head><body>
{"<div class='readonly-bar'>🔒 읽기 전용</div>" if is_readonly else ""}
<div class="toolbar">
<div class="filename">
<span class="ext-badge" style="background:{'#ff79c6' if file_ext in ('html','htm') else '#50fa7b' if file_ext in ('css',) else '#f1fa8c' if file_ext in ('js','ts','jsx','tsx') else '#bd93f9' if file_ext in ('py',) else '#8be9fd' if file_ext in ('md',) else '#ffb86c'}">.{file_ext}</span>
<span>{t.name}</span>
<span class="modified" id="modFlag">● 수정됨</span>
</div>
{"" if is_readonly else '<button class="btn-save" id="saveBtn" disabled onclick="saveFile()">💾 저장</button>'}
{"" if is_readonly else '<button class="btn-undo" id="undoBtn" onclick="editor.undo()">↩ 실행취소</button>'}
{f'<button class="btn-preview" onclick="openPreview()">👁 미리보기</button>' if file_ext in ('html','htm','md','svg') else ''}
<button class="btn-close" onclick="window.close()">✕ 닫기</button>
</div>
<div class="cm-wrap" id="editorWrap"></div>
<div class="status-bar">
<span id="statusMsg">준비</span>
<span id="cursorPos">줄 1, 열 1</span>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/codemirror.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/xml/xml.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/javascript/javascript.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/css/css.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/htmlmixed/htmlmixed.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/markdown/markdown.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/python/python.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/yaml/yaml.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/sql/sql.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/mode/shell/shell.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/edit/closebrackets.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/edit/closetag.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/edit/matchbrackets.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/search/search.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/search/searchcursor.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/dialog/dialog.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/fold/foldcode.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/fold/foldgutter.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/fold/brace-fold.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/fold/xml-fold.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/fold/indent-fold.min.js"></script>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/codemirror/5.65.18/addon/fold/foldgutter.min.css">
<script>
var filePath={json.dumps(file_path)};
var saveUrl={json.dumps(save_url)};
var ownerParam={json.dumps(owner or "")};
var isModified=false;
var originalContent={file_content_js};

var editor=CodeMirror(document.getElementById('editorWrap'),{{
value:originalContent,
mode:{json.dumps(cm_mode)},
theme:'dracula',
lineNumbers:true,
lineWrapping:true,
matchBrackets:true,
autoCloseBrackets:true,
autoCloseTags:true,
foldGutter:true,
gutters:["CodeMirror-linenumbers","CodeMirror-foldgutter"],
indentUnit:2,
tabSize:2,
indentWithTabs:false,
readOnly:{json.dumps(is_readonly)},
extraKeys:{{
"Ctrl-S":function(){{saveFile()}},
"Cmd-S":function(){{saveFile()}},
"Ctrl-Z":function(){{editor.undo()}},
"Ctrl-Shift-Z":function(){{editor.redo()}}
}}
}});

editor.on('change',function(){{
var cur=editor.getValue();
var mod=(cur!==originalContent);
if(mod!==isModified){{
isModified=mod;
document.getElementById('modFlag').style.display=mod?'inline':'none';
var btn=document.getElementById('saveBtn');
if(btn)btn.disabled=!mod;
document.title=(mod?'● ':'')+'✏️ {t.name}';
}}
}});

editor.on('cursorActivity',function(){{
var c=editor.getCursor();
document.getElementById('cursorPos').textContent='줄 '+(c.line+1)+', 열 '+(c.ch+1);
}});

function saveFile(){{
if(!isModified)return;
var btn=document.getElementById('saveBtn');
var st=document.getElementById('statusMsg');
if(btn)btn.disabled=true;
st.className='saving';st.textContent='저장 중...';
var body={{path:filePath,content:editor.getValue()}};
if(ownerParam)body.owner=ownerParam;
fetch(saveUrl+(ownerParam?'?owner='+encodeURIComponent(ownerParam):''),{{
method:'PUT',
headers:{{'Content-Type':'application/json'}},
body:JSON.stringify(body)
}}).then(function(r){{return r.json()}}).then(function(d){{
if(d.success){{
originalContent=editor.getValue();
isModified=false;
document.getElementById('modFlag').style.display='none';
document.title='✏️ {t.name}';
st.className='saved';st.textContent='✅ 저장 완료 ('+new Date().toLocaleTimeString()+')';
setTimeout(function(){{st.className='';st.textContent='준비'}},3000);
}}else{{
st.className='error';st.textContent='❌ 저장 실패: '+(d.detail||'오류');
if(btn)btn.disabled=false;
}}
}}).catch(function(e){{
st.className='error';st.textContent='❌ 네트워크 오류';
if(btn)btn.disabled=false;
}});
}}

function openPreview(){{
var ext='{file_ext}';
var previewBase=window.location.pathname.replace(/\\?.*$/,'');
if(previewBase.indexOf('/edit')>-1)previewBase=previewBase.replace(/[?&]edit=1/,'');
// 저장 안된 내용이 있으면 먼저 저장 제안
if(isModified){{
if(confirm('저장하지 않은 변경사항이 있습니다. 저장 후 미리보기할까요?'))saveFile();
}}
var url=previewBase.replace('[?&]edit=1','');
if(url.indexOf('?')>-1)url=url.replace(/[&?]edit=1/g,'');
else url=url;
window.open(url,'_blank');
}}

// 페이지 떠나기 전 경고
window.addEventListener('beforeunload',function(e){{
if(isModified){{e.preventDefault();e.returnValue='';return''}}
}});
</script>
</body></html>'''
        return HTMLResponse(editor_html)
    mime, _ = mimetypes.guess_type(str(t))

    # 인접 파일 정보 계산
    async def get_nav_info():
        parent = t.parent
        try:
            entries = await aio_listdir(parent)  # Path 객체 리스트
            files = sorted(
                [e.name for e in entries if e.is_file() and not e.name.startswith(".")],
                key=lambda n: n.lower()
            )
        except:
            files = []
        if not files or t.name not in files:
            return None, None, 0, 0, []
        idx = files.index(t.name)
        prev_f = files[idx - 1] if idx > 0 else None
        next_f = files[idx + 1] if idx < len(files) - 1 else None
        return prev_f, next_f, idx + 1, len(files), files

    def build_nav_bar(prev_f, next_f, cur_idx, total, all_files):
        """미리보기 페이지 상단에 주입할 네비게이션 바 HTML (한 줄 컴팩트 + 프레젠테이션 모드)"""
        dir_path = str(Path(file_path).parent).replace("\\", "/")
        if dir_path == ".": dir_path = ""
        owner_q = f"&owner={owner}" if owner else ""
        def make_url(fn):
            fp = f"{dir_path}/{fn}" if dir_path else fn
            return f"/{token}/api/preview/{fp}?nav=1{owner_q}"

        def file_icon(fn):
            ext = fn.rsplit('.',1)[-1].lower() if '.' in fn else ''
            icons = {'pptx':'📊','ppt':'📊','xlsx':'📗','xls':'📗','docx':'📘','doc':'📘','pdf':'📕',
                     'hwp':'📄','hwpx':'📄','png':'🖼','jpg':'🖼','jpeg':'🖼','gif':'🖼','webp':'🖼',
                     'html':'🌐','htm':'🌐','md':'📝','txt':'📄','py':'🐍','js':'⚡','css':'🎨','json':'📋'}
            return icons.get(ext, '📄')

        def nav_link(fn, direction):
            if not fn:
                arrow = '◁' if direction == 'prev' else '▷'
                return f'<span style="color:#ccc;font-size:12px;padding:0 4px">{arrow}</span>'
            arrow = '◀' if direction == 'prev' else '▶'
            url = make_url(fn)
            display = fn if len(fn) <= 18 else fn[:8] + '…' + fn[-7:]
            if direction == 'prev':
                return f'<a href="{url}" style="display:inline-flex;align-items:center;gap:4px;padding:3px 8px;border-radius:4px;background:#f3f4f6;color:#374151;font-size:11px;text-decoration:none;white-space:nowrap" title="{fn}">{arrow} {display}</a>'
            else:
                return f'<a href="{url}" style="display:inline-flex;align-items:center;gap:4px;padding:3px 8px;border-radius:4px;background:#f3f4f6;color:#374151;font-size:11px;text-decoration:none;white-space:nowrap" title="{fn}">{display} {arrow}</a>'

        owner_q_param = f"&owner={owner}" if owner else ""
        owner_q_preview = f"?owner={owner}" if owner else ""
        cur_preview_sep = "&" if owner else "?"
        cur_preview = f"/{token}/api/preview/{file_path}{owner_q_preview}{cur_preview_sep}nonav=1"
        cur_download = f"/{token}/api/download?path={file_path}{owner_q_param}"

        # 프레젠테이션 모드용 슬라이드 URL 목록 (HTML 파일만)
        html_exts = {'html', 'htm'}
        slide_files = [f for f in all_files if '.' in f and f.rsplit('.', 1)[-1].lower() in html_exts]
        slide_urls_js = json.dumps([make_url(f).replace('?nav=1', '?nonav=1').replace('&nav=1', '&nonav=1') for f in slide_files])
        slide_names_js = json.dumps(slide_files)
        cur_slide_idx = slide_files.index(t.name) if t.name in slide_files else 0

        pres_btn = ''
        if len(slide_files) > 1:
            pres_btn = f'<button onclick="startPresentation()" style="padding:2px 7px;border-radius:3px;background:#f59e0b;color:#fff;font-size:10px;border:none;cursor:pointer;font-weight:600;white-space:nowrap;flex-shrink:0" title="프레젠테이션 모드">▶ 슬라이드쇼</button>'

        nav_html = f'''<div id="file-nav-bar" style="position:sticky;top:0;z-index:9999;background:#f8fafc;border-bottom:1px solid #e2e8f0;padding:4px 12px;display:flex;align-items:center;justify-content:space-between;gap:8px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;box-shadow:0 1px 2px rgba(0,0,0,.04);height:32px;box-sizing:border-box">
<div style="display:flex;align-items:center;gap:4px;flex-shrink:0">{nav_link(prev_f, 'prev')}</div>
<div style="display:flex;align-items:center;gap:8px;flex:1;justify-content:center;min-width:0;overflow:hidden">
<span style="font-size:11px;font-weight:600;color:#6b7280;flex-shrink:0">{cur_idx}/{total}</span>
<span style="font-size:11px;color:#374151;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="{t.name}">{file_icon(t.name)} {t.name}</span>
{pres_btn}
<a href="{cur_preview}" target="_blank" style="padding:2px 7px;border-radius:3px;background:#6366f1;color:#fff;font-size:10px;text-decoration:none;font-weight:500;white-space:nowrap;flex-shrink:0">↗ 새 창</a>
<a href="{cur_download}" download style="padding:2px 7px;border-radius:3px;background:#475569;color:#fff;font-size:10px;text-decoration:none;font-weight:500;white-space:nowrap;flex-shrink:0">⬇ 다운로드</a>
</div>
<div style="display:flex;align-items:center;gap:4px;flex-shrink:0">{nav_link(next_f, 'next')}</div>
</div>
<div id="pres-overlay" style="display:none;position:fixed;top:0;left:0;width:100vw;height:100vh;background:#000;z-index:100000;flex-direction:column">
<div id="pres-toolbar" style="position:absolute;top:0;left:0;right:0;height:48px;background:linear-gradient(180deg,rgba(0,0,0,.7),transparent);display:flex;align-items:center;justify-content:space-between;padding:0 20px;z-index:100001;opacity:0;transition:opacity .3s">
<span id="pres-counter" style="color:rgba(255,255,255,.8);font-size:14px;font-weight:600"></span>
<div style="display:flex;align-items:center;gap:10px">
<span id="pres-filename" style="color:rgba(255,255,255,.6);font-size:12px"></span>
<button onclick="exitPresentation()" style="padding:6px 14px;border-radius:6px;background:rgba(255,255,255,.15);color:#fff;border:none;cursor:pointer;font-size:12px;font-weight:600;backdrop-filter:blur(4px)">✕ 닫기 (ESC)</button>
</div>
</div>
<iframe id="pres-frame" style="flex:1;border:none;width:100%;height:100%"></iframe>
<div id="pres-key-capture" tabindex="0" style="position:absolute;top:48px;left:15%;right:15%;bottom:60px;z-index:100002;cursor:default;outline:none" onclick="this.focus()"></div>
<div id="pres-nav-left" onclick="presNav(-1)" style="position:absolute;left:0;top:48px;bottom:60px;width:15%;cursor:pointer;z-index:100003"></div>
<div id="pres-nav-right" onclick="presNav(1)" style="position:absolute;right:0;top:48px;bottom:60px;width:15%;cursor:pointer;z-index:100003"></div>
<div id="pres-bottom" style="position:absolute;bottom:0;left:0;right:0;height:60px;display:flex;align-items:center;justify-content:center;gap:6px;background:linear-gradient(0deg,rgba(0,0,0,.7),transparent);z-index:100001;opacity:0;transition:opacity .3s;padding:0 20px;overflow-x:auto">
</div>
</div>
<script>
var _presSlides={slide_urls_js};
var _presNames={slide_names_js};
var _presIdx={cur_slide_idx};
var _presActive=false;
function startPresentation(){{
if(!_presSlides.length)return;
var o=document.getElementById('pres-overlay');
o.style.display='flex';_presActive=true;
presGo(_presIdx);
if(document.documentElement.requestFullscreen)document.documentElement.requestFullscreen().catch(function(){{}});
var dots=document.getElementById('pres-bottom');dots.innerHTML='';
for(var i=0;i<_presSlides.length;i++){{
var d=document.createElement('button');
d.className='pres-dot';d.dataset.i=i;
d.style.cssText='width:10px;height:10px;border-radius:50%;border:2px solid rgba(255,255,255,.5);background:'+(i===_presIdx?'#fff':'transparent')+';cursor:pointer;padding:0;flex-shrink:0;transition:all .2s';
d.onclick=function(){{presGo(parseInt(this.dataset.i))}};
d.title=_presNames[i];
dots.appendChild(d);
}}
setTimeout(function(){{var kc=document.getElementById('pres-key-capture');if(kc)kc.focus()}},200);
}}
function exitPresentation(){{
_presActive=false;
document.getElementById('pres-overlay').style.display='none';
document.getElementById('pres-frame').src='';
if(document.exitFullscreen&&document.fullscreenElement)document.exitFullscreen().catch(function(){{}});
}}
function presGo(i){{
if(i<0||i>=_presSlides.length)return;
_presIdx=i;
document.getElementById('pres-frame').src=_presSlides[i];
document.getElementById('pres-counter').textContent=(i+1)+' / '+_presSlides.length;
document.getElementById('pres-filename').textContent=_presNames[i];
var dots=document.querySelectorAll('.pres-dot');
dots.forEach(function(d,j){{d.style.background=j===i?'#fff':'transparent'}});
setTimeout(function(){{var kc=document.getElementById('pres-key-capture');if(kc)kc.focus()}},100);
}}
function presNav(dir){{presGo(_presIdx+dir)}}
document.addEventListener('keydown',function(e){{
if(!_presActive)return;
if(e.key==='Escape'){{exitPresentation();e.preventDefault()}}
else if(e.key==='ArrowRight'||e.key===' '||e.key==='Enter'){{presNav(1);e.preventDefault()}}
else if(e.key==='ArrowLeft'||e.key==='Backspace'){{presNav(-1);e.preventDefault()}}
else if(e.key==='Home'){{presGo(0);e.preventDefault()}}
else if(e.key==='End'){{presGo(_presSlides.length-1);e.preventDefault()}}
}});
(function(){{
var kc=document.getElementById('pres-key-capture');
if(kc){{
kc.addEventListener('keydown',function(e){{
if(!_presActive)return;
if(e.key==='Escape'){{exitPresentation();e.preventDefault()}}
else if(e.key==='ArrowRight'||e.key===' '||e.key==='Enter'){{presNav(1);e.preventDefault()}}
else if(e.key==='ArrowLeft'||e.key==='Backspace'){{presNav(-1);e.preventDefault()}}
else if(e.key==='Home'){{presGo(0);e.preventDefault()}}
else if(e.key==='End'){{presGo(_presSlides.length-1);e.preventDefault()}}
}});
}}
// iframe이 포커스를 가져가면 캡처 레이어로 되돌림
setInterval(function(){{
if(!_presActive)return;
if(document.activeElement&&document.activeElement.tagName==='IFRAME'){{
var kc2=document.getElementById('pres-key-capture');
if(kc2)kc2.focus();
}}
}},300);
}})();
document.addEventListener('fullscreenchange',function(){{
if(!document.fullscreenElement&&_presActive)exitPresentation();
}});
(function(){{
var tb=document.getElementById('pres-toolbar');
var bt=document.getElementById('pres-bottom');
var ov=document.getElementById('pres-overlay');
var timer;
ov.addEventListener('mousemove',function(){{
tb.style.opacity='1';bt.style.opacity='1';
clearTimeout(timer);
timer=setTimeout(function(){{if(_presActive){{tb.style.opacity='0';bt.style.opacity='0'}}}},2500);
}});
}})();
</script>'''
        return nav_html

    # nav 파라미터가 있거나, 직접 미리보기 가능한 파일이면 네비게이션 바 주입
    prev_f, next_f, cur_idx, total, all_files = await get_nav_info()
    nav_html = build_nav_bar(prev_f, next_f, cur_idx, total, all_files) if total > 1 and not nonav else ""

    def inject_nav(html_content):
        """HTML 콘텐츠의 <body> 직후에 nav bar 삽입 + body에 padding-top 추가 + Tailwind CDN 자동 주입"""
        import re as _re
        # Tailwind CDN 자동 주입: <head>가 있고 tailwindcss가 없으면 추가
        if '<head' in html_content.lower() and 'tailwindcss' not in html_content.lower():
            tw_cdn = '<script src="https://cdn.tailwindcss.com"></script>'
            head_m = _re.search(r'(<head[^>]*>)', html_content, _re.IGNORECASE)
            if head_m:
                hp = head_m.end()
                html_content = html_content[:hp] + '\n' + tw_cdn + '\n' + html_content[hp:]
        if not nav_html:
            return html_content
        # nav bar를 fixed로 고정하고 body에 상단 여백 추가
        nav_style_patch = """<style>
#file-nav-bar{position:fixed!important;top:0!important;left:0!important;right:0!important;z-index:99999!important;background:linear-gradient(135deg,#f8fafc,#eef2ff)!important}
</style>
<script>(function(){
var n=document.getElementById('file-nav-bar');if(!n)return;
function u(){
  var h=n.offsetHeight||36;
  document.body.style.paddingTop=h+'px';
  // 원본 HTML의 fixed/sticky 요소들도 아래로 밀기
  var all=document.querySelectorAll('*');
  for(var i=0;i<all.length;i++){
    if(all[i]===n||n.contains(all[i]))continue;
    var s=getComputedStyle(all[i]);
    if((s.position==='fixed'||s.position==='sticky')&&(s.top==='0px'||s.top==='0')){
      all[i].style.top=h+'px';
    }
  }
}
u();
new ResizeObserver(u).observe(n);
window.addEventListener('load',u);
setTimeout(u,500);
})()</script>"""
        m = _re.search(r'(<body[^>]*>)', html_content, _re.IGNORECASE)
        if m:
            pos = m.end()
            return html_content[:pos] + nav_html + nav_style_patch + html_content[pos:]
        return nav_html + nav_style_patch + html_content

    # HTML 파일 + 공유 모드: 상대 경로에 owner 쿼리 자동 주입
    if owner and mime and mime.startswith("text/html"):
        content = await aio_read_text(t)
        # 현재 파일의 디렉토리 기준 base path 계산
        dir_path = str(Path(file_path).parent).replace("\\", "/")
        if dir_path == ".": dir_path = ""

        HTML_EXT = {'html','htm'}
        CSS_JS_EXT = {'css','js'}
        def resolve_path(url):
            if dir_path:
                resolved = str(Path(dir_path + "/" + url)).replace("\\", "/")
            else:
                resolved = url
            parts = []
            for part in resolved.split("/"):
                if part == "..":
                    if parts: parts.pop()
                elif part and part != ".":
                    parts.append(part)
            return "/".join(parts)

        def rewrite_url(match):
            attr = match.group(1)  # src= or href=
            quote = match.group(2)  # ' or "
            url = match.group(3)
            if url.startswith(('http://', 'https://', 'data:', '#', '//', 'javascript:')):
                return match.group(0)
            if 'owner=' in url or '/api/' in url:
                return match.group(0)
            resolved = resolve_path(url)
            ext = resolved.rsplit('.', 1)[-1].lower() if '.' in resolved else ''
            # HTML → preview, CSS/JS → preview (내부 url 재작성 필요), 나머지 → download
            if ext in HTML_EXT or ext in CSS_JS_EXT:
                new_url = f"/{token}/api/preview/{resolved}?owner={owner}"
            else:
                new_url = f"/{token}/api/download?path={resolved}&owner={owner}"
            return f"{attr}{quote}{new_url}{quote}"
        content = re.sub(r'((?:src|href)\s*=\s*)(["\'])([^"\']*?)\2', rewrite_url, content, flags=re.IGNORECASE)
        # CSS url() 안의 경로도 재작성 (이미지/폰트 → download)
        def rewrite_css_url(match):
            prefix = match.group(1)
            url = match.group(2)
            if url.startswith(('http://', 'https://', 'data:', '#', '//')):
                return match.group(0)
            if 'owner=' in url or '/api/' in url:
                return match.group(0)
            resolved = resolve_path(url)
            return f"{prefix}/{token}/api/download?path={resolved}&owner={owner}"
        content = re.sub(r'(url\s*\(\s*["\']?)([^"\')\s]+?)(?=["\']?\s*\))', rewrite_css_url, content, flags=re.IGNORECASE)
        return HTMLResponse(content=inject_nav(content))
    # CSS 파일: nav=1이면 코드 미리보기 (nav bar 포함), 아니면 raw CSS
    if mime and mime == "text/css":
        content = await aio_read_text(t)
        dir_path = str(Path(file_path).parent).replace("\\", "/")
        if dir_path == ".": dir_path = ""
        owner_q_css = f"&owner={owner}" if owner else ""
        def resolve_css_path(url):
            if dir_path:
                resolved = str(Path(dir_path + "/" + url)).replace("\\", "/")
            else:
                resolved = url
            parts = []
            for part in resolved.split("/"):
                if part == "..":
                    if parts: parts.pop()
                elif part and part != ".":
                    parts.append(part)
            return "/".join(parts)
        def rewrite_css_url2(match):
            prefix = match.group(1)
            url = match.group(2)
            if url.startswith(('http://', 'https://', 'data:', '#', '//')):
                return match.group(0)
            if '/api/' in url:
                return match.group(0)
            resolved = resolve_css_path(url)
            ext = resolved.rsplit('.', 1)[-1].lower() if '.' in resolved else ''
            if ext == 'css':
                return f"{prefix}/{token}/api/preview/{resolved}?owner={owner}" if owner else f"{prefix}/{token}/api/preview/{resolved}"
            return f"{prefix}/{token}/api/download?path={resolved}{owner_q_css}"
        content = re.sub(r'(url\s*\(\s*["\']?)([^"\')\s]+?)(?=["\']?\s*\))', rewrite_css_url2, content, flags=re.IGNORECASE)
        def rewrite_css_import(match):
            prefix = match.group(1)
            quote = match.group(2)
            url = match.group(3)
            if url.startswith(('http://', 'https://', '//')):
                return match.group(0)
            if '/api/' in url:
                return match.group(0)
            resolved = resolve_css_path(url)
            if owner:
                new_url = f"/{token}/api/preview/{resolved}?owner={owner}"
            else:
                new_url = f"/{token}/api/preview/{resolved}"
            return f"{prefix}{quote}{new_url}{quote}"
        content = re.sub(r'(@import\s+)(["\'])([^"\']+?)\2', rewrite_css_import, content, flags=re.IGNORECASE)
        # nav=1 파라미터가 있으면 코드 미리보기 HTML로 래핑 (nav bar 포함)
        if nav and nav_html:
            import html as _html
            escaped = _html.escape(content)
            code_html = f'''<!DOCTYPE html><html><head><meta charset="utf-8"><title>{t.name}</title>
<style>body{{margin:0;font-family:monospace}}pre{{margin:0;padding:16px;background:#1e1e1e;color:#d4d4d4;font-size:13px;line-height:1.5;overflow-x:auto;min-height:calc(100vh - 40px)}}</style></head>
<body>{nav_html}<pre>{escaped}</pre></body></html>'''
            return HTMLResponse(content=code_html)
        return Response(content=content, media_type="text/css")
    # Markdown 파일: HTML로 변환하여 미리보기
    if str(t).lower().endswith(".md"):
        md_content = await aio_read_text(t)
        md_inner = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{t.name}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/github-markdown-css/5.5.1/github-markdown-light.min.css">
<style>
body{{max-width:880px;margin:0 auto;padding:32px 24px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,Arial,sans-serif}}
.markdown-body{{font-size:15px;line-height:1.7}}
.markdown-body pre{{background:#f6f8fa;border-radius:6px;padding:16px;overflow-x:auto}}
.markdown-body code{{background:#f0f0f0;padding:2px 6px;border-radius:3px;font-size:13px}}
.markdown-body pre code{{background:none;padding:0}}
.markdown-body table{{border-collapse:collapse;width:100%}}
.markdown-body th,.markdown-body td{{border:1px solid #d0d7de;padding:8px 12px}}
.markdown-body th{{background:#f6f8fa;font-weight:600}}
.markdown-body img{{max-width:100%}}
.markdown-body blockquote{{border-left:4px solid #d0d7de;margin:0;padding:0 16px;color:#656d76}}
</style>
<script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
</head><body>
<article class="markdown-body" id="content"></article>
<script>
var raw = {json.dumps(md_content)};
document.getElementById('content').innerHTML = marked.parse(raw);
</script></body></html>"""
        if nav_html:
            return HTMLResponse(inject_nav(md_inner))
        return HTMLResponse(md_inner)
    # 일반 HTML 파일: 상대 경로 재작성 + nav bar
    if mime and mime.startswith("text/html"):
        content = await aio_read_text(t)
        dir_path = str(Path(file_path).parent).replace("\\", "/")
        if dir_path == ".": dir_path = ""
        owner_q = f"&owner={owner}" if owner else ""
        owner_q_preview = f"?owner={owner}" if owner else ""
        HTML_EXT2 = {'html','htm'}
        CSS_JS_EXT2 = {'css','js'}

        def resolve_path2(url):
            if dir_path:
                resolved = str(Path(dir_path + "/" + url)).replace("\\", "/")
            else:
                resolved = url
            parts = []
            for part in resolved.split("/"):
                if part == "..":
                    if parts: parts.pop()
                elif part and part != ".":
                    parts.append(part)
            return "/".join(parts)

        def rewrite_html_url(match):
            attr = match.group(1)
            quote = match.group(2)
            url = match.group(3)
            if url.startswith(('http://', 'https://', 'data:', '#', '//', 'javascript:')):
                return match.group(0)
            if '/api/' in url:
                return match.group(0)
            resolved = resolve_path2(url)
            ext = resolved.rsplit('.', 1)[-1].lower() if '.' in resolved else ''
            if ext in HTML_EXT2 or ext in CSS_JS_EXT2:
                new_url = f"/{token}/api/preview/{resolved}{owner_q_preview}"
            else:
                new_url = f"/{token}/api/download?path={resolved}{owner_q}"
            return f"{attr}{quote}{new_url}{quote}"
        content = re.sub(r'((?:src|href)\s*=\s*)(["\'])([^"\']*?)\2', rewrite_html_url, content, flags=re.IGNORECASE)
        # CSS url() 재작성 (이미지/폰트 → download)
        def rewrite_html_css_url(match):
            prefix = match.group(1)
            url = match.group(2)
            if url.startswith(('http://', 'https://', 'data:', '#', '//')):
                return match.group(0)
            if '/api/' in url:
                return match.group(0)
            resolved = resolve_path2(url)
            return f"{prefix}/{token}/api/download?path={resolved}{owner_q}"
        content = re.sub(r'(url\s*\(\s*["\']?)([^"\')\s]+?)(?=["\']?\s*\))', rewrite_html_css_url, content, flags=re.IGNORECASE)
        if nav_html:
            return HTMLResponse(inject_nav(content))
        return HTMLResponse(content)
    # 이미지 파일: HTML 래퍼로 감싸 nav 제공
    if nav_html and mime and mime.startswith("image/"):
        img_path = f"/{token}/api/download?path={file_path}"
        if owner: img_path += f"&owner={owner}"
        img_inner = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>{t.name}</title>
<style>body{{margin:0;background:#1a1a2e;display:flex;align-items:center;justify-content:center;min-height:100vh}}
img{{max-width:95%;max-height:95vh;object-fit:contain;border-radius:4px;box-shadow:0 4px 20px rgba(0,0,0,.3)}}</style>
</head><body><img src="{img_path}"></body></html>"""
        return HTMLResponse(inject_nav(img_inner))
    # 텍스트/코드 파일: HTML 래퍼로 감싸 nav 제공
    if nav_html and mime and (mime.startswith("text/") or mime in ("application/json","application/xml","application/javascript")):
        txt_content = await aio_read_text(t)
        import html as html_mod
        escaped = html_mod.escape(txt_content)
        ext = t.name.rsplit('.',1)[-1].lower() if '.' in t.name else ''
        txt_inner = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>{t.name}</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>
<style>body{{margin:0;font-family:-apple-system,sans-serif}} pre{{margin:0;padding:20px;font-size:13px;line-height:1.6;overflow-x:auto}}</style>
</head><body><pre><code class="language-{ext}">{escaped}</code></pre>
<script>hljs.highlightAll();</script></body></html>"""
        return HTMLResponse(inject_nav(txt_inner))
    # 오피스/PDF 파일: iframe으로 문서뷰어 임베드 + nav bar
    OFFICE_PREVIEW_EXT = {'pptx','ppt','xlsx','xls','docx','doc','pdf','hwp','hwpx','cell','show'}
    file_ext = t.name.rsplit('.',1)[-1].lower() if '.' in t.name else ''
    if nav_html and file_ext in OFFICE_PREVIEW_EXT:
        owner_param = f'"owner":"{owner}",' if owner else ''
        file_icons = {'pptx':'📊','ppt':'📊','xlsx':'📗','xls':'📗','docx':'📘','doc':'📘',
                      'pdf':'📕','hwp':'📄','hwpx':'📄','cell':'📗','show':'📊'}
        icon = file_icons.get(file_ext, '📄')
        office_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>{t.name}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;height:100vh;display:flex;flex-direction:column;background:#f1f5f9}}
#viewer-frame{{flex:1;border:none;width:100%}}
.loading-wrap{{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:16px}}
.loading-spinner{{width:40px;height:40px;border:4px solid #e2e8f0;border-top-color:#3b82f6;border-radius:50%;animation:spin .8s linear infinite}}
@keyframes spin{{to{{transform:rotate(360deg)}}}}
.loading-text{{font-size:14px;color:#64748b}}
.file-icon{{font-size:48px;margin-bottom:8px}}
</style></head><body>
{nav_html}
<div class="loading-wrap" id="loading">
<div class="file-icon">{icon}</div>
<div class="loading-spinner"></div>
<div class="loading-text">{t.name} 문서 뷰어를 로딩 중입니다...</div>
</div>
<iframe id="viewer-frame" style="display:none"></iframe>
<script>
(function(){{
  var fp="{file_path}";
  fetch("/{token}/api/temp-link",{{method:"POST",headers:{{"Content-Type":"application/json"}},body:JSON.stringify({{{owner_param}"path":fp}})}})
  .then(function(r){{return r.json()}}).then(function(d){{
    var pubUrl="{PORTAL_URL}"+d.url;
    var fn="{t.name}";
    var ext=fn.split(".").pop().toLowerCase();
    var ch="0123456789abcdef",h="";for(var i=0;i<32;i++)h+=ch.charAt(Math.floor(Math.random()*16));
    var dockey="upload_"+h+"."+ext;
    var vUrl="{PORTAL_URL}/officeview/ov.jsp?url="+encodeURIComponent(pubUrl)+"&filename="+encodeURIComponent(fn)+"&dockey="+encodeURIComponent(dockey);
    var f=document.getElementById("viewer-frame");
    f.src=vUrl;
    f.style.display="block";
    document.getElementById("loading").style.display="none";
  }}).catch(function(){{
    document.getElementById("loading").innerHTML='<div class="file-icon">⚠️</div><div class="loading-text">문서를 로드할 수 없습니다</div>';
  }});
}})();
</script></body></html>"""
        return HTMLResponse(office_html)
    return FileResponse(str(t), media_type=mime or "application/octet-stream", headers={"Content-Disposition":"inline"})

@app.post("/{token}/api/create-folder")
async def u_create_folder(token: str, payload: dict):
    owner = payload.get("owner")
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read": raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    name = payload.get("name","").strip(); base = payload.get("path",".")
    if not name: raise HTTPException(400, "이름 필요")
    full = os.path.join(base, name) if base != "." else name; t = safe_path(full, ws)
    if await aio_exists(t): raise HTTPException(400, "이미 존재")
    await aio_mkdir(t); return {"success":True,"path":full}

@app.delete("/{token}/api/file")
async def u_delete(token: str, path: str, owner: str = None):
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read": raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    t = safe_path(path, ws)
    if not await aio_exists(t): raise HTTPException(404)
    if await aio_is_dir(t): await aio_rmtree(t)
    else: await aio_unlink(t)
    return {"success":True}

@app.post("/{token}/api/delete-all")
async def u_delete_all(token: str, payload: dict):
    owner = payload.get("owner")
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read": raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    t = safe_path(payload.get("path","."), ws)
    if not await aio_exists(t) or not await aio_is_dir(t): raise HTTPException(404)
    deleted = []
    for e in await aio_listdir(t):
        try:
            if await aio_is_dir(e): await aio_rmtree(e)
            else: await aio_unlink(e)
            deleted.append(e.name)
        except: pass
    return {"success":True,"count":len(deleted)}

@app.post("/{token}/api/rename")
async def u_rename(token: str, payload: dict):
    """파일 또는 폴더 이름 변경"""
    owner = payload.get("owner")
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read": raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    old_path = payload.get("path", "").strip()
    new_name = payload.get("newName", "").strip()
    if not old_path or not new_name:
        raise HTTPException(400, "경로와 새 이름을 입력해주세요")
    if '/' in new_name or '\\' in new_name or '..' in new_name:
        raise HTTPException(400, "이름에 경로 구분자를 포함할 수 없습니다")
    src = safe_path(old_path, ws)
    if not await aio_exists(src):
        raise HTTPException(404, "파일 또는 폴더를 찾을 수 없습니다")
    dst = src.parent / new_name
    if await aio_exists(dst):
        raise HTTPException(409, "같은 이름의 파일/폴더가 이미 존재합니다")
    await asyncio.to_thread(src.rename, dst)
    return {"success": True, "oldPath": old_path, "newName": new_name}

@app.post("/{token}/api/move")
async def u_move(token: str, payload: dict):
    """파일/폴더를 다른 폴더로 이동"""
    owner = payload.get("owner")
    _, ws, perm = await _resolve_workspace(token, owner)
    if perm == "read": raise HTTPException(403, "읽기 전용 공유 폴더입니다")
    items = payload.get("items", [])
    dest_folder = payload.get("destFolder", "").strip()
    if not items:
        raise HTTPException(400, "이동할 항목을 지정해주세요")
    dest = safe_path(dest_folder, ws) if dest_folder and dest_folder != "." else Path(ws)
    if not await aio_exists(dest) or not await aio_is_dir(dest):
        raise HTTPException(404, "대상 폴더를 찾을 수 없습니다")
    moved = []
    errors = []
    for item_path in items:
        try:
            src = safe_path(item_path, ws)
            if not await aio_exists(src):
                errors.append({"path": item_path, "error": "파일 없음"})
                continue
            target = dest / src.name
            if await aio_exists(target):
                errors.append({"path": item_path, "error": "대상 폴더에 같은 이름 존재"})
                continue
            if str(dest).startswith(str(src)):
                errors.append({"path": item_path, "error": "자기 자신 안으로 이동 불가"})
                continue
            await asyncio.to_thread(shutil.move, str(src), str(target))
            moved.append(item_path)
        except Exception as e:
            errors.append({"path": item_path, "error": str(e)})
    return {"success": True, "moved": moved, "errors": errors}

# ============================================================
# 조직도 사용자 검색 API
# ============================================================
@app.get("/{token}/api/org/search")
async def u_org_search(token: str, q: str = ""):
    """조직도에서 이름으로 사용자 검색"""
    _resolve_user(token)  # 인증 확인
    if not q or len(q.strip()) < 1:
        return {"users": []}
    if not MONGO_OK or org_user_collection is None:
        raise HTTPException(500, "조직도 DB 미연결")
    query = q.strip()
    # 이름(nm)으로 검색 (정규식 부분일치)
    cursor = org_user_collection.find(
        {"nm": {"$regex": query, "$options": "i"}},
        {"_id": 0, "nm": 1, "lid": 1, "dp": 1}
    ).limit(50)
    users = []
    async for doc in cursor:
        if doc.get("lid"):
            users.append({
                "name": doc.get("nm", ""),
                "lid": doc.get("lid", ""),
                "dept": doc.get("dp", ""),
            })
    return {"users": users}

@app.get("/{token}/api/org/user")
async def u_org_user(token: str, lid: str = ""):
    """조직도에서 lid로 사용자 정보 조회"""
    _resolve_user(token)
    if not lid or not lid.strip():
        raise HTTPException(400, "lid 필요")
    if not MONGO_OK or org_user_collection is None:
        raise HTTPException(500, "조직도 DB 미연결")
    doc = await org_user_collection.find_one(
        {"lid": lid.strip()},
        {"_id": 0, "nm": 1, "lid": 1, "dp": 1}
    )
    if not doc:
        return {"found": False}
    return {
        "found": True,
        "name": doc.get("nm", ""),
        "lid": doc.get("lid", ""),
        "dept": doc.get("dp", ""),
    }

# ============================================================
# 폴더 공유 API
# ============================================================
@app.post("/{token}/api/share")
async def u_share_folder(token: str, payload: dict):
    """폴더를 다른 사용자에게 공유 (단일 또는 복수)"""
    owner, ws = _resolve_user(token)
    folder_path = payload.get("path", "").strip()
    permission = payload.get("permission", "read")  # read | write
    # 복수 사용자 지원: targetUsers(배열) 또는 targetUser(단일)
    target_users = payload.get("targetUsers", [])
    if not target_users:
        single = payload.get("targetUser", "").strip()
        if single: target_users = [single]
    if not folder_path or not target_users:
        raise HTTPException(400, "폴더 경로와 대상 사용자를 지정해주세요")
    # 폴더 존재 확인
    t = safe_path(folder_path, ws)
    if not await aio_exists(t) or not await aio_is_dir(t):
        raise HTTPException(404, "폴더를 찾을 수 없습니다")
    if not MONGO_OK or shared_folders_collection is None:
        raise HTTPException(500, "DB 미연결")
    results = []
    for target_user in target_users:
        target_user = target_user.strip()
        if not target_user: continue
        if target_user == owner:
            results.append({"user": target_user, "status": "skipped", "reason": "자기 자신"}); continue
        existing = await shared_folders_collection.find_one({
            "owner": owner, "folder_path": folder_path, "shared_with": target_user
        })
        if existing:
            await shared_folders_collection.update_one(
                {"_id": existing["_id"]},
                {"$set": {"permission": permission, "updated_at": datetime.now(timezone.utc)}}
            )
            results.append({"user": target_user, "status": "updated"})
        else:
            await shared_folders_collection.insert_one({
                "owner": owner,
                "folder_path": folder_path,
                "folder_name": os.path.basename(folder_path) or folder_path,
                "shared_with": target_user,
                "permission": permission,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc),
            })
            results.append({"user": target_user, "status": "shared"})
    shared_count = sum(1 for r in results if r["status"] in ("shared","updated"))
    return {"success": True, "message": f"{shared_count}명에게 공유되었습니다", "results": results}

@app.delete("/{token}/api/share")
async def u_unshare_folder(token: str, share_id: str):
    """공유 해제 (소유자 또는 공유받은 사용자 모두 가능)"""
    user, _ = _resolve_user(token)
    if not MONGO_OK or shared_folders_collection is None:
        raise HTTPException(500, "DB 미연결")
    from bson import ObjectId
    try:
        doc = await shared_folders_collection.find_one({"_id": ObjectId(share_id)})
    except Exception:
        raise HTTPException(400, "유효하지 않은 공유 ID")
    if not doc:
        raise HTTPException(404, "공유 항목을 찾을 수 없습니다")
    if doc["owner"] != user and doc["shared_with"] != user:
        raise HTTPException(403, "권한이 없습니다")
    await shared_folders_collection.delete_one({"_id": ObjectId(share_id)})
    return {"success": True}

@app.get("/{token}/api/shares/my")
async def u_my_shares(token: str):
    """내가 공유한 폴더 목록"""
    user, _ = _resolve_user(token)
    if not MONGO_OK or shared_folders_collection is None:
        return {"shares": []}
    cursor = shared_folders_collection.find({"owner": user}).sort("created_at", -1)
    shares = []
    async for d in cursor:
        shares.append({
            "id": str(d["_id"]),
            "folder_path": d["folder_path"],
            "folder_name": d.get("folder_name", ""),
            "shared_with": d["shared_with"],
            "permission": d.get("permission", "read"),
        })
    return {"shares": shares}

@app.get("/{token}/api/shares/received")
async def u_received_shares(token: str):
    """나에게 공유된 폴더 목록"""
    user, _ = _resolve_user(token)
    if not MONGO_OK or shared_folders_collection is None:
        return {"shares": []}
    cursor = shared_folders_collection.find({"shared_with": user}).sort("created_at", -1)
    shares = []
    async for d in cursor:
        shares.append({
            "id": str(d["_id"]),
            "owner": d["owner"],
            "folder_path": d["folder_path"],
            "folder_name": d.get("folder_name", ""),
            "permission": d.get("permission", "read"),
        })
    return {"shares": shares}

@app.get("/{token}/api/shares/files")
async def u_shared_files(token: str, owner: str, path: str = "."):
    """공유받은 폴더의 파일 목록 조회"""
    user, _ = _resolve_user(token)
    if not MONGO_OK or shared_folders_collection is None:
        raise HTTPException(500, "DB 미연결")
    # 공유 권한 확인
    share = await shared_folders_collection.find_one({"owner": owner, "shared_with": user})
    if not share:
        raise HTTPException(403, "공유받지 않은 폴더입니다")
    # 요청 경로가 공유된 폴더 범위 안인지 확인
    shared_root = share["folder_path"]
    if path == ".":
        actual_path = shared_root
    else:
        actual_path = path
    # actual_path가 shared_root 하위인지 확인 (보안)
    if actual_path != shared_root and not actual_path.startswith(shared_root + "/"):
        raise HTTPException(403, "공유 범위를 벗어난 접근입니다")
    owner_ws = get_user_workspace(owner)
    return JSONResponse(json.loads(await execute_tool("list_files", {"path": actual_path}, owner_ws)))

@app.post("/{token}/api/shares/copy")
async def u_shared_copy(token: str, payload: dict):
    """공유받은 파일을 내 작업 공간으로 복사"""
    user, my_ws = _resolve_user(token)
    owner = payload.get("owner", "").strip()
    src_path = payload.get("srcPath", "").strip()
    dest_path = payload.get("destPath", ".").strip()
    if not owner or not src_path:
        raise HTTPException(400, "소유자와 원본 경로를 지정해주세요")
    if not MONGO_OK or shared_folders_collection is None:
        raise HTTPException(500, "DB 미연결")
    share = await shared_folders_collection.find_one({"owner": owner, "shared_with": user})
    if not share:
        raise HTTPException(403, "공유받지 않은 폴더입니다")
    shared_root = share["folder_path"]
    if src_path != shared_root and not src_path.startswith(shared_root + "/"):
        raise HTTPException(403, "공유 범위를 벗어난 접근입니다")
    owner_ws = get_user_workspace(owner)
    src = safe_path(src_path, owner_ws)
    dest_dir = safe_path(dest_path, my_ws) if dest_path != "." else Path(my_ws)
    if not await aio_exists(src):
        raise HTTPException(404, "원본 파일/폴더를 찾을 수 없습니다")
    target = dest_dir / src.name
    if await aio_exists(target):
        raise HTTPException(409, "같은 이름의 항목이 이미 존재합니다")
    if await aio_is_dir(src):
        await asyncio.to_thread(shutil.copytree, str(src), str(target))
    else:
        await aio_mkdir(dest_dir)
        await asyncio.to_thread(shutil.copy2, str(src), str(target))
    return {"success": True, "copied": src_path, "dest": str(target.name)}

@app.get("/{token}/api/download-folder")
async def u_dl_folder(token: str, path: str, owner: str = None):
    import tempfile; _, ws, _ = await _resolve_workspace(token, owner); t = safe_path(path, ws)
    if not await aio_exists(t) or not await aio_is_dir(t): raise HTTPException(404)
    def _z(): td = tempfile.mkdtemp(); zp = os.path.join(td, t.name or "ws"); shutil.make_archive(zp, 'zip', str(t)); return zp + ".zip"
    return FileResponse(await asyncio.to_thread(_z), filename=(t.name or "ws")+".zip", media_type="application/zip")

@app.get("/{token}/api/download-multi")
async def u_dl_multi(token: str, paths: str):
    """변경된 파일 일괄 다운로드 (zip)"""
    import tempfile, zipfile
    u, _ = _resolve_user(token)
    ws = Path(get_user_workspace(u))
    file_paths = [p.strip() for p in paths.split(",") if p.strip()]
    if not file_paths: raise HTTPException(400, "파일 경로가 없습니다")
    def _make_zip():
        td = tempfile.mkdtemp()
        zp = os.path.join(td, "modified_files.zip")
        with zipfile.ZipFile(zp, "w", zipfile.ZIP_DEFLATED) as zf:
            for fp in file_paths:
                full = safe_path(fp, str(ws))
                if os.path.isfile(str(full)):
                    # zip 내 경로: _projects/xxx/ prefix 제거
                    arc_name = fp
                    if "_projects/" in arc_name:
                        parts = arc_name.split("/", 2)  # _projects/{id}/나머지
                        if len(parts) >= 3: arc_name = parts[2]
                    zf.write(str(full), arc_name)
        return zp
    zip_path = await asyncio.to_thread(_make_zip)
    return FileResponse(zip_path, filename="modified_files.zip", media_type="application/zip")

@app.get("/{token}/api/chat-logs")
async def u_chat_logs(token: str, skip: int = 0, limit: int = 13):
    u, _ = _resolve_user(token)
    if not MONGO_OK: return {"logs":[], "total": 0, "skip": skip, "limit": limit}
    total = await chat_collection.count_documents({"username": u})
    cursor = chat_collection.find({"username":u}, {"messages":0,"api_history":0}).sort("updated_at",-1).skip(skip).limit(limit); logs = []
    async for d in cursor:
        logs.append({"session_id":d.get("session_id",""),"title":d.get("title",""),"updated_at":str(d.get("updated_at","")),"current_folder":d.get("current_folder",".")})
    return {"logs":logs, "total": total, "skip": skip, "limit": limit}

@app.get("/{token}/api/chat-log/{session_id}")
async def u_chat_detail(token: str, session_id: str):
    u, _ = _resolve_user(token)
    if not MONGO_OK: raise HTTPException(500)
    d = await chat_collection.find_one({"session_id":session_id,"username":u}, {"api_history":0})
    if not d: raise HTTPException(404)
    return {"session_id":d["session_id"],"title":d.get("title",""),"messages":d.get("messages",[])}

@app.delete("/{token}/api/chat-log/{session_id}")
async def u_del_log(token: str, session_id: str):
    u, _ = _resolve_user(token)
    if not MONGO_OK: raise HTTPException(500)
    return {"success":(await chat_collection.delete_one({"session_id":session_id,"username":u})).deleted_count > 0}

# ============ 프로젝트 API ============

@app.get("/{token}/api/projects")
async def get_projects(token: str, skip: int = 0, limit: int = 10):
    """프로젝트 목록 조회 (페이징)"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        return {"projects": [], "total": 0}
    total = await projects_collection.count_documents({"username": u})
    cursor = projects_collection.find({"username": u}, {"files_content": 0}).sort("updated_at", -1).skip(skip).limit(limit)
    projects = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        projects.append(doc)
    return {"projects": projects, "total": total}

@app.get("/{token}/api/projects/{project_id}")
async def get_project_detail(token: str, project_id: str):
    """프로젝트 상세 조회"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(404)
    from bson import ObjectId
    doc = await projects_collection.find_one({"_id": ObjectId(project_id), "username": u})
    if not doc:
        raise HTTPException(404, "프로젝트를 찾을 수 없습니다")
    doc["_id"] = str(doc["_id"])
    return doc

@app.post("/{token}/api/projects")
async def create_project(token: str, payload: dict):
    """프로젝트 생성"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(500)
    now = datetime.now(timezone.utc)
    doc = {
        "username": u,
        "name": payload.get("name", "").strip(),
        "description": payload.get("description", "").strip(),
        "instructions": payload.get("instructions", "").strip(),
        "files": [],  # [{name, path, size, uploaded_at}]
        "files_content": [],  # [{name, content}] 텍스트 파일 내용
        "output_files": [],  # [{name, path, size, created_at, source_file}] 수정/생성된 파일
        "created_at": now,
        "updated_at": now
    }
    if not doc["name"]:
        raise HTTPException(400, "프로젝트명은 필수입니다")
    result = await projects_collection.insert_one(doc)
    doc["_id"] = str(result.inserted_id)
    return doc

@app.put("/{token}/api/projects/{project_id}")
async def update_project(token: str, project_id: str, payload: dict):
    """프로젝트 수정 (이름, 설명, 지침)"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(500)
    from bson import ObjectId
    update = {"updated_at": datetime.now(timezone.utc)}
    if "name" in payload:
        update["name"] = payload["name"].strip()
    if "description" in payload:
        update["description"] = payload["description"].strip()
    if "instructions" in payload:
        update["instructions"] = payload["instructions"].strip()
    result = await projects_collection.update_one(
        {"_id": ObjectId(project_id), "username": u},
        {"$set": update}
    )
    if result.matched_count == 0:
        raise HTTPException(404)
    return {"success": True}

@app.delete("/{token}/api/projects/{project_id}")
async def delete_project(token: str, project_id: str):
    """프로젝트 삭제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(500)
    from bson import ObjectId
    # 프로젝트 파일 폴더 삭제
    ws = get_user_workspace(u)
    proj_dir = Path(ws) / "_projects" / project_id
    if proj_dir.exists():
        import shutil
        shutil.rmtree(str(proj_dir), ignore_errors=True)
    result = await projects_collection.delete_one({"_id": ObjectId(project_id), "username": u})
    return {"success": result.deleted_count > 0}

@app.post("/{token}/api/projects/{project_id}/files")
async def upload_project_files(token: str, project_id: str, files: list[UploadFile] = File(...), subpath: str = Form(".")):
    """프로젝트 파일 업로드 (멀티 파일, 폴더 구조 보존, 중복명 자동 번호)"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(500)
    from bson import ObjectId
    ws = get_user_workspace(u)
    proj_dir = Path(ws) / "_projects" / project_id
    target_dir = proj_dir / subpath if subpath and subpath != "." else proj_dir
    os.makedirs(str(target_dir), exist_ok=True)
    text_exts = {'.md','.txt','.py','.js','.ts','.html','.css','.json','.xml','.csv','.yaml','.yml','.sh','.sql','.java','.jsx','.tsx','.ini','.cfg','.conf','.env'}
    uploaded = []
    files_to_push = []
    content_to_push = []
    for file in files:
        fname = file.filename or "file"
        # 폴더 구조가 포함된 경우 하위 디렉토리 생성
        if '/' in fname:
            file_target_dir = target_dir / str(Path(fname).parent)
            os.makedirs(str(file_target_dir), exist_ok=True)
        else:
            file_target_dir = target_dir
        base_name = Path(fname).name
        # 중복 파일명 처리
        stem = Path(base_name).stem
        ext_s = Path(base_name).suffix
        final_name = base_name
        counter = 1
        while (file_target_dir / final_name).exists():
            final_name = f"{stem}({counter}){ext_s}"
            counter += 1
        if '/' in fname:
            rel_path = str(Path(fname).parent / final_name)
        else:
            rel_path = final_name
        if subpath and subpath != ".":
            full_rel = f"{subpath}/{rel_path}"
        else:
            full_rel = rel_path
        fpath = file_target_dir / final_name
        content = await file.read()
        await aio_write_bytes(fpath, content)
        file_info = {
            "name": final_name,
            "rel_path": full_rel,
            "path": f"_projects/{project_id}/{full_rel}",
            "size": len(content),
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        }
        files_to_push.append(file_info)
        # 텍스트 파일 내용 저장
        if Path(final_name).suffix.lower() in text_exts:
            try:
                text_content = content.decode('utf-8')[:50000]
            except:
                try:
                    text_content = content.decode('euc-kr')[:50000]
                except:
                    text_content = None
            if text_content:
                content_to_push.append({"name": full_rel, "content": text_content})
        uploaded.append(file_info)
    # 배치 DB 업데이트
    push_ops = {}
    if files_to_push:
        push_ops["files"] = {"$each": files_to_push}
    if content_to_push:
        push_ops["files_content"] = {"$each": content_to_push}
    if push_ops:
        await projects_collection.update_one(
            {"_id": ObjectId(project_id), "username": u},
            {"$push": push_ops, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    return {"success": True, "files": uploaded, "count": len(uploaded)}

@app.get("/{token}/api/projects/{project_id}/files")
async def list_project_files(token: str, project_id: str, subpath: str = "."):
    """프로젝트 파일/폴더 목록 (디스크 기반 탐색)"""
    u, _ = _resolve_user(token)
    ws = get_user_workspace(u)
    proj_dir = Path(ws) / "_projects" / project_id
    target = proj_dir / subpath if subpath and subpath != "." else proj_dir
    if not target.exists():
        os.makedirs(str(target), exist_ok=True)
    items = []
    try:
        for entry in sorted(target.iterdir(), key=lambda e: (not e.is_dir(), e.name.lower())):
            if entry.name.startswith('.') or entry.name.startswith('_'):
                continue
            rel = str(entry.relative_to(proj_dir)).replace('\\', '/')
            if entry.is_dir():
                child_count = sum(1 for c in entry.iterdir() if not c.name.startswith('.'))
                items.append({"name": entry.name, "type": "directory", "rel_path": rel, "child_count": child_count})
            else:
                items.append({"name": entry.name, "type": "file", "rel_path": rel, "size": entry.stat().st_size})
    except:
        pass
    return {"items": items, "subpath": subpath}

@app.post("/{token}/api/projects/{project_id}/mkdir")
async def create_project_folder(token: str, project_id: str, payload: dict):
    """프로젝트 내 폴더 생성"""
    u, _ = _resolve_user(token)
    ws = get_user_workspace(u)
    folder_name = payload.get("name", "").strip()
    subpath = payload.get("subpath", ".")
    if not folder_name:
        raise HTTPException(400, "폴더명 필요")
    proj_dir = Path(ws) / "_projects" / project_id
    target = proj_dir / subpath / folder_name if subpath and subpath != "." else proj_dir / folder_name
    os.makedirs(str(target), exist_ok=True)
    return {"success": True}

@app.post("/{token}/api/projects/{project_id}/copy-from-workspace")
async def copy_files_to_project(token: str, project_id: str, payload: dict):
    """파일 탐색기에서 프로젝트로 파일/폴더 복사"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(500)
    from bson import ObjectId
    items = payload.get("items", [])  # 워크스페이스 상대 경로 배열
    subpath = payload.get("subpath", ".")
    owner = payload.get("owner")  # 공유 파일의 경우 소유자
    if not items:
        raise HTTPException(400, "복사할 파일을 지정해주세요")
    # 소스 워크스페이스 결정
    if owner:
        src_ws = get_user_workspace(owner)
    else:
        src_ws = get_user_workspace(u)
    # 대상 프로젝트 디렉토리
    proj_dir = Path(get_user_workspace(u)) / "_projects" / project_id
    target_dir = proj_dir / subpath if subpath and subpath != "." else proj_dir
    os.makedirs(str(target_dir), exist_ok=True)
    text_exts = {'.md','.txt','.py','.js','.ts','.html','.css','.json','.xml','.csv','.yaml','.yml','.sh','.sql','.java','.jsx','.tsx','.ini','.cfg','.conf','.env'}
    copied = []
    files_to_push = []
    content_to_push = []
    for item_path in items:
        try:
            src = safe_path(item_path, src_ws)
            if not src.exists():
                continue
            dest_name = src.name
            # 중복 처리
            stem = Path(dest_name).stem
            ext_s = Path(dest_name).suffix
            counter = 1
            dest = target_dir / dest_name
            while dest.exists():
                dest_name = f"{stem}({counter}){ext_s}"
                dest = target_dir / dest_name
                counter += 1
            if src.is_dir():
                await asyncio.to_thread(shutil.copytree, str(src), str(dest))
                copied.append({"name": dest_name, "type": "directory"})
                # 폴더 내 텍스트 파일들 DB에 등록
                for root, dirs, fls in os.walk(str(dest)):
                    for fn in fls:
                        fp = Path(root) / fn
                        rel = str(fp.relative_to(proj_dir)).replace('\\', '/')
                        # 원본 경로 계산: item_path/상대경로
                        orig_rel = str(fp.relative_to(dest)).replace('\\', '/')
                        orig_path = f"{item_path}/{orig_rel}" if orig_rel != fn else f"{item_path}/{fn}"
                        fi = {"name": fn, "rel_path": rel, "path": f"_projects/{project_id}/{rel}", "size": fp.stat().st_size, "uploaded_at": datetime.now(timezone.utc).isoformat()}
                        files_to_push.append(fi)
                        if Path(fn).suffix.lower() in text_exts:
                            try:
                                tc = fp.read_text('utf-8')[:50000]
                                content_to_push.append({"name": rel, "content": tc})
                            except: pass
            else:
                await asyncio.to_thread(shutil.copy2, str(src), str(dest))
                rel = str(dest.relative_to(proj_dir)).replace('\\', '/')
                fi = {"name": dest_name, "rel_path": rel, "path": f"_projects/{project_id}/{rel}", "size": dest.stat().st_size, "uploaded_at": datetime.now(timezone.utc).isoformat()}
                files_to_push.append(fi)
                copied.append({"name": dest_name, "type": "file"})
                if Path(dest_name).suffix.lower() in text_exts:
                    try:
                        tc = dest.read_text('utf-8')[:50000]
                        content_to_push.append({"name": rel, "content": tc})
                    except: pass
        except Exception:
            continue
    # DB 업데이트
    push_ops = {}
    if files_to_push:
        push_ops["files"] = {"$each": files_to_push}
    if content_to_push:
        push_ops["files_content"] = {"$each": content_to_push}
    if push_ops:
        await projects_collection.update_one(
            {"_id": ObjectId(project_id), "username": u},
            {"$push": push_ops, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    return {"success": True, "copied": copied, "count": len(copied)}

@app.delete("/{token}/api/projects/{project_id}/files/{filepath:path}")
async def delete_project_file(token: str, project_id: str, filepath: str):
    """프로젝트 파일 또는 폴더 삭제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(500)
    from bson import ObjectId
    ws = get_user_workspace(u)
    target = Path(ws) / "_projects" / project_id / filepath
    if target.is_dir():
        import shutil
        shutil.rmtree(str(target), ignore_errors=True)
        # DB에서 해당 경로 하위 파일 모두 제거
        await projects_collection.update_one(
            {"_id": ObjectId(project_id), "username": u},
            {
                "$pull": {
                    "files": {"rel_path": {"$regex": f"^{re.escape(filepath)}"}},
                    "files_content": {"name": {"$regex": f"^{re.escape(filepath)}"}}
                },
                "$set": {"updated_at": datetime.now(timezone.utc)}
            }
        )
    elif target.exists():
        os.remove(str(target))
        await projects_collection.update_one(
            {"_id": ObjectId(project_id), "username": u},
            {
                "$pull": {"files": {"rel_path": filepath}, "files_content": {"name": filepath}},
                "$set": {"updated_at": datetime.now(timezone.utc)}
            }
        )
    return {"success": True}

@app.get("/{token}/api/projects/{project_id}/chats")
async def get_project_chats(token: str, project_id: str, skip: int = 0, limit: int = 20):
    """프로젝트에 속한 대화 로그 목록"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or chat_collection is None:
        return {"logs": [], "total": 0}
    total = await chat_collection.count_documents({"username": u, "project_id": project_id})
    cursor = chat_collection.find(
        {"username": u, "project_id": project_id},
        {"messages": 0, "api_history": 0}
    ).sort("updated_at", -1).skip(skip).limit(limit)
    logs = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        logs.append(doc)
    return {"logs": logs, "total": total}

@app.delete("/{token}/api/projects/{project_id}/chats/{session_id}")
async def delete_project_chat(token: str, project_id: str, session_id: str):
    """프로젝트 대화 개별 삭제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or chat_collection is None:
        raise HTTPException(500)
    r = await chat_collection.delete_one({"session_id": session_id, "username": u, "project_id": project_id})
    return {"success": r.deleted_count > 0}

@app.delete("/{token}/api/projects/{project_id}/chats")
async def delete_all_project_chats(token: str, project_id: str):
    """프로젝트 대화 전체 삭제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or chat_collection is None:
        raise HTTPException(500)
    r = await chat_collection.delete_many({"username": u, "project_id": project_id})
    return {"success": True, "deleted": r.deleted_count}

@app.post("/{token}/api/projects/{project_id}/chats/delete-selected")
async def delete_selected_project_chats(token: str, project_id: str, payload: dict):
    """프로젝트 대화 선택 삭제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or chat_collection is None:
        raise HTTPException(500)
    session_ids = payload.get("session_ids", [])
    if not session_ids:
        raise HTTPException(400, "삭제할 대화를 선택해주세요")
    r = await chat_collection.delete_many({"session_id": {"$in": session_ids}, "username": u, "project_id": project_id})
    return {"success": True, "deleted": r.deleted_count}

@app.get("/{token}/api/projects/{project_id}/snapshots")
async def get_project_snapshots(token: str, project_id: str):
    """프로젝트 원본 버전 스냅샷 목록"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        return {"snapshots": []}
    from bson import ObjectId
    doc = await projects_collection.find_one(
        {"_id": ObjectId(project_id), "username": u},
        {"snapshots": 1}
    )
    snapshots = doc.get("snapshots", []) if doc else []
    snapshots.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return {"snapshots": snapshots}

@app.get("/{token}/api/projects/{project_id}/snapshots/{folder_key}/files")
async def list_snapshot_files(token: str, project_id: str, folder_key: str, subpath: str = "."):
    """스냅샷 폴더의 파일/폴더 목록 (구조 탐색)"""
    u, _ = _resolve_user(token)
    ws = get_user_workspace(u)
    snap_dir = Path(ws) / "_projects" / project_id / "_snapshots" / folder_key
    if not snap_dir.exists():
        raise HTTPException(404, "스냅샷 없음")
    target = snap_dir / subpath if subpath and subpath != "." else snap_dir
    if not target.exists() or not target.is_dir():
        raise HTTPException(404, "경로 없음")
    items = []
    for p in sorted(target.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
        if p.name.startswith(".") or p.name.startswith("_"):
            continue
        if p.is_dir():
            cnt = sum(1 for _ in p.rglob("*") if _.is_file())
            items.append({"name": p.name, "type": "directory", "children": cnt})
        else:
            rel = str(p.relative_to(snap_dir)).replace("\\", "/")
            items.append({"name": p.name, "type": "file", "size": p.stat().st_size, "rel_path": rel, "path": f"_projects/{project_id}/_snapshots/{folder_key}/{rel}"})
    return {"items": items, "subpath": subpath}

@app.get("/{token}/api/projects/{project_id}/snapshots/{folder_key}/download")
async def download_snapshot(token: str, project_id: str, folder_key: str):
    """스냅샷 폴더를 zip으로 다운로드"""
    u, _ = _resolve_user(token)
    ws = get_user_workspace(u)
    snap_dir = Path(ws) / "_projects" / project_id / "_snapshots" / folder_key
    if not snap_dir.exists() or not snap_dir.is_dir():
        raise HTTPException(404, "스냅샷 폴더 없음")
    import zipfile, io
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(str(snap_dir)):
            for fn in files:
                fp = Path(root) / fn
                arcname = str(fp.relative_to(snap_dir))
                zf.write(str(fp), arcname)
    zip_buf.seek(0)
    from starlette.responses import StreamingResponse
    return StreamingResponse(zip_buf, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{folder_key}.zip"'})

@app.post("/{token}/api/projects/{project_id}/snapshots/{folder_key}/restore")
async def restore_snapshot(token: str, project_id: str, folder_key: str):
    """스냅샷을 프로젝트 원본으로 복원 (현재 파일을 스냅샷 내용으로 덮어씀)"""
    u, _ = _resolve_user(token)
    ws = get_user_workspace(u)
    proj_dir = Path(ws) / "_projects" / project_id
    snap_dir = proj_dir / "_snapshots" / folder_key
    if not snap_dir.exists():
        raise HTTPException(404, "스냅샷 없음")
    # 현재 프로젝트 파일 삭제 (_snapshots 제외)
    for item in proj_dir.iterdir():
        if item.name in ("_snapshots", "_outputs"):
            continue
        if item.is_dir():
            shutil.rmtree(str(item), ignore_errors=True)
        else:
            item.unlink(missing_ok=True)
    # 스냅샷 복원
    for item in snap_dir.iterdir():
        dest = proj_dir / item.name
        if item.is_dir():
            if dest.exists():
                shutil.rmtree(str(dest), ignore_errors=True)
            shutil.copytree(str(item), str(dest))
        else:
            shutil.copy2(str(item), str(dest))
    return {"success": True}

@app.delete("/{token}/api/projects/{project_id}/snapshots/{folder_key}")
async def delete_snapshot(token: str, project_id: str, folder_key: str):
    """스냅샷 삭제"""
    u, _ = _resolve_user(token)
    if not MONGO_OK or projects_collection is None:
        raise HTTPException(500)
    from bson import ObjectId
    ws = get_user_workspace(u)
    snap_dir = Path(ws) / "_projects" / project_id / "_snapshots" / folder_key
    if snap_dir.exists():
        shutil.rmtree(str(snap_dir), ignore_errors=True)
    await projects_collection.update_one(
        {"_id": ObjectId(project_id), "username": u},
        {"$pull": {"snapshots": {"folder_key": folder_key}}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    return {"success": True}

# ============ 캘린더 API ============
calendar_collection = mongo_db["calendar_events"] if MONGO_OK else None

@app.get("/{token}/api/calendar/events")
async def get_calendar_events(token: str, year: int = 2026, month: int = 1):
    u, _ = _resolve_user(token)
    if not MONGO_OK or calendar_collection is None:
        return {"events": []}
    cursor = calendar_collection.find({"username": u, "year": year, "month": month}, {"_id": 0})
    events = []
    async for doc in cursor:
        events.append(doc)
    return {"events": events}

@app.post("/{token}/api/calendar/events")
async def create_calendar_event(token: str, payload: dict):
    u, _ = _resolve_user(token)
    if not MONGO_OK or calendar_collection is None:
        return {"success": False}
    date_str = payload.get("date", "")
    parts = date_str.split("-")
    payload["username"] = u
    payload["year"] = int(parts[0]) if len(parts) >= 1 else 2026
    payload["month"] = int(parts[1]) if len(parts) >= 2 else 1
    await calendar_collection.replace_one(
        {"username": u, "id": payload.get("id", "")},
        payload, upsert=True
    )
    return {"success": True}

@app.delete("/{token}/api/calendar/events")
async def delete_calendar_event(token: str, id: str = ""):
    u, _ = _resolve_user(token)
    if not MONGO_OK or calendar_collection is None:
        return {"success": False}
    result = await calendar_collection.delete_one({"username": u, "id": id})
    return {"success": result.deleted_count > 0}

# 작업 기록 조회
@app.get("/{token}/api/task-history")
async def u_task_history(token: str):
    u, _ = _resolve_user(token)
    if not MONGO_OK: return {"tasks":[]}
    cursor = task_collection.find({"username":u}).sort("started_at",-1).limit(50); tasks = []
    async for d in cursor:
        tasks.append({
            "task_id":d.get("task_id",""),"message":d.get("message","")[:80],
            "status":d.get("status",""),"started_at":str(d.get("started_at","")),"completed_at":str(d.get("completed_at","") or ""),
            "duration_seconds":d.get("duration_seconds",0)
        })
    return {"tasks":tasks}

# ============================================================
# REST API 백그라운드 작업 (WebSocket 불필요)
# ============================================================
rest_task_buffers: dict = {}  # {task_id: [log_entry, ...]} 메모리 버퍼

async def save_task_log(task_id: str, log_type: str, content: str, meta: dict = None):
    """작업 로그를 MongoDB와 메모리 버퍼에 저장"""
    entry = {
        "task_id": task_id,
        "type": log_type,
        "content": content,
        "meta": meta or {},
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    if task_id not in rest_task_buffers:
        rest_task_buffers[task_id] = []
    rest_task_buffers[task_id].append(entry)
    if MONGO_OK and task_log_collection is not None:
        try:
            await task_log_collection.insert_one(entry.copy())
        except:
            pass

async def run_rest_agent_background(task_id: str, user_message: str, ws_dir: str, current_folder: str, username: str):
    """REST API용 백그라운드 에이전트 - WebSocket 없이 독립 실행, 로그를 DB에 저장"""
    selected_model = select_model(user_message)
    api_key = await get_next_api_key_async()
    client = anthropic.AsyncAnthropic(api_key=api_key)
    fc = ""
    if current_folder and current_folder != ".":
        fc = f"\n\n현재 사용자가 선택한 작업 폴더: {current_folder}\n이 폴더 내의 파일만 우선적으로 참조하여 작업하세요."
    skills_prompt, skill_names = await _get_active_skills_prompt(username, user_message)
    system_prompt = f"""당신은 {APP_ASSISTANT_NAME}입니다.
작업 공간({ws_dir})에서 파일 관리, 코드 작성, 명령어 실행을 도와줍니다.
규칙: 파일 경로는 상대 경로 사용, 수정 전 내용 확인, 한국어 응답, 임시 스크립트는 write_temp_file 사용
Figma 변환 요청 시: figma_get_file로 디자인 구조를 먼저 가져온 후, figma_get_styles로 스타일 정보를 확인하고, 필요한 이미지는 figma_get_images로 추출하세요. URL에서 file_key와 node-id를 파싱하세요.

[Tailwind CSS 필수 규칙]
HTML 파일을 생성할 때 반드시 Tailwind CSS를 사용하세요:
1. <head> 안에 반드시 <script src="https://cdn.tailwindcss.com"></script>를 포함하세요.
2. 인라인 <style> 대신 Tailwind 유틸리티 클래스를 우선 사용하세요.
3. Tailwind으로 불가능한 커스텀 스타일만 <style> 태그로 보완하세요.
4. 반응형: sm:, md:, lg:, xl: 브레이크포인트 적극 활용.
이 규칙은 모든 HTML 생성에 적용됩니다.{fc}
현재: {datetime.now().isoformat()} | OS: {"Windows" if IS_WINDOWS else "Linux/Mac"}{skills_prompt}"""

    history = [{"role": "user", "content": user_message}]
    full_response = ""
    started_at = datetime.now(timezone.utc)
    _key_idx = ANTHROPIC_API_KEYS.index(api_key) + 1 if api_key in ANTHROPIC_API_KEYS else 0
    _model_label = "Opus" if selected_model == MODEL_OPUS else "Sonnet"

    if MONGO_OK and task_collection is not None:
        await task_collection.insert_one({
            "task_id": task_id, "username": username, "message": user_message,
            "status": "running", "started_at": started_at, "completed_at": None,
            "response_summary": "", "current_folder": current_folder, "source": "rest_api",
            "model": selected_model, "key_index": _key_idx
        })

    await save_task_log(task_id, "info", f"작업 시작: {user_message[:200]}", {"username": username, "folder": current_folder, "model": _model_label, "key": _key_idx})

    try:
        for step in range(1, 11):
            await save_task_log(task_id, "progress", f"분석 중... (단계 {step})", {"step": step})

            final_message = None
            for retry in range(5):
                try:
                    step_text = ""
                    async with client.messages.stream(model=selected_model, max_tokens=get_max_tokens(selected_model), system=system_prompt, tools=TOOLS, messages=history) as stream:
                        async for event in stream:
                            if event.type == "content_block_delta" and event.delta.type == "text_delta":
                                chunk = event.delta.text
                                full_response += chunk
                                step_text += chunk
                            elif event.type == "content_block_start":
                                if event.content_block.type == "tool_use":
                                    await save_task_log(task_id, "tool_start", f"도구 호출: {event.content_block.name}", {"tool": event.content_block.name, "id": event.content_block.id})
                        final_message = await stream.get_final_message()
                        stop_reason = final_message.stop_reason
                        if final_message and hasattr(final_message, 'usage'):
                            await record_token_usage(username=username, task_id=task_id, session_id=task_id, service_type="rest_task", model=selected_model, usage=final_message.usage, step=step, key_index=_key_idx)
                    if step_text:
                        await save_task_log(task_id, "text", step_text[:2000], {"step": step})
                    break
                except anthropic.RateLimitError:
                    new_key = await get_next_api_key_async()
                    if new_key and new_key != api_key:
                        client = anthropic.AsyncAnthropic(api_key=new_key)
                        api_key = new_key
                        _key_idx = ANTHROPIC_API_KEYS.index(api_key) + 1 if api_key in ANTHROPIC_API_KEYS else 0
                    wait = min(2 ** retry * 10, 120)
                    await save_task_log(task_id, "warning", f"API 사용량 초과, Key #{_key_idx}로 전환 후 {wait}초 대기 중... ({retry+1}/5)")
                    for remaining in range(wait, 0, -1):
                        await asyncio.sleep(1)
                    await save_task_log(task_id, "info", f"재시도 중... ({retry+1}/5)")
                    if retry == 4:
                        await save_task_log(task_id, "error", "API 사용량 제한으로 요청을 처리할 수 없습니다. 잠시 후 다시 시도해주세요."); break
                except anthropic.APIError as e:
                    await save_task_log(task_id, "error", f"API 오류: {e}"); break

            if final_message is None: break

            history.append({"role": "assistant", "content": final_message.content})
            tool_uses = [b for b in final_message.content if b.type == "tool_use"]

            if not tool_uses: break

            tool_results = []
            for tu in tool_uses:
                await save_task_log(task_id, "tool_executing", f"도구 실행: {tu.name}", {"tool": tu.name, "id": tu.id, "input": tu.input})
                rs = await execute_tool(tu.name, tu.input, ws_dir, username)
                rj = json.loads(rs)
                # tool_result 크기 제한 (8000자)
                if len(rs) > 8000:
                    rs = rs[:8000] + '...(truncated)'
                tool_results.append({"type": "tool_result", "tool_use_id": tu.id, "content": rs})
                success = "error" not in rj
                await save_task_log(task_id, "tool_result", f"도구 완료: {tu.name} ({'성공' if success else '실패'})",
                                    {"tool": tu.name, "id": tu.id, "success": success, "result_summary": json.dumps(rj, ensure_ascii=False)[:500]})
            history.append({"role": "user", "content": tool_results})
            if stop_reason == "end_turn": break

        await tm.set_task_status(task_id, "done")
        await save_task_log(task_id, "done", f"작업 완료 (총 {step}단계)", {"steps": step})

    except Exception as e:
        await tm.set_task_status(task_id, "error")
        await save_task_log(task_id, "error", f"작업 오류: {e}", {"traceback": traceback.format_exc()})

    finally:
        completed_at = datetime.now(timezone.utc)
        if MONGO_OK and task_collection is not None:
            await task_collection.update_one({"task_id": task_id}, {"$set": {
                "status": tm.task_status.get(task_id, "done"),
                "completed_at": completed_at,
                "duration_seconds": (completed_at - started_at).total_seconds(),
                "response_summary": full_response[:500]
            }})
        # 대화 로그(chat_logs)에도 저장 → 좌측 '요청한 업무' 탭에서 확인 가능
        if MONGO_OK and chat_collection is not None:
            session_id = task_id
            await save_history_to_db(session_id, username, history, user_message, full_response, started_at, completed_at, current_folder=current_folder)
        await save_task_log(task_id, "complete", f"총 소요시간: {(completed_at - started_at).total_seconds():.1f}초")
        tm.cleanup_task(task_id)


# REST API: 백그라운드 작업 시작
@app.post("/{token}/api/task")
async def create_rest_task(token: str, payload: dict):
    """JWT 인증 후 백그라운드 작업 시작. WebSocket 불필요."""
    username, ws_dir = _resolve_user(token)
    message = payload.get("message", "").strip()
    current_folder = payload.get("currentFolder", ".")
    if not message:
        raise HTTPException(400, "message 필드가 필요합니다")
    if not ANTHROPIC_API_KEYS:
        raise HTTPException(500, "API KEY 미설정")

    task_id = str(uuid.uuid4())
    rest_task_buffers[task_id] = []
    await tm.set_task_status(task_id, "running")
    tm.task_username[task_id] = username

    async_task = asyncio.create_task(
        run_rest_agent_background(task_id, message, ws_dir, current_folder, username)
    )
    tm.running_tasks[task_id] = async_task

    return {
        "task_id": task_id,
        "status": "running",
        "username": username,
        "message": message[:200],
        "created_at": datetime.now(timezone.utc).isoformat()
    }

# REST API: 작업 상태 조회
@app.get("/{token}/api/task/{task_id}")
async def get_rest_task(token: str, task_id: str):
    """작업 상태 및 요약 정보 조회"""
    username, _ = _resolve_user(token)

    # 메모리에서 상태 확인
    status = tm.task_status.get(task_id, None)

    # MongoDB에서 상세 정보 조회
    task_doc = None
    if MONGO_OK and task_collection is not None:
        task_doc = await task_collection.find_one({"task_id": task_id, "username": username})

    if not task_doc and status is None:
        raise HTTPException(404, "작업을 찾을 수 없습니다")

    log_count = len(rest_task_buffers.get(task_id, []))

    return {
        "task_id": task_id,
        "status": status or (task_doc.get("status") if task_doc else "unknown"),
        "message": task_doc.get("message", "") if task_doc else "",
        "started_at": str(task_doc.get("started_at", "")) if task_doc else "",
        "completed_at": str(task_doc.get("completed_at", "") or "") if task_doc else "",
        "duration_seconds": task_doc.get("duration_seconds", 0) if task_doc else 0,
        "response_summary": task_doc.get("response_summary", "") if task_doc else "",
        "log_count": log_count
    }

# REST API: 작업 로그 조회
@app.get("/{token}/api/task/{task_id}/logs")
async def get_rest_task_logs(token: str, task_id: str, since: int = 0):
    """
    작업 로그 조회.
    since 파라미터: 해당 인덱스 이후의 로그만 반환 (폴링용)
    """
    username, _ = _resolve_user(token)

    # 메모리 버퍼에서 조회 (빠름)
    buffer = rest_task_buffers.get(task_id, [])
    if buffer:
        logs = buffer[since:] if since < len(buffer) else []
        return {
            "task_id": task_id,
            "status": tm.task_status.get(task_id, "unknown"),
            "total_logs": len(buffer),
            "since": since,
            "logs": logs
        }

    # 메모리에 없으면 MongoDB에서 조회
    if MONGO_OK and task_log_collection is not None:
        # 권한 확인
        task_doc = await task_collection.find_one({"task_id": task_id, "username": username})
        if not task_doc:
            raise HTTPException(404, "작업을 찾을 수 없습니다")

        cursor = task_log_collection.find({"task_id": task_id}).sort("timestamp", 1)
        logs = []
        async for doc in cursor:
            doc.pop("_id", None)
            # datetime 객체를 문자열로 변환
            for k, v in doc.items():
                if isinstance(v, datetime):
                    doc[k] = v.isoformat()
            logs.append(doc)
        return {
            "task_id": task_id,
            "status": task_doc.get("status", "unknown"),
            "total_logs": len(logs),
            "since": since,
            "logs": logs[since:] if since < len(logs) else []
        }

    raise HTTPException(404, "로그를 찾을 수 없습니다")

# REST API: 작업 취소
@app.post("/{token}/api/task/{task_id}/cancel")
async def cancel_rest_task(token: str, task_id: str):
    """실행 중인 작업 취소"""
    username, _ = _resolve_user(token)
    if tm.task_username.get(task_id) != username:
        raise HTTPException(403, "권한이 없습니다")
    task = tm.running_tasks.get(task_id)
    if task and not task.done():
        task.cancel()
        await tm.set_task_status(task_id, "cancelled")
        await save_task_log(task_id, "cancelled", "사용자에 의해 작업이 취소되었습니다")
        if MONGO_OK and task_collection is not None:
            await task_collection.update_one({"task_id": task_id}, {"$set": {"status": "cancelled", "completed_at": datetime.now(timezone.utc)}})
        return {"task_id": task_id, "status": "cancelled"}
    return {"task_id": task_id, "status": tm.task_status.get(task_id, "unknown"), "message": "작업이 이미 완료되었거나 존재하지 않습니다"}

# REST API: 사용자의 모든 REST 작업 목록
@app.get("/{token}/api/tasks")
async def list_rest_tasks(token: str, skip: int = 0, limit: int = 13):
    """사용자의 REST API 작업 목록 (최근순, 페이징)"""
    username, _ = _resolve_user(token)
    if not MONGO_OK:
        return {"tasks": [], "total": 0, "skip": skip, "limit": limit}
    total = await task_collection.count_documents({"username": username, "source": "rest_api"})
    cursor = task_collection.find({"username": username, "source": "rest_api"}).sort("started_at", -1).skip(skip).limit(limit)
    tasks = []
    async for d in cursor:
        tasks.append({
            "task_id": d.get("task_id", ""),
            "message": d.get("message", "")[:120],
            "status": d.get("status", ""),
            "started_at": str(d.get("started_at", "")),
            "completed_at": str(d.get("completed_at", "") or ""),
            "duration_seconds": d.get("duration_seconds", 0),
            "response_summary": d.get("response_summary", "")[:200]
        })
    return {"tasks": tasks, "total": total, "skip": skip, "limit": limit}


# ============================================================
# WebSocket (백그라운드 작업 + 재접속 복원)
# ============================================================
user_histories: dict = {}  # {username: {"history":[], "session_id":""}}

def serialize_history(history: list) -> list:
    """Anthropic content block 객체를 JSON 직렬화 가능한 dict로 변환"""
    result = []
    for msg in history:
        role = msg.get("role", "") if isinstance(msg, dict) else getattr(msg, "role", "")
        content = msg.get("content", "") if isinstance(msg, dict) else getattr(msg, "content", "")
        if isinstance(content, str):
            result.append({"role": role, "content": content})
        elif isinstance(content, list):
            serialized_content = []
            for block in content:
                if isinstance(block, dict):
                    serialized_content.append(block)
                elif hasattr(block, 'type'):
                    if block.type == "text":
                        serialized_content.append({"type": "text", "text": block.text})
                    elif block.type == "tool_use":
                        serialized_content.append({"type": "tool_use", "id": block.id, "name": block.name, "input": block.input})
                    elif block.type == "tool_result":
                        serialized_content.append({"type": "tool_result", "tool_use_id": block.tool_use_id, "content": block.content if isinstance(block.content, str) else str(block.content)})
                    else:
                        serialized_content.append({"type": block.type, "text": getattr(block, "text", "")})
                else:
                    serialized_content.append(block)
            result.append({"role": role, "content": serialized_content})
        else:
            result.append({"role": role, "content": str(content)})
    return result

def truncate_history_for_db(history: list, max_pairs: int = 20) -> list:
    """DB 저장용: 최근 N쌍만 유지하여 도큐먼트 크기 제한 (tool 중간 과정 제거하고 핵심만)"""
    serialized = serialize_history(history)
    # 대화쌍 기준으로 최근 것만 유지
    if len(serialized) > max_pairs * 2:
        serialized = serialized[-(max_pairs * 2):]
    # 각 content 블록의 이미지/tool_result 내용을 축약
    for msg in serialized:
        if isinstance(msg.get("content"), list):
            for i, block in enumerate(msg["content"]):
                if isinstance(block, dict):
                    # 이미지 블록: base64 데이터를 제거하고 플레이스홀더로 교체
                    if block.get("type") == "image":
                        media_type = block.get("source", {}).get("media_type", "image/png")
                        msg["content"][i] = {"type": "text", "text": f"[이미지 첨부됨: {media_type}]"}
                    elif block.get("type") == "tool_result":
                        c = block.get("content", "")
                        if isinstance(c, str) and len(c) > 500:
                            block["content"] = c[:500] + "...(truncated)"
    return serialized

async def save_history_to_db(session_id: str, username: str, history: list, user_message: str = "", full_response: str = "", started_at=None, completed_at=None, current_folder: str = ".", project_id: str = ""):
    """대화 히스토리를 MongoDB에 저장 (api_history 포함)"""
    if not MONGO_OK or chat_collection is None:
        return
    title = user_message[:80] + ("..." if len(user_message) > 80 else "") if user_message else ""
    me = {"role": "user", "content": user_message, "timestamp": (started_at or datetime.now(timezone.utc)).isoformat()}
    ae = {"role": "assistant", "content": full_response[:2000], "timestamp": (completed_at or datetime.now(timezone.utc)).isoformat()}
    api_history = truncate_history_for_db(history)
    existing = await chat_collection.find_one({"session_id": session_id})
    if existing:
        update_set = {"updated_at": completed_at or datetime.now(timezone.utc), "api_history": api_history}
        if current_folder and current_folder != ".":
            update_set["current_folder"] = current_folder
        await chat_collection.update_one(
            {"session_id": session_id},
            {"$push": {"messages": {"$each": [me, ae]}},
             "$set": update_set}
        )
    else:
        doc = {
            "session_id": session_id, "username": username, "title": title,
            "messages": [me, ae], "api_history": api_history,
            "current_folder": current_folder if current_folder and current_folder != "." else ".",
            "created_at": started_at or datetime.now(timezone.utc),
            "updated_at": completed_at or datetime.now(timezone.utc)
        }
        if project_id:
            doc["project_id"] = project_id
        await chat_collection.insert_one(doc)

async def load_history_from_db(session_id: str) -> list:
    """MongoDB에서 api_history를 복원"""
    if not MONGO_OK or chat_collection is None:
        return []
    doc = await chat_collection.find_one({"session_id": session_id}, {"api_history": 1})
    if doc and doc.get("api_history"):
        return sanitize_history(doc["api_history"])
    return []

def sanitize_history(history: list) -> list:
    """API에 전달하기 전에 tool_use/tool_result 쌍 무결성 검증.
    tool_use가 없는 tool_result, tool_result가 없는 tool_use를 제거하여
    Anthropic API의 유효성 검증을 통과하도록 보장."""
    if not history:
        return history
    # 1단계: 모든 tool_use id 수집 (assistant 메시지에서)
    tool_use_ids = set()
    for msg in history:
        if msg.get("role") == "assistant" and isinstance(msg.get("content"), list):
            for block in msg["content"]:
                if isinstance(block, dict) and block.get("type") == "tool_use" and block.get("id"):
                    tool_use_ids.add(block["id"])
    # 2단계: 모든 tool_result의 tool_use_id 수집 (user 메시지에서)
    tool_result_ids = set()
    for msg in history:
        if msg.get("role") == "user" and isinstance(msg.get("content"), list):
            for block in msg["content"]:
                if isinstance(block, dict) and block.get("type") == "tool_result" and block.get("tool_use_id"):
                    tool_result_ids.add(block["tool_use_id"])
    # 3단계: 고아 블록 제거
    cleaned = []
    for msg in history:
        role = msg.get("role", "")
        content = msg.get("content")
        if isinstance(content, list):
            new_content = []
            for block in content:
                if isinstance(block, dict):
                    # tool_result인데 대응 tool_use가 없으면 제거
                    if block.get("type") == "tool_result":
                        if block.get("tool_use_id") in tool_use_ids:
                            new_content.append(block)
                        # else: 고아 tool_result → 스킵
                    # tool_use인데 대응 tool_result가 없으면 텍스트로 변환
                    elif block.get("type") == "tool_use":
                        if block.get("id") in tool_result_ids:
                            new_content.append(block)
                        else:
                            new_content.append({"type": "text", "text": f"[도구 호출: {block.get('name', '?')}]"})
                    else:
                        new_content.append(block)
                else:
                    new_content.append(block)
            if new_content:
                cleaned.append({"role": role, "content": new_content})
            # content가 비면 빈 텍스트로 대체 (메시지 순서 유지)
            else:
                cleaned.append({"role": role, "content": "[이전 대화]"})
        else:
            cleaned.append(msg)
    # 4단계: 연속 role 방지 (user-user, assistant-assistant)
    final = []
    for msg in cleaned:
        if final and final[-1].get("role") == msg.get("role"):
            # 같은 role 연속이면 병합
            prev = final[-1]
            if isinstance(prev.get("content"), str) and isinstance(msg.get("content"), str):
                prev["content"] = prev["content"] + "\n" + msg["content"]
            else:
                continue  # 복잡한 블록은 뒤 것을 스킵
        else:
            final.append(msg)
    # 5단계: 첫 메시지가 user가 아니면 앞에 더미 추가
    if final and final[0].get("role") != "user":
        final.insert(0, {"role": "user", "content": "(이전 대화 계속)"})
    # 6단계: 마지막이 assistant면 유지 (다음 user 입력 대기)
    return final

@app.websocket("/ws/chat")
async def ws_chat(websocket: WebSocket):
    await _handle_ws(websocket, WORKSPACE_ROOT, "default", "")

@app.websocket("/ws/chat/{token}")
async def ws_chat_user(websocket: WebSocket, token: str):
    if token.count(".") == 2:
        username = userid_from_jwt(token)
        if not username:
            await websocket.accept()
            expired = is_jwt_expired(token)
            _email = email_from_jwt(token)
            if expired:
                print(f"[WS REJECTED] 만료된 토큰: {_email}")
                await websocket.send_json({"type":"auth_expired","content":"세션이 만료되었습니다. K-Portal에서 다시 접속해주세요."})
            else:
                print(f"[WS REJECTED] 유효하지 않은 토큰: {_email}")
                await websocket.send_json({"type":"auth_error","content":"유효하지 않은 토큰입니다."})
            await websocket.close(code=4001, reason="auth_failed"); return
    elif USERNAME_PATTERN.match(token):
        username = token
    else:
        await websocket.accept()
        await websocket.send_json({"type":"error","content":"유효하지 않은 접근"})
        await websocket.close(); return
    await _handle_ws(websocket, get_user_workspace(username), username, token)

async def _handle_ws(websocket, ws_dir, username, token):
    await websocket.accept()
    tm.register_ws(username, websocket)
    # 사용자 접속 로그 (이메일 포함)
    _user_email = email_from_jwt(token) if token and token.count(".") == 2 else username
    print(f"[WS CONNECT] {username} ({_user_email}) - ws_dir: {ws_dir}")

    # 분산 세션에 이메일 저장
    if MONGO_OK and active_sessions_col is not None:
        try:
            await active_sessions_col.update_one(
                {"username": username},
                {"$set": {"last_active": datetime.now(timezone.utc), "email": _user_email, "server_id": SERVER_ID}},
                upsert=True
            )
        except: pass

    # 세션 초기화/복원
    if username not in user_histories:
        # 서버 재시작 등으로 메모리에 없는 경우 DB에서 마지막 세션 복원 시도
        restored = False
        if MONGO_OK and chat_collection is not None:
            try:
                last_doc = await chat_collection.find_one(
                    {"username": username},
                    {"session_id": 1, "api_history": 1, "messages": 1},
                    sort=[("updated_at", -1)]
                )
                if last_doc and last_doc.get("api_history"):
                    api_history = sanitize_history(last_doc["api_history"])
                    api_user_count = len([m for m in api_history if m.get("role") == "user"])
                    session_messages = last_doc.get("messages", [])
                    session_user_count = len([m for m in session_messages if m.get("role") == "user"])
                    # api_history가 세션 메시지 수보다 많으면 다른 세션 히스토리 섞임 → 재구성
                    if api_user_count > session_user_count and session_messages:
                        rebuilt = []
                        for msg in session_messages:
                            role = msg.get("role", "")
                            content = msg.get("content", "")
                            if role in ("user", "assistant") and content:
                                rebuilt.append({"role": role, "content": content})
                        api_history = rebuilt
                    user_histories[username] = {
                        "history": api_history,
                        "session_id": last_doc["session_id"]
                    }
                    restored = True
            except:
                pass
        if not restored:
            user_histories[username] = {"history": [], "session_id": str(uuid.uuid4())}
    session_id = user_histories[username]["session_id"]

    # 현재 활성 작업이 있으면 버퍼 전송 (재접속 복원)
    active_task = tm.user_active_task.get(username)
    if active_task and active_task in tm.task_buffers:
        status = tm.task_status.get(active_task, "unknown")
        await websocket.send_json({"type":"reconnect","task_id":active_task,"status":status,"buffered_count":len(tm.task_buffers[active_task])})
        # 버퍼된 메시지 모두 전송
        for msg in tm.task_buffers[active_task]:
            try: await websocket.send_json(msg)
            except: break
        if status in ("done","error"):
            # 완료된 작업이면 정리
            await tm.clear_active_task(username)
    else:
        await websocket.send_json({"type":"session_init","session_id":session_id,"username":username})

    try:
        # ping/pong keepalive: 30초마다 ping 전송
        async def keepalive():
            try:
                while True:
                    await asyncio.sleep(30)
                    try:
                        await websocket.send_json({"type": "ping"})
                    except:
                        break
            except asyncio.CancelledError:
                pass

        ping_task = asyncio.create_task(keepalive())
        try:
            while True:
                data = await websocket.receive_json()

                if data.get("type") == "pong":
                    continue

                if data.get("type") == "clear":
                    new_sid = str(uuid.uuid4())
                    user_histories[username] = {"history":[], "session_id":new_sid}
                    await tm.clear_active_task(username)
                    await websocket.send_json({"type":"cleared","session_id":new_sid})
                    continue

                if data.get("type") == "load_session":
                    old_sid = data.get("session_id","")
                    if MONGO_OK:
                        doc = await chat_collection.find_one({"session_id":old_sid})
                        if doc:
                            # 해당 세션의 messages에서 실제 대화 수 확인
                            session_messages = doc.get("messages", [])
                            session_user_count = len([m for m in session_messages if m.get("role") == "user"])
                            
                            # api_history 복원
                            restored_history = sanitize_history(doc.get("api_history", []))
                            api_user_count = len([m for m in restored_history if m.get("role") == "user"])
                            
                            # api_history의 user 수가 세션 messages의 user 수보다 많으면
                            # 다른 세션의 히스토리가 섞인 것이므로, messages에서 히스토리를 재구성
                            if api_user_count > session_user_count and session_messages:
                                rebuilt = []
                                for msg in session_messages:
                                    role = msg.get("role", "")
                                    content = msg.get("content", "")
                                    if role in ("user", "assistant") and content:
                                        rebuilt.append({"role": role, "content": content})
                                restored_history = rebuilt
                                api_user_count = session_user_count
                            
                            _loaded_proj_id = doc.get("project_id", "")
                            user_histories[username] = {"history": restored_history, "session_id": old_sid, "project_id": _loaded_proj_id}
                            
                            # 프로젝트 정보 조회
                            _proj_info = None
                            if _loaded_proj_id and projects_collection is not None:
                                try:
                                    from bson import ObjectId
                                    _proj_doc = await projects_collection.find_one({"_id": ObjectId(_loaded_proj_id)})
                                    if _proj_doc:
                                        _proj_info = {"id": _loaded_proj_id, "name": _proj_doc.get("name",""), "description": _proj_doc.get("description","")}
                                except: pass
                            
                            resp = {
                                "type":"session_loaded","session_id":old_sid,
                                "messages":session_messages,
                                "context_restored":api_user_count,
                                "current_folder":doc.get("current_folder",".")
                            }
                            if _loaded_proj_id:
                                resp["project_id"] = _loaded_proj_id
                            if _proj_info:
                                resp["project"] = _proj_info
                            await websocket.send_json(resp)
                            continue
                    await websocket.send_json({"type":"error","content":"세션 없음"})
                    continue

                if data.get("type") == "cancel":
                    active = tm.user_active_task.get(username)
                    if active and active in tm.running_tasks and not tm.running_tasks[active].done():
                        tm.running_tasks[active].cancel()
                        await tm.set_task_status(active, "cancelled")
                        await tm.clear_active_task(username)
                        if MONGO_OK and task_collection is not None:
                            await task_collection.update_one({"task_id": active}, {"$set": {"status": "cancelled", "completed_at": datetime.now(timezone.utc)}})
                        await websocket.send_json({"type":"cancelled","task_id":active})
                    else:
                        await websocket.send_json({"type":"error","content":"진행 중인 작업이 없습니다."})
                    continue

                if data.get("type") == "compress_context":
                    # 대화 컨텍스트 압축
                    history = user_histories.get(username, {}).get("history", [])
                    if len(history) < 4:
                        await websocket.send_json({"type":"compress_result","success":False,"message":"압축할 대화가 충분하지 않습니다."})
                        continue
                    try:
                        await websocket.send_json({"type":"compress_progress","progress":10,"message":"대화를 계속하기 위해 압축하고 있습니다..."})

                        # 히스토리에서 텍스트만 추출
                        conversation_text = ""
                        msg_count = 0
                        for m in history:
                            role = m.get("role", "")
                            content = m.get("content", "")
                            if isinstance(content, str) and content.strip():
                                conversation_text += f"\n[{role}]: {content[:2000]}\n"
                                msg_count += 1
                            elif isinstance(content, list):
                                for block in content:
                                    if isinstance(block, dict):
                                        if block.get("type") == "text":
                                            conversation_text += f"\n[{role}]: {block['text'][:2000]}\n"
                                            msg_count += 1
                                        elif block.get("type") == "tool_use":
                                            conversation_text += f"\n[tool_use: {block.get('name','')}]\n"
                                        elif block.get("type") == "tool_result":
                                            ct = block.get("content","")
                                            if isinstance(ct, str):
                                                conversation_text += f"\n[tool_result]: {ct[:500]}\n"

                        await websocket.send_json({"type":"compress_progress","progress":30,"message":f"{msg_count}개 메시지를 분석하고 있습니다..."})

                        # Claude API로 대화 요약
                        api_key = await get_next_api_key_async()
                        compress_client = anthropic.AsyncAnthropic(api_key=api_key)

                        await websocket.send_json({"type":"compress_progress","progress":50,"message":"AI가 대화 내용을 요약하고 있습니다..."})

                        summary_response = await compress_client.messages.create(
                            model=MODEL_SONNET,
                            max_tokens=4096,
                            system="You are a conversation summarizer. Summarize the entire conversation concisely but completely, preserving:\n1. All key decisions, results, and conclusions\n2. Important file paths, folder names, and technical details\n3. Any pending tasks or next steps\n4. User preferences expressed during the conversation\nWrite in the same language as the conversation. Be thorough but concise.",
                            messages=[{"role":"user","content":f"다음 대화를 요약해주세요. 핵심 내용, 결정사항, 작업 결과, 파일 경로 등 중요한 세부 사항을 모두 포함해야 합니다:\n\n{conversation_text[:50000]}"}]
                        )

                        summary = ""
                        for block in summary_response.content:
                            if hasattr(block, "text"):
                                summary += block.text

                        await websocket.send_json({"type":"compress_progress","progress":80,"message":"압축된 컨텍스트를 적용하고 있습니다..."})

                        # 새 히스토리: 요약 + 최근 2턴 유지
                        new_history = [
                            {"role":"user","content":f"[이전 대화 요약]\n{summary}\n\n위 요약은 이전 대화의 압축된 컨텍스트입니다. 이 맥락을 기반으로 대화를 이어가겠습니다."},
                            {"role":"assistant","content":"네, 이전 대화 내용을 이해했습니다. 요약된 맥락을 바탕으로 계속 진행하겠습니다. 무엇을 도와드릴까요?"}
                        ]
                        # 최근 2턴(user+assistant) 유지
                        recent = []
                        for m in reversed(history):
                            if isinstance(m.get("content"), str) and m.get("role") in ("user","assistant"):
                                recent.insert(0, m)
                                if len(recent) >= 4: break
                            elif isinstance(m.get("content"), list):
                                # tool_result 등은 제외
                                has_text = any(b.get("type") == "text" for b in m["content"] if isinstance(b, dict))
                                if has_text and m.get("role") in ("user","assistant"):
                                    text_only = [b for b in m["content"] if isinstance(b, dict) and b.get("type") == "text"]
                                    recent.insert(0, {"role": m["role"], "content": text_only})
                                    if len(recent) >= 4: break
                        new_history.extend(recent)

                        old_count = len(history)
                        user_histories[username]["history"] = new_history

                        # DB에도 업데이트
                        if MONGO_OK and chat_collection is not None:
                            sid = user_histories[username].get("session_id","")
                            if sid:
                                await chat_collection.update_one(
                                    {"session_id": sid},
                                    {"$set": {"api_history": new_history, "compressed": True, "compressed_at": datetime.now(timezone.utc)}}
                                )

                        await websocket.send_json({"type":"compress_progress","progress":100,"message":"압축 완료!"})
                        await websocket.send_json({
                            "type":"compress_result",
                            "success":True,
                            "message":f"대화 컨텍스트가 압축되었습니다. ({old_count}개 → {len(new_history)}개 메시지)",
                            "old_count": old_count,
                            "new_count": len(new_history)
                        })
                    except Exception as e:
                        await websocket.send_json({"type":"compress_result","success":False,"message":f"압축 실패: {str(e)}"})
                    continue

                msg = data.get("message","").strip()
                cf = data.get("currentFolder",".")
                share_owner = data.get("shareOwner", None)
                images = data.get("images", None)  # [{data, media_type, name}]
                forced_skill = data.get("forcedSkill", None)  # 슬래시 명령으로 선택된 스킬
                project_id = data.get("projectId", "") or ""  # 프로젝트 ID
                if not msg: continue
                if not ANTHROPIC_API_KEYS:
                    await websocket.send_json({"type":"error","content":"API KEY 미설정"})
                    await websocket.send_json({"type":"done","steps":0}); continue

                # 프로젝트 ID를 user_histories에 저장
                if project_id:
                    user_histories[username]["project_id"] = project_id

                # 공유 폴더 모드: 소유자의 워크스페이스 사용
                effective_ws = ws_dir
                if share_owner and share_owner != username:
                    if MONGO_OK and shared_folders_collection is not None:
                        share = await shared_folders_collection.find_one({"owner": share_owner, "shared_with": username})
                        if share:
                            effective_ws = get_user_workspace(share_owner)

                # 이미 실행 중인 작업이 있으면 거부 (분산 체크)
                if await tm.is_user_busy(username):
                    await websocket.send_json({"type":"error","content":"이전 작업이 진행 중입니다. 완료를 기다려주세요."})
                    continue

                # 백그라운드 작업 시작
                task_id = str(uuid.uuid4())
                tm.task_buffers[task_id] = []
                await tm.set_task_status(task_id, "running")
                tm.task_username[task_id] = username
                await tm.set_active_task(username, task_id)
                # session_id는 대화 단위로 유지 (첫 메시지일 때만 새로 생성)
                if not user_histories[username].get("session_id"):
                    user_histories[username]["session_id"] = task_id

                async_task = asyncio.create_task(
                    run_agent_background(task_id, msg, user_histories[username]["history"], effective_ws, cf, username, images=images, forced_skill_name=forced_skill, project_id=project_id)
                )
                tm.running_tasks[task_id] = async_task

                # Task 완료 시 history 업데이트를 위한 콜백
                def on_task_done(t, tid=task_id, uname=username):
                    try:
                        result = t.result()
                        if result and uname in user_histories:
                            user_histories[uname]["history"] = result
                    except: pass
                async_task.add_done_callback(on_task_done)

        except WebSocketDisconnect: pass
        except (ConnectionResetError, OSError): pass
        except: pass
        finally:
            ping_task.cancel()
    except WebSocketDisconnect: pass
    except (ConnectionResetError, OSError): pass
    except: pass
    finally:
        tm.unregister_ws(username, websocket)

# ============================================================
# Static & Start
# ============================================================
app.mount("/static", StaticFiles(directory="static"), name="static")

if __name__ == "__main__":
    import uvicorn
    if IS_WINDOWS:
        import logging; logging.getLogger("asyncio").setLevel(logging.CRITICAL)
        # Windows CMD UTF-8 출력 보장
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except:
            pass
    print("=" * 56)
    print(f"  [*] {APP_TITLE} v5")
    print("=" * 56)
    print(f"  [DIR]     Workspace : {WORKSPACE_ROOT}")
    print(f"  [JWT]     Key       : {JWT_SECRET[:4]}****")
    print(f"  [DB]      MongoDB   : {'OK' if MONGO_OK else 'FAIL'} ({MONGO_DB_NAME})")
    print(f"  [AI]      API Keys  : {len(ANTHROPIC_API_KEYS)}")
    print(f"  [VER]     Version   : {_app_version or 'auto'}")
    print(f"  [ENV]     .env      : {'OK' if _env_loaded else 'NOT LOADED'}")
    _ssl_args = {}
    _ssl_cert = os.environ.get("SSL_CERTFILE", "")
    _ssl_key = os.environ.get("SSL_KEYFILE", "")
    if _ssl_cert and _ssl_key:
        if os.path.isfile(_ssl_cert) and os.path.isfile(_ssl_key):
            _ssl_args["ssl_certfile"] = _ssl_cert
            _ssl_args["ssl_keyfile"] = _ssl_key
            _ssl_pw = os.environ.get("SSL_KEYFILE_PASSWORD", "")
            if _ssl_pw:
                _ssl_args["ssl_keyfile_password"] = _ssl_pw
            print(f"  [SSL]     Cert      : OK ({Path(_ssl_cert).name})")
        else:
            print(f"  [SSL]     Cert      : FAIL - file not found!")
            if not os.path.isfile(_ssl_cert):
                print(f"            -> CERT: {_ssl_cert}")
            if not os.path.isfile(_ssl_key):
                print(f"            -> KEY : {_ssl_key}")
    else:
        print(f"  [SSL]     Cert      : DISABLED (HTTP mode)")
    _port = int(os.environ.get("PORT", "5012"))
    _host = os.environ.get("HOST", "0.0.0.0")
    print(f"  [URL]     {'https' if _ssl_args else 'http'}://{_host}:{_port}")
    print("=" * 56)
    uvicorn.run(app, host=_host, port=_port, **_ssl_args)
